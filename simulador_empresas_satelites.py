import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import pandas as pd
from datetime import datetime
from scipy import stats # Para Pruebas de Hip√≥tesis
import json
import tempfile
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



class OptimizadorPro:
    def __init__(self, root):
        self.root = root
        self.root.title("Estrategia Fiscal Avanzada - Multi-Holding")
        self.root.geometry("1500x950")
        
        # Par√°metros Base
        self.UIT = tk.DoubleVar(value=5500)
        self.TASA_GENERAL = tk.DoubleVar(value=29.5)
        self.TASA_ESPECIAL = tk.DoubleVar(value=10.0)
        self.LIMITE_UTILIDAD_UIT = tk.DoubleVar(value=15)
        self.LIMITE_INGRESOS_UIT = tk.DoubleVar(value=1700)
        self.num_satelites = tk.IntVar(value=4)
        # --- NUEVAS VARIABLES EDITABLES ---
        self.CONF_NIVEL = tk.DoubleVar(value=95.0)     # Para Simulaci√≥n
        self.ALFA_SIG = tk.DoubleVar(value=0.05)       # Para Prueba Hip√≥tesis
        self.COMP_LIMITE_PCT = tk.DoubleVar(value=20.0)# % de incremento en comparativo
        self.COMP_TASA_DELTA = tk.DoubleVar(value=2.0) # puntos de reducci√≥n en comparativo
        self.TASA_PAGO_CUENTA = tk.DoubleVar(value=1.5)  # Tasa pago a cuenta IR (%)

        # --- PARAMETROS DE SIMULACION PARA MATRIZ Y CORRELACION TERRITORIAL ---
        self.VAR_MATRIZ_COSTOS = tk.BooleanVar(value=False)
        self.PCT_VAR_MATRIZ_COSTOS = tk.DoubleVar(value=0.0)
        self.VAR_MATRIZ_INGRESOS = tk.BooleanVar(value=False)
        self.PCT_VAR_MATRIZ_INGRESOS = tk.DoubleVar(value=0.0)
        self.DISTRIBUCION_FACTORES = tk.StringVar(value="Log-normal")
        self.CORRELACION_MATRIZ_SAT = tk.DoubleVar(value=0.0)
        # Widgets Entry que necesitan habilitarse/deshabilitarse
        self.entry_pct_var_matriz_costos = None
        self.entry_pct_var_matriz_ingresos = None

        
        self.satelites_entries = []
        # Almacenar figuras para exportaci√≥n Excel
        self.fig_resultados = None
        self.fig_simulacion = None
        self.fig_sensibilidad = None
        self.fig_comparativo = None
        self.datos_simulacion = None  # Estad√≠sticas Monte Carlo
        self.fig_pagos_cuenta = None
        self.datos_pagos_cuenta = None
        self.fig_equilibrio_sim = None
        self.datos_equilibrio_sim = None
        self.crear_interfaz()
        
    def crear_interfaz(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Pesta√±as
        tab_config = ttk.Frame(notebook)
        notebook.add(tab_config, text='Configuraci√≥n Tributaria')
        
        tab_datos = ttk.Frame(notebook)
        notebook.add(tab_datos, text='Datos del Grupo')
        
        tab_resultados = ttk.Frame(notebook)
        notebook.add(tab_resultados, text='Resultados y An√°lisis')
        
        tab_simulacion = ttk.Frame(notebook)
        notebook.add(tab_simulacion, text='Simulaci√≥n Monte Carlo')
        
        tab_sensibilidad = ttk.Frame(notebook)
        notebook.add(tab_sensibilidad, text='An√°lisis de Sensibilidad')
        
        tab_comparativo = ttk.Frame(notebook)
        notebook.add(tab_comparativo, text='An√°lisis Comparativo')

        tab_pagos_cuenta = ttk.Frame(notebook)
        notebook.add(tab_pagos_cuenta, text='Pagos a Cuenta y Equilibrio')

        self.crear_tab_configuracion(tab_config)
        self.crear_tab_datos(tab_datos)
        self.crear_tab_resultados(tab_resultados)
        self.crear_tab_simulacion(tab_simulacion)
        self.crear_tab_sensibilidad(tab_sensibilidad)
        self.crear_tab_comparativo(tab_comparativo)
        self.crear_tab_pagos_cuenta(tab_pagos_cuenta)
        
    def crear_tab_configuracion(self, parent):
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # SECCI√ìN: Par√°metros Tributarios
        frame_parametros = ttk.LabelFrame(scrollable_frame, text="Par√°metros del Sistema Tributario", padding=20)
        frame_parametros.grid(row=0, column=0, padx=10, pady=10, sticky='ew', columnspan=2)
        
        ttk.Label(frame_parametros, text="Unidad de Referencia (UIT):").grid(row=0, column=0, sticky='w', pady=8)
        entry_uit = ttk.Entry(frame_parametros, textvariable=self.UIT, width=20)
        entry_uit.grid(row=0, column=1, pady=8, padx=5)
        
        ttk.Label(frame_parametros, text="Tasa R√©gimen General (%):").grid(row=1, column=0, sticky='w', pady=8)
        entry_tasa_gral = ttk.Entry(frame_parametros, textvariable=self.TASA_GENERAL, width=20)
        entry_tasa_gral.grid(row=1, column=1, pady=8, padx=5)
        
        ttk.Label(frame_parametros, text="Tasa R√©gimen Especial/MYPE (%):").grid(row=2, column=0, sticky='w', pady=8)
        entry_tasa_esp = ttk.Entry(frame_parametros, textvariable=self.TASA_ESPECIAL, width=20)
        entry_tasa_esp.grid(row=2, column=1, pady=8, padx=5)
        
        ttk.Label(frame_parametros, text="L√≠mite Utilidad R√©gimen Especial (UIT):").grid(row=3, column=0, sticky='w', pady=8)
        entry_lim_util = ttk.Entry(frame_parametros, textvariable=self.LIMITE_UTILIDAD_UIT, width=20)
        entry_lim_util.grid(row=3, column=1, pady=8, padx=5)
        
        ttk.Label(frame_parametros, text="L√≠mite Ingresos R√©gimen Especial (UIT):").grid(row=4, column=0, sticky='w', pady=8)
        entry_lim_ing = ttk.Entry(frame_parametros, textvariable=self.LIMITE_INGRESOS_UIT, width=20)
        entry_lim_ing.grid(row=4, column=1, pady=8, padx=5)

        ttk.Label(frame_parametros, text="Tasa Pago a Cuenta IR (%):").grid(row=5, column=0, sticky='w', pady=8)
        entry_tasa_pc = ttk.Entry(frame_parametros, textvariable=self.TASA_PAGO_CUENTA, width=20)
        entry_tasa_pc.grid(row=5, column=1, pady=8, padx=5)

        # SECCI√ìN: Estructura del Grupo
        frame_estructura = ttk.LabelFrame(scrollable_frame, text="Estructura del Grupo Empresarial", padding=20)
        frame_estructura.grid(row=1, column=0, padx=10, pady=10, sticky='ew', columnspan=2)
        
        ttk.Label(frame_estructura, text="N√∫mero de Empresas Sat√©lites:").grid(row=0, column=0, sticky='w', pady=8)
        spinbox_satelites = ttk.Spinbox(frame_estructura, from_=1, to=20, textvariable=self.num_satelites, width=18)
        spinbox_satelites.grid(row=0, column=1, pady=8, padx=5)
        
        ttk.Button(frame_estructura, text="Aplicar Cambios en Estructura", 
                  command=self.aplicar_estructura).grid(row=1, column=0, columnspan=3, pady=15)
        
        # SECCI√ìN NUEVA: Control de Escenarios Comparativos
        frame_comp_vars = ttk.LabelFrame(scrollable_frame, text="Variables para An√°lisis Comparativo", padding=20)
        frame_comp_vars.grid(row=2, column=0, padx=10, pady=10, sticky='ew', columnspan=2)

        ttk.Label(frame_comp_vars, text="Escenario L√≠mite (incremento %):").grid(row=0, column=0, sticky='w', pady=5)
        ttk.Entry(frame_comp_vars, textvariable=self.COMP_LIMITE_PCT, width=15).grid(row=0, column=1, padx=5)

        ttk.Label(frame_comp_vars, text="Escenario Tasa (reducci√≥n puntos):").grid(row=1, column=0, sticky='w', pady=5)
        ttk.Entry(frame_comp_vars, textvariable=self.COMP_TASA_DELTA, width=15).grid(row=1, column=1, padx=5)
        
        
        # SECCI√ìN: Presets
        frame_presets = ttk.LabelFrame(scrollable_frame, text="Configuraciones Predefinidas", padding=20)
        frame_presets.grid(row=2, column=0, padx=10, pady=10, sticky='ew', columnspan=2)
        
        btn_frame = ttk.Frame(frame_presets)
        btn_frame.grid(row=1, column=0, columnspan=3, pady=10)
        
        ttk.Button(btn_frame, text="Per√∫", command=lambda: self.cargar_preset('peru')).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Colombia", command=lambda: self.cargar_preset('colombia')).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="M√©xico", command=lambda: self.cargar_preset('mexico')).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Chile", command=lambda: self.cargar_preset('chile')).pack(side='left', padx=5)
        
        # Informaci√≥n actual
        frame_info = ttk.LabelFrame(scrollable_frame, text="Par√°metros Calculados", padding=20)
        frame_info.grid(row=4, column=0, padx=10, pady=10, sticky='ew', columnspan=2)
        
        self.label_info = ttk.Label(frame_info, text="", justify='left', font=('Courier', 9))
        self.label_info.pack(fill='both', expand=True)
        
        self.actualizar_info_parametros()
        
        ttk.Button(scrollable_frame, text="Actualizar Informaci√≥n", 
                  command=self.actualizar_info_parametros).grid(row=5, column=0, columnspan=2, pady=10)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def crear_tab_datos(self, parent):
        self.canvas_datos = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.canvas_datos.yview)
        self.scrollable_datos = ttk.Frame(self.canvas_datos)
        
        self.scrollable_datos.bind(
            "<Configure>",
            lambda e: self.canvas_datos.configure(scrollregion=self.canvas_datos.bbox("all"))
        )
        
        self.canvas_datos.create_window((0, 0), window=self.scrollable_datos, anchor="nw")
        self.canvas_datos.configure(yscrollcommand=scrollbar.set)
        
        # SECCI√ìN MATRIZ
        self.frame_matriz = ttk.LabelFrame(self.scrollable_datos, text="Empresa Matriz (R√©gimen General)", padding=15)
        self.frame_matriz.grid(row=0, column=0, padx=10, pady=10, sticky='ew')
        
        ttk.Label(self.frame_matriz, text="Nombre Empresa Matriz:").grid(row=0, column=0, sticky='w', pady=5)
        self.entry_nombre_matriz = ttk.Entry(self.frame_matriz, width=30)
        self.entry_nombre_matriz.grid(row=0, column=1, pady=5, columnspan=2)
        self.entry_nombre_matriz.insert(0, "KALLPA")
        
        ttk.Label(self.frame_matriz, text="Ingresos Totales:").grid(row=1, column=0, sticky='w', pady=5)
        self.entry_ingresos_matriz = ttk.Entry(self.frame_matriz, width=20)
        self.entry_ingresos_matriz.grid(row=1, column=1, pady=5)
        self.entry_ingresos_matriz.insert(0, "10000000")
        
        ttk.Label(self.frame_matriz, text="Costos Externos:").grid(row=2, column=0, sticky='w', pady=5)
        self.entry_costos_matriz = ttk.Entry(self.frame_matriz, width=20)
        self.entry_costos_matriz.grid(row=2, column=1, pady=5)
        self.entry_costos_matriz.insert(0, "6000000")
        
        # SECCI√ìN SAT√âLITES
        self.frame_satelites_container = ttk.LabelFrame(self.scrollable_datos, text="Empresas Sat√©lites", padding=15)
        self.frame_satelites_container.grid(row=1, column=0, padx=10, pady=10, sticky='ew')
        
        self.generar_campos_satelites()
        
        # Botones
        frame_botones = ttk.Frame(self.scrollable_datos)
        frame_botones.grid(row=2, column=0, pady=20)
        
        ttk.Button(frame_botones, text="Calcular M√°rgenes √ìptimos", 
                  command=self.calcular_margenes_optimos).pack(side='left', padx=5)
        ttk.Button(frame_botones, text="Limpiar", command=self.limpiar_datos).pack(side='left', padx=5)
        ttk.Button(frame_botones, text="Cargar Ejemplo", command=self.cargar_ejemplo).pack(side='left', padx=5)
        
        self.canvas_datos.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def generar_campos_satelites(self):
        for widget in self.frame_satelites_container.winfo_children():
            widget.destroy()
        
        # Header con nueva columna para r√©gimen
        headers = [
            ("Nombre", 0),
            ("Costos Compra", 1),
            ("Tipo", 2),
            ("Gastos Operativos", 3),
            ("R√©gimen General", 4)  # Nueva columna
        ]
        
        for text, col in headers:
            ttk.Label(self.frame_satelites_container, text=text, font='TkDefaultFont 9 bold').grid(
                row=0, column=col, padx=5, pady=5)
        
        self.satelites_entries = []
        num_sat = self.num_satelites.get()
        
        nombres_default = ["IG PAME", "SUMAQ WARMY", "MISKY SONKO", "KHASP", "EMPRESA E", "EMPRESA F"]
        tipos_default = ["Servicios", "Productos", "Productos", "Productos", "Productos", "Servicios"]
        
        for i in range(num_sat):
            nombre_default = nombres_default[i] if i < len(nombres_default) else f"SATELITE {i+1}"
            tipo_default = tipos_default[i] if i < len(tipos_default) else "Productos"
            
            entry_nombre = ttk.Entry(self.frame_satelites_container, width=18)
            entry_nombre.grid(row=i+1, column=0, pady=5, padx=5)
            entry_nombre.insert(0, nombre_default)
            
            entry_costo = ttk.Entry(self.frame_satelites_container, width=18)
            entry_costo.grid(row=i+1, column=1, pady=5, padx=5)
            entry_costo.insert(0, "500000")
            
            combo_tipo = ttk.Combobox(self.frame_satelites_container, 
                                     values=["Productos", "Servicios", "Manufactura", "Distribuci√≥n"],
                                     state='readonly', width=15)
            combo_tipo.grid(row=i+1, column=2, pady=5, padx=5)
            combo_tipo.set(tipo_default)
            
            entry_gastos = ttk.Entry(self.frame_satelites_container, width=18)
            entry_gastos.grid(row=i+1, column=3, pady=5, padx=5)
            entry_gastos.insert(0, "0")
            
            # --- AQU√ç EST√Å LA L√ìGICA QUIR√öRGICA ---
            var_regimen_general = tk.BooleanVar(value=False)
            check_regimen = ttk.Checkbutton(self.frame_satelites_container, 
                                           variable=var_regimen_general,
                                           text="Usar tasa general")
            check_regimen.grid(row=i+1, column=4, pady=5, padx=5)

            # Definimos la validaci√≥n interna para esta fila espec√≠fica
            def validar_uit(event, v_reg=var_regimen_general, e_c=entry_costo, e_g=entry_gastos, chk=check_regimen):
                try:
                    uit_actual = self.UIT.get()
                    limite_anual = 1700 * uit_actual
                    
                    # El ingreso que el sat√©lite cobrar√≠a a la matriz para ser eficiente es:
                    # Ingreso = Costos + Utilidad M√°xima (15 UIT) + Gastos Operativos
                    costo_val = float(e_c.get() or 0)
                    gastos_val = float(e_g.get() or 0)
                    utilidad_max = 15 * uit_actual
                    
                    ingreso_proyectado = costo_val + utilidad_max + gastos_val
                    
                    if ingreso_proyectado > limite_anual:
                        v_reg.set(True) # Forzamos el check
                        chk.config(state='disabled') # Bloqueamos el bot√≥n
                    else:
                        chk.config(state='normal') # Lo liberamos si baja el monto
                except ValueError:
                    pass # Evita errores si el campo est√° vac√≠o mientras escribes

            # "Escuchamos" cuando el usuario suelta una tecla en los campos de dinero
            entry_costo.bind("<KeyRelease>", validar_uit)
            entry_gastos.bind("<KeyRelease>", validar_uit)
            # ---------------------------------------

            self.satelites_entries.append({
                'entry_nombre': entry_nombre,
                'entry_costo': entry_costo,
                'combo_tipo': combo_tipo,
                'entry_gastos': entry_gastos,
                'var_regimen_general': var_regimen_general,
                'check_widget': check_regimen # Guardamos el widget para controlarlo
            })
    
    def crear_tab_resultados(self, parent):
        # 1. Botones de Exportaci√≥n arriba
        frame_export = ttk.Frame(parent)
        frame_export.pack(fill='x', padx=10, pady=5)
        ttk.Button(frame_export, text="üìä Exportar Reporte Ejecutivo Excel", command=self.exportar_excel).pack(side='left', padx=5)
        ttk.Button(frame_export, text="üìã Copiar Resultados", command=self.copiar_resultados).pack(side='left', padx=5)

        # 2. Texto de Resultados (Reducimos un poco el alto para dar espacio a gr√°ficos)
        frame_resultados = ttk.LabelFrame(parent, text="Reporte Detallado", padding=10)
        frame_resultados.pack(fill='x', padx=10, pady=5)
        self.text_resultados = tk.Text(frame_resultados, height=12, width=110, wrap='word', font=('Courier', 9))
        self.text_resultados.pack(side='left', fill='both', expand=True)

        # 3. Contenedor de Gr√°ficos
        self.frame_graficos_container = ttk.LabelFrame(parent, text="Visualizaci√≥n Estrat√©gica", padding=10)
        self.frame_graficos_container.pack(fill='both', expand=True, padx=10, pady=5)
        self.canvas_graficos = None
        
    def crear_tab_simulacion(self, parent):
        frame_config = ttk.LabelFrame(parent, text="Configuraci√≥n Cient√≠fica Monte Carlo", padding=15)
        frame_config.pack(fill='x', padx=10, pady=10)
        
        # Fila 1
        ttk.Label(frame_config, text="N¬∞ Simulaciones:").grid(row=0, column=0, sticky='w')
        self.entry_num_sim = ttk.Entry(frame_config, width=12); self.entry_num_sim.insert(0, "10000")
        self.entry_num_sim.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_config, text="Variabilidad (¬±%):").grid(row=0, column=2, sticky='w', padx=10)
        self.entry_variabilidad = ttk.Entry(frame_config, width=12); self.entry_variabilidad.insert(0, "10")
        self.entry_variabilidad.grid(row=0, column=3, padx=5, pady=5)

        # Fila 2: Nuevos Controles
        ttk.Label(frame_config, text="Nivel Confianza (%):").grid(row=1, column=0, sticky='w')
        ttk.Entry(frame_config, textvariable=self.CONF_NIVEL, width=12).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame_config, text="Alfa (Significancia):").grid(row=1, column=2, sticky='w', padx=10)
        ttk.Entry(frame_config, textvariable=self.ALFA_SIG, width=12).grid(row=1, column=3, padx=5, pady=5)

        ttk.Button(frame_config, text="EJECUTAR ANALISIS DE RIESGO", command=self.ejecutar_simulacion).grid(row=2, column=0, columnspan=4, pady=10)

        # --- MARCO: INCERTIDUMBRE DE LA MATRIZ Y CORRELACION TERRITORIAL ---
        frame_matriz = ttk.LabelFrame(frame_config, text="Incertidumbre de la Matriz y Correlacion Territorial", padding=10)
        frame_matriz.grid(row=3, column=0, columnspan=4, pady=10, sticky='ew')
        frame_matriz.columnconfigure(1, weight=1)

        # Fila 0: Checkbutton costos matriz
        ttk.Checkbutton(frame_matriz, text="Incluir variabilidad en costos externos de la matriz",
                        variable=self.VAR_MATRIZ_COSTOS).grid(row=0, column=0, columnspan=2, sticky='w', pady=5)

        # Fila 1: Variabilidad costos matriz
        ttk.Label(frame_matriz, text="Variabilidad costos matriz (+/-%):"
                  ).grid(row=1, column=0, sticky='w', padx=20, pady=5)
        self.entry_pct_var_matriz_costos = ttk.Entry(frame_matriz, textvariable=self.PCT_VAR_MATRIZ_COSTOS, width=12)
        self.entry_pct_var_matriz_costos.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        self.entry_pct_var_matriz_costos.config(state='disabled')

        # Fila 2: Checkbutton ingresos matriz
        ttk.Checkbutton(frame_matriz, text="Incluir variabilidad en ingresos de la matriz",
                        variable=self.VAR_MATRIZ_INGRESOS).grid(row=2, column=0, columnspan=2, sticky='w', pady=5)

        # Fila 3: Variabilidad ingresos matriz
        ttk.Label(frame_matriz, text="Variabilidad ingresos matriz (+/-%):"
                  ).grid(row=3, column=0, sticky='w', padx=20, pady=5)
        self.entry_pct_var_matriz_ingresos = ttk.Entry(frame_matriz, textvariable=self.PCT_VAR_MATRIZ_INGRESOS, width=12)
        self.entry_pct_var_matriz_ingresos.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        self.entry_pct_var_matriz_ingresos.config(state='disabled')

        # Fila 4: Distribucion de factores
        ttk.Label(frame_matriz, text="Distribucion de factores:").grid(row=4, column=0, sticky='w', pady=5)
        ttk.Combobox(frame_matriz, textvariable=self.DISTRIBUCION_FACTORES,
                     values=["Log-normal", "Normal"], state='readonly', width=15
                     ).grid(row=4, column=1, sticky='w', padx=5, pady=5)

        # Fila 5: Correlacion matriz-satelites
        ttk.Label(frame_matriz, text="Correlacion matriz-satelites (rho):"
                  ).grid(row=5, column=0, sticky='w', pady=5)
        ttk.Entry(frame_matriz, textvariable=self.CORRELACION_MATRIZ_SAT, width=12
                  ).grid(row=5, column=1, sticky='w', padx=5, pady=5)

        # Callbacks para habilitar/deshabilitar Entries
        def _toggle_var_costos(*args):
            state = 'normal' if self.VAR_MATRIZ_COSTOS.get() else 'disabled'
            self.entry_pct_var_matriz_costos.config(state=state)
        self.VAR_MATRIZ_COSTOS.trace_add('write', _toggle_var_costos)

        def _toggle_var_ingresos(*args):
            state = 'normal' if self.VAR_MATRIZ_INGRESOS.get() else 'disabled'
            self.entry_pct_var_matriz_ingresos.config(state=state)
        self.VAR_MATRIZ_INGRESOS.trace_add('write', _toggle_var_ingresos)

        self.frame_sim_container = ttk.LabelFrame(parent, text="Resultados Estadisticos", padding=15)
        self.frame_sim_container.pack(fill='both', expand=True, padx=10, pady=10)
        self.canvas_simulacion = None
    
    def crear_tab_sensibilidad(self, parent):
        frame_config = ttk.LabelFrame(parent, text="An√°lisis de Sensibilidad", padding=15)
        frame_config.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(frame_config, text="Variable:").grid(row=0, column=0, sticky='w', pady=5)
        self.combo_variable = ttk.Combobox(frame_config, 
                                          values=["Costos Sat√©lites", "Tasa R√©gimen Especial", "L√≠mite Utilidad"],
                                          state='readonly', width=25)
        self.combo_variable.grid(row=0, column=1, pady=5)
        self.combo_variable.current(0)
        
        ttk.Label(frame_config, text="Rango (¬±%):").grid(row=1, column=0, sticky='w', pady=5)
        self.entry_rango_sens = ttk.Entry(frame_config, width=15)
        self.entry_rango_sens.grid(row=1, column=1, pady=5)
        self.entry_rango_sens.insert(0, "30")
        
        ttk.Button(frame_config, text="Ejecutar", command=self.ejecutar_sensibilidad).grid(row=2, column=0, columnspan=2, pady=10)
        
        frame_sens_resultados = ttk.LabelFrame(parent, text="Resultados", padding=15)
        frame_sens_resultados.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.canvas_sensibilidad = None
        self.frame_sens_container = frame_sens_resultados
        
    def crear_tab_comparativo(self, parent):
        frame_config = ttk.LabelFrame(parent, text="An√°lisis Comparativo", padding=15)
        frame_config.pack(fill='x', padx=10, pady=10)
        
        ttk.Button(frame_config, text="Generar An√°lisis", 
                  command=self.generar_comparativo).pack(pady=10)
        
        frame_comp_resultados = ttk.LabelFrame(parent, text="Resultados", padding=15)
        frame_comp_resultados.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.canvas_comparativo = None
        self.frame_comp_container = frame_comp_resultados
        
    def crear_tab_pagos_cuenta(self, parent):
        canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        frame_config_pc = ttk.LabelFrame(scrollable_frame,
                                         text="Analisis de Pagos a Cuenta y Equilibrio Fiscal (Teoria de Juegos)",
                                         padding=15)
        frame_config_pc.grid(row=0, column=0, padx=10, pady=10, sticky='ew')

        ttk.Label(frame_config_pc, text="Tasa Pago a Cuenta (%):").grid(row=0, column=0, sticky='w', pady=5)
        ttk.Entry(frame_config_pc, textvariable=self.TASA_PAGO_CUENTA, width=12).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_config_pc, text="Pasos Sensibilidad:").grid(row=0, column=2, sticky='w', padx=10, pady=5)
        self.entry_pasos_sens_pc = ttk.Entry(frame_config_pc, width=12)
        self.entry_pasos_sens_pc.insert(0, "15")
        self.entry_pasos_sens_pc.grid(row=0, column=3, padx=5, pady=5)

        ttk.Button(frame_config_pc, text="Ejecutar Analisis Pagos a Cuenta y Sensibilidad",
                   command=self.analizar_pagos_cuenta).grid(row=1, column=0, columnspan=2, pady=10)

        ttk.Label(frame_config_pc, text="Sim. Equilibrio (N):").grid(row=1, column=2, sticky='w', padx=10)
        self.entry_n_sim_eq = ttk.Entry(frame_config_pc, width=12)
        self.entry_n_sim_eq.insert(0, "5000")
        self.entry_n_sim_eq.grid(row=1, column=3, padx=5, pady=5)

        ttk.Button(frame_config_pc, text="Simular Equilibrio (Monte Carlo)",
                   command=self.simular_equilibrio_financiero).grid(row=2, column=0, columnspan=4, pady=10)

        # Texto de resultados
        frame_texto_pc = ttk.LabelFrame(scrollable_frame, text="Analisis Detallado por Satelite", padding=10)
        frame_texto_pc.grid(row=1, column=0, padx=10, pady=5, sticky='ew')
        self.text_pagos_cuenta = tk.Text(frame_texto_pc, height=18, width=140, wrap='none', font=('Courier', 8))
        scroll_h = ttk.Scrollbar(frame_texto_pc, orient='horizontal', command=self.text_pagos_cuenta.xview)
        scroll_v = ttk.Scrollbar(frame_texto_pc, orient='vertical', command=self.text_pagos_cuenta.yview)
        self.text_pagos_cuenta.configure(xscrollcommand=scroll_h.set, yscrollcommand=scroll_v.set)
        self.text_pagos_cuenta.grid(row=0, column=0, sticky='nsew')
        scroll_v.grid(row=0, column=1, sticky='ns')
        scroll_h.grid(row=1, column=0, sticky='ew')
        frame_texto_pc.columnconfigure(0, weight=1)
        frame_texto_pc.rowconfigure(0, weight=1)

        # Graficos
        self.frame_pagos_graficos = ttk.LabelFrame(scrollable_frame, text="Visualizacion Pagos a Cuenta vs IR", padding=10)
        self.frame_pagos_graficos.grid(row=2, column=0, padx=10, pady=5, sticky='ew')
        self.canvas_pagos_cuenta = None

        # Graficos simulacion equilibrio
        self.frame_eq_sim_graficos = ttk.LabelFrame(scrollable_frame,
                                                     text="Simulacion Monte Carlo - Equilibrio Financiero", padding=10)
        self.frame_eq_sim_graficos.grid(row=3, column=0, padx=10, pady=5, sticky='ew')
        self.canvas_equilibrio_sim = None

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def actualizar_info_parametros(self):
        uit = self.UIT.get()
        tasa_gral = self.TASA_GENERAL.get()
        tasa_esp = self.TASA_ESPECIAL.get()
        lim_util_uit = self.LIMITE_UTILIDAD_UIT.get()
        lim_ing_uit = self.LIMITE_INGRESOS_UIT.get()
        
        lim_util_soles = lim_util_uit * uit
        lim_ing_soles = lim_ing_uit * uit
        diferencial = tasa_gral - tasa_esp
        
        ahorro_max = lim_util_soles * (diferencial / 100) * self.num_satelites.get()
        
        tasa_pc = self.TASA_PAGO_CUENTA.get()

        info_text = f"""
PARAMETROS ACTUALES:

UIT:                                {uit:>15,.2f}
Tasa Regimen General:               {tasa_gral:>15.2f}%
Tasa Regimen Especial:              {tasa_esp:>15.2f}%
Diferencial:                        {diferencial:>15.2f} puntos
Tasa Pago a Cuenta IR:              {tasa_pc:>15.2f}%

Limite Utilidad:                    {lim_util_soles:>15,.2f}
Limite Ingresos:                    {lim_ing_soles:>15,.2f}

Ahorro maximo total:                {ahorro_max:>15,.2f}
Numero satelites:                   {self.num_satelites.get():>15}

FORMULA MARGEN OPTIMO (con gastos):
m* = (Limite_Utilidad + Gastos_Operativos) / Costos

FORMULA MARGEN EQUILIBRIO (IR = Pago a Cuenta):
m_eq = (t_pc * C + t_esp * G) / (C * (t_esp - t_pc))

Margen Equilibrio sin gastos:       {(tasa_pc / (tasa_esp - tasa_pc) * 100):>15.2f}%
        """
        
        self.label_info.config(text=info_text)
        
    def cargar_preset(self, pais):
        presets = {
            'peru': {'UIT': 5500, 'TASA_GENERAL': 29.5, 'TASA_ESPECIAL': 10.0, 'LIMITE_UTILIDAD_UIT': 15, 'LIMITE_INGRESOS_UIT': 1700},
            'colombia': {'UIT': 44000, 'TASA_GENERAL': 35.0, 'TASA_ESPECIAL': 9.0, 'LIMITE_UTILIDAD_UIT': 100, 'LIMITE_INGRESOS_UIT': 1400},
            'mexico': {'UIT': 96000, 'TASA_GENERAL': 30.0, 'TASA_ESPECIAL': 10.0, 'LIMITE_UTILIDAD_UIT': 50, 'LIMITE_INGRESOS_UIT': 1500},
            'chile': {'UIT': 600000, 'TASA_GENERAL': 27.0, 'TASA_ESPECIAL': 10.0, 'LIMITE_UTILIDAD_UIT': 20, 'LIMITE_INGRESOS_UIT': 1000}
        }
        
        if pais in presets:
            config = presets[pais]
            self.UIT.set(config['UIT'])
            self.TASA_GENERAL.set(config['TASA_GENERAL'])
            self.TASA_ESPECIAL.set(config['TASA_ESPECIAL'])
            self.LIMITE_UTILIDAD_UIT.set(config['LIMITE_UTILIDAD_UIT'])
            self.LIMITE_INGRESOS_UIT.set(config['LIMITE_INGRESOS_UIT'])
            self.actualizar_info_parametros()
            messagebox.showinfo("Preset Cargado", f"Configuraci√≥n de {pais.upper()} cargada")
        
    def aplicar_estructura(self):
        self.generar_campos_satelites()
        self.actualizar_info_parametros()
        messagebox.showinfo("Actualizado", f"Estructura con {self.num_satelites.get()} sat√©lites")
        
  
    def calcular_margenes_optimos(self):
        try:
            # Par√°metros
            uit = self.UIT.get()
            tasa_general = self.TASA_GENERAL.get() / 100
            tasa_especial = self.TASA_ESPECIAL.get() / 100
            limite_utilidad = self.LIMITE_UTILIDAD_UIT.get() * uit
            limite_ingresos = self.LIMITE_INGRESOS_UIT.get() * uit
            
            # Matriz
            nombre_matriz = self.entry_nombre_matriz.get()
            ingresos = float(self.entry_ingresos_matriz.get())
            costos_externos = float(self.entry_costos_matriz.get())
            
            utilidad_sin_estructura = ingresos - costos_externos
            impuesto_sin_estructura = utilidad_sin_estructura * tasa_general
            
            # Sat√©lites
            satelites_data = []
            advertencias = []
            
            for idx, sat in enumerate(self.satelites_entries, 1):
                nombre = sat['entry_nombre'].get()
                costo = float(sat['entry_costo'].get())
                tipo = sat['combo_tipo'].get()
                gastos = float(sat['entry_gastos'].get())
                usar_regimen_general = sat['var_regimen_general'].get()
                
                # VERIFICAR SI PUEDE ESTAR EN R√âGIMEN ESPECIAL
                margen_para_limite = (limite_utilidad + gastos) / costo
                precio_venta_estimado = costo * (1 + margen_para_limite)
                
                puede_regimen_especial = True
                razon_no_puede = ""
                
                # Verificar si el margen ser√≠a muy alto (> 100%)
                if margen_para_limite > 1.0:
                    puede_regimen_especial = False
                    razon_no_puede = f"Margen requerido muy alto ({margen_para_limite*100:.1f}%)"
                
                # Verificar l√≠mite de ingresos
                if precio_venta_estimado > limite_ingresos:
                    puede_regimen_especial = False
                    razon_no_puede = f"Supera l√≠mite ingresos"
                
                # ADVERTENCIA: Si marca r√©gimen general pudiendo estar en especial
                if usar_regimen_general and puede_regimen_especial and gastos == 0:
                    ahorro_potencial = limite_utilidad * (tasa_general - tasa_especial)
                    advertencias.append(
                        f"‚ö†Ô∏è {nombre}: Marcado R√âGIMEN GENERAL sin necesidad.\n"
                        f"   Ahorro potencial perdido: {ahorro_potencial:,.2f}\n"
                        f"   RECOMENDACI√ìN: Desmarcar y usar R√©gimen Especial\n"
                    )
                
                # ADVERTENCIA: Si NO puede especial pero no est√° marcado
                if not usar_regimen_general and not puede_regimen_especial:
                    advertencias.append(
                        f"‚ö†Ô∏è {nombre}: NO califica para R√©gimen Especial.\n"
                        f"   Raz√≥n: {razon_no_puede}\n"
                        f"   Se aplicar√° autom√°ticamente R√©gimen General\n"
                    )
                    usar_regimen_general = True
                
                satelites_data.append({
                    'nombre': nombre,
                    'tipo': tipo,
                    'costo': costo,
                    'gastos_operativos': gastos,
                    'usar_regimen_general': usar_regimen_general,
                    'puede_regimen_especial': puede_regimen_especial,
                    'razon_no_puede': razon_no_puede
                })
            
            # Mostrar advertencias
            if advertencias:
                mensaje = "ADVERTENCIAS DEL SISTEMA:\n\n" + "\n".join(advertencias) + "\n\n¬øContinuar?"
                if messagebox.askquestion("Advertencias", mensaje, icon='warning') == 'no':
                    return
            
            # CALCULAR M√ÅRGENES √ìPTIMOS
            resultados_satelites = []
            total_utilidad_satelites = 0
            total_impuesto_satelites = 0
            total_compras_matriz = 0
            
            for sat in satelites_data:
                costo = sat['costo']
                gastos = sat['gastos_operativos']
                usar_regimen_general = sat['usar_regimen_general']
                
                if usar_regimen_general:
                    # ============================================
                    # R√âGIMEN GENERAL - L√ìGICA MEJORADA
                    # ============================================
                    
                    if usar_regimen_general:
                        # ============================================
                        # R√âGIMEN GENERAL - L√ìGICA CORREGIDA
                        # ============================================
                        
                        if gastos > 0:
                            # CON GASTOS: Margen al punto de equilibrio
                            margen_optimo = gastos / costo
                            utilidad_bruta = costo * margen_optimo
                            utilidad_neta = 0  # Punto equilibrio
                            precio_venta = costo * (1 + margen_optimo)
                            
                            impuesto_satelite = 0
                            tasa_efectiva = 0
                            
                            # AHORRO: KALLPA deja de pagar 29.5% sobre el gasto
                            ahorro_individual = gastos * tasa_general
                            
                            regimen = "GENERAL (Punto Equilibrio)"
                            ajuste_aplicado = f"Margen {margen_optimo*100:.2f}% para cubrir gastos ({gastos:,.0f}). Ahorro: {ahorro_individual:,.0f}"
                            
                        else:
                            # SIN GASTOS: Margen CERO (solo traspaso)
                            margen_optimo = 0.0
                            utilidad_bruta = 0
                            utilidad_neta = 0
                            precio_venta = costo  # Solo traspasa el costo
                            
                            impuesto_satelite = 0
                            tasa_efectiva = 0
                            ahorro_individual = 0
                            
                            regimen = "GENERAL (Traspaso sin margen)"
                            ajuste_aplicado = "‚ö†Ô∏è SIN MARGEN - No genera ahorro. RECOMENDACI√ìN: Eliminar de estructura o desmarcar r√©gimen general"
                    
                else:
                    # ============================================
                    # R√âGIMEN ESPECIAL - L√ìGICA ORIGINAL
                    # ============================================
                    
                    # F√≥rmula mejorada: m* = (Limite_Utilidad + Gastos) / Costos
                    margen_optimo = (limite_utilidad + gastos) / costo
                    
                    # Verificar l√≠mite de ingresos
                    precio_venta_tentativo = costo * (1 + margen_optimo)
                    if precio_venta_tentativo > limite_ingresos:
                        margen_optimo = (limite_ingresos / costo) - 1
                        ajuste_aplicado = f"Ajustado por l√≠mite ingresos"
                    else:
                        if gastos > 0:
                            ajuste_aplicado = f"Margen incrementado por gastos ({gastos:,.0f})"
                        else:
                            ajuste_aplicado = "Margen √≥ptimo est√°ndar"
                    
                    utilidad_bruta = costo * margen_optimo
                    utilidad_neta = utilidad_bruta - gastos
                    precio_venta = costo * (1 + margen_optimo)
                    
                    # Calcular impuesto
                    if utilidad_neta <= limite_utilidad:
                        impuesto_satelite = utilidad_neta * tasa_especial
                        tasa_efectiva = tasa_especial * 100
                    else:
                        impuesto_base = limite_utilidad * tasa_especial
                        impuesto_exceso = (utilidad_neta - limite_utilidad) * tasa_general
                        impuesto_satelite = impuesto_base + impuesto_exceso
                        tasa_efectiva = (impuesto_satelite / utilidad_neta * 100) if utilidad_neta > 0 else 0
                    
                    # Ahorro: Diferencial de tasas
                    impuesto_si_fuera_general = utilidad_neta * tasa_general
                    ahorro_individual = impuesto_si_fuera_general - impuesto_satelite
                    
                    regimen = "ESPECIAL"
                
                # Pagos a cuenta (1.5% del precio de venta)
                tasa_pc = self.TASA_PAGO_CUENTA.get() / 100
                pago_a_cuenta = tasa_pc * precio_venta
                saldo_regularizacion = impuesto_satelite - pago_a_cuenta

                # Margen de equilibrio: donde IR = pago a cuenta
                margen_equilibrio_pc = None
                if not usar_regimen_general and (tasa_especial - tasa_pc) > 0:
                    m_eq = (tasa_pc * costo + tasa_especial * gastos) / (costo * (tasa_especial - tasa_pc))
                    util_eq = costo * m_eq - gastos
                    if util_eq <= limite_utilidad and costo * (1 + m_eq) <= limite_ingresos:
                        margen_equilibrio_pc = m_eq * 100
                elif usar_regimen_general and (tasa_general - tasa_pc) > 0:
                    m_eq = (tasa_pc * costo + tasa_general * gastos) / (costo * (tasa_general - tasa_pc))
                    margen_equilibrio_pc = m_eq * 100

                # Acumular totales
                total_utilidad_satelites += max(0, utilidad_neta)
                total_impuesto_satelites += impuesto_satelite
                total_compras_matriz += precio_venta

                resultados_satelites.append({
                    'nombre': sat['nombre'],
                    'tipo': sat['tipo'],
                    'regimen': regimen,
                    'costo': costo,
                    'gastos_operativos': gastos,
                    'margen_optimo': margen_optimo * 100,
                    'precio_venta': precio_venta,
                    'utilidad_bruta': utilidad_bruta,
                    'utilidad_neta': max(0, utilidad_neta),
                    'impuesto': impuesto_satelite,
                    'tasa_efectiva': tasa_efectiva,
                    'rentabilidad': (max(0, utilidad_neta) / precio_venta * 100) if precio_venta > 0 else 0,
                    'ahorro_individual': ahorro_individual,
                    'ajuste_aplicado': ajuste_aplicado,
                    'pago_a_cuenta': pago_a_cuenta,
                    'saldo_regularizacion': saldo_regularizacion,
                    'margen_equilibrio_pc': margen_equilibrio_pc
                })
            
            # Calcular nueva utilidad matriz
            nueva_utilidad_matriz = utilidad_sin_estructura - total_compras_matriz + sum([s['costo'] for s in satelites_data])
            impuesto_matriz = nueva_utilidad_matriz * tasa_general
            
            # Totales
            impuesto_total_grupo = impuesto_matriz + total_impuesto_satelites
            ahorro_tributario = impuesto_sin_estructura - impuesto_total_grupo
            ahorro_porcentual = (ahorro_tributario / impuesto_sin_estructura * 100) if impuesto_sin_estructura > 0 else 0
            
            # Almacenar resultados
            self.resultados = {
                'parametros': {
                    'uit': uit,
                    'tasa_general': tasa_general * 100,
                    'tasa_especial': tasa_especial * 100,
                    'limite_utilidad': limite_utilidad,
                    'limite_ingresos': limite_ingresos
                },
                'matriz': {
                    'nombre': nombre_matriz,
                    'ingresos': ingresos,
                    'costos_externos': costos_externos,
                    'utilidad_sin_estructura': utilidad_sin_estructura,
                    'impuesto_sin_estructura': impuesto_sin_estructura,
                    'nueva_utilidad': nueva_utilidad_matriz,
                    'impuesto_con_estructura': impuesto_matriz,
                    'total_compras': total_compras_matriz
                },
                'satelites': resultados_satelites,
                'grupo': {
                    'total_utilidad_satelites': total_utilidad_satelites,
                    'total_impuesto_satelites': total_impuesto_satelites,
                    'impuesto_total': impuesto_total_grupo,
                    'ahorro_tributario': ahorro_tributario,
                    'ahorro_porcentual': ahorro_porcentual
                }
            }
            
            self.mostrar_resultados()
            self.generar_graficos()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Valores inv√°lidos:\n{str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Error inesperado:\n{str(e)}")

    
    def mostrar_resultados(self):
        self.text_resultados.delete(1.0, tk.END)
        
        r = self.resultados
        p = r['parametros']
        
        texto = f"""
{'='*110}
                            REPORTE DE OPTIMIZACI√ìN TRIBUTARIA CON AJUSTE POR GASTOS
                                    AN√ÅLISIS DE GRUPOS EMPRESARIALES
{'='*110}
Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

PAR√ÅMETROS TRIBUTARIOS:
  UIT:                                 {p['uit']:>15,.2f}
  Tasa R√©gimen General:                {p['tasa_general']:>15.2f}%
  Tasa R√©gimen Especial:               {p['tasa_especial']:>15.2f}%
  Diferencial:                         {p['tasa_general'] - p['tasa_especial']:>15.2f} puntos
  L√≠mite Utilidad Especial:            {p['limite_utilidad']:>15,.2f}
  L√≠mite Ingresos Especial:            {p['limite_ingresos']:>15,.2f}
{'='*110}

ESCENARIO 1: SIN ESTRUCTURA SATELITAL
{'='*110}
{r['matriz']['nombre']} - R√©gimen General
  Ingresos:                            {r['matriz']['ingresos']:>15,.2f}
  Costos Externos:                     {r['matriz']['costos_externos']:>15,.2f}
  Utilidad:                            {r['matriz']['utilidad_sin_estructura']:>15,.2f}
  Impuesto:                            {r['matriz']['impuesto_sin_estructura']:>15,.2f}
  Utilidad Neta:                       {r['matriz']['utilidad_sin_estructura'] - r['matriz']['impuesto_sin_estructura']:>15,.2f}

{'='*110}
ESCENARIO 2: CON ESTRUCTURA SATELITAL OPTIMIZADA (AJUSTE POR GASTOS OPERATIVOS)
{'='*110}

EMPRESAS SAT√âLITES
{'-'*110}
"""
        
        for i, sat in enumerate(r['satelites'], 1):
            texto += f"""
{i}. {sat['nombre']} - {sat['tipo']} [{sat['regimen']}]
   Costos Compra:                      {sat['costo']:>15,.2f}
   Gastos Operativos:                  {sat['gastos_operativos']:>15,.2f}
   Margen √ìptimo:                      {sat['margen_optimo']:>15.2f}%
   Precio Venta:                       {sat['precio_venta']:>15,.2f}
   Utilidad Bruta:                     {sat['utilidad_bruta']:>15,.2f}
   Utilidad Neta:                      {sat['utilidad_neta']:>15,.2f}
   Impuesto (Tasa: {sat['tasa_efectiva']:.2f}%):           {sat['impuesto']:>15,.2f}
   Ahorro Individual:                  {sat['ahorro_individual']:>15,.2f}
   ROI:                                {sat['rentabilidad']:>15.2f}%
   Pago a Cuenta ({self.TASA_PAGO_CUENTA.get():.1f}%):            {sat['pago_a_cuenta']:>15,.2f}
   Saldo Regularizacion:               {sat['saldo_regularizacion']:>15,.2f}  {'(Favor)' if sat['saldo_regularizacion'] < 0 else '(Pagar)'}
   Margen Equilibrio (IR=PaC):         {f"{sat['margen_equilibrio_pc']:.4f}%" if sat['margen_equilibrio_pc'] is not None else "N/A":>15}
   Ajuste: {sat['ajuste_aplicado']}
"""
        
        utilidad_total = r['matriz']['nueva_utilidad'] + r['grupo']['total_utilidad_satelites']
        
        texto += f"""
{'-'*110}
CONSOLIDADO SATELITES:
  Total Utilidad Neta:                 {r['grupo']['total_utilidad_satelites']:>15,.2f}
  Total Impuesto:                      {r['grupo']['total_impuesto_satelites']:>15,.2f}
  Ahorro Total Satelites:              {sum([s['ahorro_individual'] for s in r['satelites']]):>15,.2f}
  Total Pagos a Cuenta:                {sum([s['pago_a_cuenta'] for s in r['satelites']]):>15,.2f}
  Total Saldo Regularizacion:          {sum([s['saldo_regularizacion'] for s in r['satelites']]):>15,.2f}

{'-'*110}
{r['matriz']['nombre']} (con estructura):
  Compras a Sat√©lites:                 {r['matriz']['total_compras']:>15,.2f}
  Nueva Utilidad:                      {r['matriz']['nueva_utilidad']:>15,.2f}
  Impuesto:                            {r['matriz']['impuesto_con_estructura']:>15,.2f}

{'='*110}
CONSOLIDADO GRUPO
{'='*110}
  Utilidad Total:                      {utilidad_total:>15,.2f}
  Impuesto Total:                      {r['grupo']['impuesto_total']:>15,.2f}
  
  AHORRO TRIBUTARIO:                   {r['grupo']['ahorro_tributario']:>15,.2f}
  AHORRO PORCENTUAL:                   {r['grupo']['ahorro_porcentual']:>15.2f}%
  
  Tasa Efectiva Grupo:                 {(r['grupo']['impuesto_total'] / utilidad_total * 100):>15.2f}%
  Reducci√≥n vs Sin Estructura:         {p['tasa_general'] - (r['grupo']['impuesto_total'] / utilidad_total * 100):>15.2f} puntos

{'='*110}
F√ìRMULA APLICADA (MEJORADA CON GASTOS):
{'='*110}
Margen √ìptimo = (L√≠mite_Utilidad + Gastos_Operativos) / Costos

Esta f√≥rmula garantiza que la utilidad neta (despu√©s de gastos) alcance exactamente el l√≠mite
permitido en el r√©gimen especial, maximizando el ahorro tributario.
{'='*110}
"""
        
        self.text_resultados.insert(1.0, texto)
    
    def generar_graficos(self):
        if self.canvas_graficos:
            self.canvas_graficos.get_tk_widget().destroy()
        
        r = self.resultados
        
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('An√°lisis Tributario - Grupo Empresarial', fontsize=16, fontweight='bold')
        
        # Gr√°fico 1: Comparaci√≥n impuestos
        escenarios = ['Sin Estructura', 'Con Estructura']
        impuestos = [r['matriz']['impuesto_sin_estructura'], r['grupo']['impuesto_total']]
        colores = ['#e74c3c', '#2ecc71']
        
        bars1 = ax1.bar(escenarios, impuestos, color=colores, alpha=0.7, edgecolor='black', linewidth=2)
        ax1.set_ylabel('Impuesto Total', fontsize=10)
        ax1.set_title('Comparaci√≥n de Carga Tributaria', fontsize=12, fontweight='bold')
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.2f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
        
        for bar in bars1:
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold')
        
        ax1.text(0.5, max(impuestos) * 0.5, 
                f'Ahorro: {r["grupo"]["ahorro_tributario"]:,.0f}\n({r["grupo"]["ahorro_porcentual"]:.2f}%)',
                ha='center', fontsize=11, fontweight='bold',
                bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.5))
        
        # Gr√°fico 2: M√°rgenes por sat√©lite con colores seg√∫n r√©gimen
        nombres_sat = [s['nombre'][:12] for s in r['satelites']]
        margenes = [s['margen_optimo'] for s in r['satelites']]
        colores_regimen = ['#3498db' if s['regimen'] == 'ESPECIAL' else '#e67e22' for s in r['satelites']]
        
        bars2 = ax2.barh(nombres_sat, margenes, color=colores_regimen, alpha=0.7, edgecolor='black')
        ax2.set_xlabel('Margen √ìptimo (%)', fontsize=10)
        ax2.set_title('M√°rgenes √ìptimos (Azul=Especial, Naranja=General)', fontsize=11, fontweight='bold')
        ax2.invert_yaxis()
        
        for i, bar in enumerate(bars2):
            width = bar.get_width()
            ax2.text(width + max(margenes)*0.02, bar.get_y() + bar.get_height()/2.,
                    f'{width:.1f}%',
                    ha='left', va='center', fontsize=8, fontweight='bold')
        
        # Gr√°fico 3: Ahorro individual por sat√©lite
        ahorros_ind = [max(0, s['ahorro_individual']) for s in r['satelites']]
        nombres_cortos = [s['nombre'].split()[0] for s in r['satelites']]
        
        bars3 = ax3.bar(nombres_cortos, ahorros_ind, color='#9b59b6', alpha=0.7, edgecolor='black')
        ax3.set_ylabel('Ahorro Tributario', fontsize=10)
        ax3.set_title('Ahorro Individual por Sat√©lite', fontsize=12, fontweight='bold')
        ax3.tick_params(axis='x', rotation=45)
        ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
        ax3.grid(axis='y', alpha=0.3)
        
        # Gr√°fico 4: Utilidad Neta vs Gastos
        utilidades = [max(0, s['utilidad_neta']) for s in r['satelites']]
        gastos = [s['gastos_operativos'] for s in r['satelites']]
        
        x = np.arange(len(nombres_sat))
        width = 0.35
        
        ax4.bar(x - width/2, utilidades, width, label='Utilidad Neta', color='#2ecc71', alpha=0.7, edgecolor='black')
        ax4.bar(x + width/2, gastos, width, label='Gastos Operativos', color='#e74c3c', alpha=0.7, edgecolor='black')
        
        ax4.set_ylabel('Monto', fontsize=10)
        ax4.set_title('Utilidad Neta vs Gastos Operativos', fontsize=12, fontweight='bold')
        ax4.set_xticks(x)
        ax4.set_xticklabels(nombres_cortos, rotation=45, ha='right')
        ax4.legend()
        ax4.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
        ax4.grid(axis='y', alpha=0.3)
        
        plt.tight_layout()

        self.fig_resultados = fig  # Guardar para Excel

        self.canvas_graficos = FigureCanvasTkAgg(fig, self.frame_graficos_container)
        self.canvas_graficos.draw()
        self.canvas_graficos.get_tk_widget().pack(fill='both', expand=True)

    def _generar_factores_simulacion(self, n_satelites, variabilidad_sat,
                                      var_matriz_costos, pct_var_costos,
                                      var_matriz_ingresos, pct_var_ingresos,
                                      distribucion, rho):
        """Genera factores aleatorios correlacionados para una iteracion de simulacion.

        Retorna: (factor_matriz_costos, factor_matriz_ingresos, lista_factores_satelites)
        - Correlacion rho se aplica entre Z_matriz (costos) y Z de cada satelite.
        - Ingresos de matriz se generan independientemente.
        - Si var_matriz_costos=False, correlacion se ignora.
        """
        def _factor_desde_z(z, cv, dist):
            """Convierte un valor Z~N(0,1) en un factor multiplicativo."""
            if cv <= 0:
                return 1.0
            if dist == "Log-normal":
                sigma_ln = np.sqrt(np.log(1 + cv ** 2))
                mu_ln = -0.5 * sigma_ln ** 2
                return float(np.exp(mu_ln + sigma_ln * z))
            else:  # Normal
                return max(0.0, 1.0 + cv * z)

        cv_sat = variabilidad_sat / 3.0  # CV efectivo de satelites
        cv_costos = pct_var_costos / 3.0
        cv_ingresos = pct_var_ingresos / 3.0

        # Factor costos matriz
        Z_matriz = None
        if var_matriz_costos:
            Z_matriz = np.random.randn()
            factor_costos = _factor_desde_z(Z_matriz, cv_costos, distribucion)
        else:
            factor_costos = 1.0

        # Factor ingresos matriz (independiente)
        if var_matriz_ingresos:
            Z_ing = np.random.randn()
            factor_ingresos = _factor_desde_z(Z_ing, cv_ingresos, distribucion)
        else:
            factor_ingresos = 1.0

        # Factores por satelite (correlacionados con Z_matriz si aplica)
        factores_sat = []
        for _ in range(n_satelites):
            if var_matriz_costos and Z_matriz is not None and rho != 0:
                eps = np.random.randn()
                Z_i = rho * Z_matriz + np.sqrt(1 - rho ** 2) * eps
            else:
                Z_i = np.random.randn()
            factor_sat = _factor_desde_z(Z_i, cv_sat, distribucion)
            factores_sat.append(factor_sat)

        return factor_costos, factor_ingresos, factores_sat

    def ejecutar_simulacion(self):
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Advertencia", "Primero debe calcular los margenes optimos")
                return

            n_sim = int(self.entry_num_sim.get())
            variabilidad = float(self.entry_variabilidad.get()) / 100

            # Leer parametros de variabilidad de matriz y correlacion
            var_matriz_costos = self.VAR_MATRIZ_COSTOS.get()
            pct_var_costos = self.PCT_VAR_MATRIZ_COSTOS.get() / 100.0
            var_matriz_ingresos = self.VAR_MATRIZ_INGRESOS.get()
            pct_var_ingresos = self.PCT_VAR_MATRIZ_INGRESOS.get() / 100.0
            distribucion = self.DISTRIBUCION_FACTORES.get()
            rho = self.CORRELACION_MATRIZ_SAT.get()

            if rho < -1.0 or rho > 1.0:
                raise ValueError("La correlacion (rho) debe estar entre -1 y 1.")
            if pct_var_costos < 0 or pct_var_ingresos < 0:
                raise ValueError("Los porcentajes de variabilidad deben ser >= 0.")

            ahorros = []
            tasa_gral = self.resultados['parametros']['tasa_general'] / 100
            tasa_esp = self.resultados['parametros']['tasa_especial'] / 100
            limite_util = self.resultados['parametros']['limite_utilidad']

            # Valores base de la matriz
            ingresos_base = self.resultados['matriz']['ingresos']
            costos_ext_base = self.resultados['matriz']['costos_externos']
            util_sin_est_base = self.resultados['matriz']['utilidad_sin_estructura']
            imp_sin_est_base = self.resultados['matriz']['impuesto_sin_estructura']
            n_sats = len(self.resultados['satelites'])

            for _ in range(n_sim):
                # Generar factores correlacionados para esta iteracion
                f_costos, f_ingresos, factores_sat = self._generar_factores_simulacion(
                    n_sats, variabilidad, var_matriz_costos, pct_var_costos,
                    var_matriz_ingresos, pct_var_ingresos, distribucion, rho
                )

                # Aplicar variabilidad a la matriz
                ingresos_sim = ingresos_base * f_ingresos
                costos_ext_sim = costos_ext_base * f_costos
                util_sin_estructura_sim = ingresos_sim - costos_ext_sim
                imp_sin_estructura_sim = max(0, util_sin_estructura_sim) * tasa_gral

                satelites_sim = []
                for idx, sat in enumerate(self.resultados['satelites']):
                    factor = factores_sat[idx]
                    costo_sim = sat['costo'] * factor
                    gastos_sim = sat['gastos_operativos'] * factor

                    if sat['regimen'].startswith('GENERAL'):
                        if gastos_sim > 0:
                            margen = gastos_sim / costo_sim
                            utilidad_neta = 0
                            impuesto = 0
                        else:
                            margen = 0
                            utilidad_neta = 0
                            impuesto = 0
                    else:
                        margen = (limite_util + gastos_sim) / costo_sim
                        utilidad_bruta = costo_sim * margen
                        utilidad_neta = utilidad_bruta - gastos_sim

                        if utilidad_neta <= limite_util:
                            impuesto = utilidad_neta * tasa_esp
                        else:
                            impuesto = limite_util * tasa_esp + (utilidad_neta - limite_util) * tasa_gral

                    precio_venta = costo_sim * (1 + margen)

                    satelites_sim.append({
                        'utilidad': utilidad_neta,
                        'impuesto': impuesto,
                        'precio_venta': precio_venta,
                        'costo': costo_sim
                    })

                total_compras = sum(s['precio_venta'] for s in satelites_sim)
                total_costos_sat = sum(s['costo'] for s in satelites_sim)

                nueva_utilidad_matriz = util_sin_estructura_sim - total_compras + total_costos_sat
                impuesto_matriz = max(0, nueva_utilidad_matriz) * tasa_gral

                impuesto_total = impuesto_matriz + sum(s['impuesto'] for s in satelites_sim)
                ahorro = imp_sin_estructura_sim - impuesto_total

                ahorros.append(ahorro)

            ahorros = np.array(ahorros)

            # --- CALCULOS ESTADISTICOS AVANZADOS ---
            media = np.mean(ahorros)
            mediana = np.median(ahorros)
            std = np.std(ahorros, ddof=1)
            cv = (std / media * 100) if media != 0 else 0
            conf_val = self.CONF_NIVEL.get() / 100
            alfa_val = self.ALFA_SIG.get()

            # Bootstrap percentil para IC y p-valor no parametrico
            n_boot = 5000
            boot_medias = np.array([
                np.mean(np.random.choice(ahorros, size=len(ahorros), replace=True))
                for _ in range(n_boot)
            ])
            alpha_boot = 1 - conf_val
            ic_boot_lower = np.percentile(boot_medias, 100 * alpha_boot / 2)
            ic_boot_upper = np.percentile(boot_medias, 100 * (1 - alpha_boot / 2))

            # P-valor no parametrico: proporcion de bootstrap medias <= 0
            p_valor_boot = np.mean(boot_medias <= 0)
            p_valor_boot = max(p_valor_boot, 1 / n_boot)  # floor

            # T-test clasico (referencia)
            t_stat, p_value_t = stats.ttest_1samp(ahorros, 0)

            decision = "RECHAZAR H0 (Ahorro Significativo)" if p_valor_boot < alfa_val else "NO RECHAZAR H0 (Incierto)"

            # Ajuste Log-Normal
            ahorros_positivos = ahorros[ahorros > 0]
            ln_ajustado = False
            ln_shape = ln_loc = ln_scale = 0
            if len(ahorros_positivos) > 50:
                try:
                    ln_shape, ln_loc, ln_scale = stats.lognorm.fit(ahorros_positivos, floc=0)
                    ln_ajustado = True
                except Exception:
                    pass

            # --- VISUALIZACION MEJORADA: 2x2 ---
            if self.canvas_simulacion:
                self.canvas_simulacion.get_tk_widget().destroy()

            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
            fig.suptitle(f'Simulacion Monte Carlo ({n_sim:,} iter.) - Analisis de Riesgo',
                         fontsize=14, fontweight='bold')

            # [1] Histograma + curva Log-Normal ajustada
            ax1.hist(ahorros, bins=60, color='#3498db', alpha=0.6, edgecolor='black',
                     density=True, label='Datos simulados')
            if ln_ajustado:
                x_ln = np.linspace(max(0.01, ahorros.min()), ahorros.max(), 300)
                pdf_ln = stats.lognorm.pdf(x_ln, ln_shape, ln_loc, ln_scale)
                ax1.plot(x_ln, pdf_ln, 'r-', linewidth=2,
                         label=f'Log-Normal (s={ln_shape:.3f})')
            ax1.axvline(media, color='orange', linestyle='--', linewidth=2,
                        label=f'Media: {media:,.0f}')
            ax1.axvline(mediana, color='green', linestyle=':', linewidth=2,
                        label=f'Mediana: {mediana:,.0f}')
            ax1.fill_betweenx([0, ax1.get_ylim()[1] if ax1.get_ylim()[1] > 0 else 0.001],
                              ic_boot_lower, ic_boot_upper, alpha=0.15, color='red',
                              label=f'IC {conf_val*100:.0f}% Boot')
            ax1.set_xlabel('Ahorro Tributario (S/)', fontsize=10)
            ax1.set_ylabel('Densidad', fontsize=10)
            ax1.set_title('Histograma + Ajuste Log-Normal', fontsize=11, fontweight='bold')
            ax1.legend(fontsize=7, loc='upper right')
            ax1.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
            ax1.grid(alpha=0.3)

            # [2] Q-Q Plot (Log-Normal)
            if ln_ajustado and len(ahorros_positivos) > 10:
                ahorros_sorted = np.sort(ahorros_positivos)
                n_qq = len(ahorros_sorted)
                teoricos = stats.lognorm.ppf(
                    (np.arange(1, n_qq + 1) - 0.5) / n_qq,
                    ln_shape, ln_loc, ln_scale
                )
                ax2.scatter(teoricos, ahorros_sorted, s=4, alpha=0.4, color='#2E75B6')
                lim_min = min(teoricos.min(), ahorros_sorted.min())
                lim_max = max(teoricos.max(), ahorros_sorted.max())
                ax2.plot([lim_min, lim_max], [lim_min, lim_max], 'r--', linewidth=1.5,
                         label='Linea 45')
                ax2.set_xlabel('Cuantiles Teoricos (Log-Normal)', fontsize=10)
                ax2.set_ylabel('Cuantiles Observados', fontsize=10)
                ax2.set_title('Q-Q Plot Log-Normal', fontsize=11, fontweight='bold')
                ax2.legend(fontsize=8)
            else:
                sorted_ahorros = np.sort(ahorros)
                cumulative = np.arange(1, len(sorted_ahorros) + 1) / len(sorted_ahorros)
                ax2.plot(sorted_ahorros, cumulative * 100, color='#9b59b6', linewidth=2)
                ax2.axhline(50, color='red', linestyle='--', alpha=0.5, label='Mediana')
                ax2.axhline(95, color='green', linestyle='--', alpha=0.5, label='P95')
                ax2.set_xlabel('Ahorro Tributario', fontsize=10)
                ax2.set_ylabel('Prob. Acumulada (%)', fontsize=10)
                ax2.set_title('CDF (sin ajuste Log-Normal)', fontsize=11, fontweight='bold')
                ax2.legend(fontsize=8)
            ax2.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
            ax2.grid(alpha=0.3)

            # [3] Violin + Boxplot
            parts = ax3.violinplot(ahorros, positions=[1], showmeans=True,
                                   showmedians=True, showextrema=False)
            for pc in parts['bodies']:
                pc.set_facecolor('#3498db')
                pc.set_alpha(0.4)
            parts['cmeans'].set_color('red')
            parts['cmedians'].set_color('green')
            bp = ax3.boxplot(ahorros, positions=[1], widths=0.15, patch_artist=True,
                             showfliers=False)
            bp['boxes'][0].set_facecolor('#2ecc71')
            bp['boxes'][0].set_alpha(0.6)
            ax3.set_ylabel('Ahorro Tributario (S/)', fontsize=10)
            ax3.set_title('Violin + Box Plot', fontsize=11, fontweight='bold')
            ax3.set_xticks([1])
            ax3.set_xticklabels(['Ahorro'])
            ax3.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
            ax3.grid(axis='y', alpha=0.3)

            # [4] Panel de estadisticos
            mat_info = ""
            if var_matriz_costos:
                mat_info += f"\nVar.Costos Mat: +/-{pct_var_costos*100:.1f}%"
            if var_matriz_ingresos:
                mat_info += f"\nVar.Ing. Mat:   +/-{pct_var_ingresos*100:.1f}%"
            if var_matriz_costos and rho != 0:
                mat_info += f"\nCorrelacion:    rho={rho:.2f}"
            if mat_info:
                mat_info = f"\nMATRIZ:{mat_info}\nDistribucion:   {distribucion}"

            stats_text = f"""ESTADISTICAS SIMULACION
{'='*34}
Iteraciones:  {n_sim:,}
Variabilidad: +/-{variabilidad*100:.1f}%{mat_info}

AHORRO TRIBUTARIO:
  Media:      {media:>12,.0f}
  Mediana:    {mediana:>12,.0f}
  Desv.Est.:  {std:>12,.0f}
  Coef.Var.:  {cv:>12,.2f}%

PERCENTILES:
  P5:         {np.percentile(ahorros, 5):>12,.0f}
  P25:        {np.percentile(ahorros, 25):>12,.0f}
  P75:        {np.percentile(ahorros, 75):>12,.0f}
  P95:        {np.percentile(ahorros, 95):>12,.0f}

BOOTSTRAP IC {conf_val*100:.0f}%:
  [{ic_boot_lower:>12,.0f},
   {ic_boot_upper:>12,.0f}]

PRUEBA HIPOTESIS (H0: ahorro=0):
  P-valor Boot:  {p_valor_boot:.4e}
  t-stat ref:    {t_stat:.4f}
  P-valor t:     {p_value_t:.4e}
  Decision:      {decision}"""
            if ln_ajustado:
                stats_text += f"""

LOG-NORMAL FIT:
  Shape(s):  {ln_shape:.4f}
  Scale:     {ln_scale:>12,.0f}"""

            ax4.axis('off')
            ax4.text(0.05, 0.95, stats_text, transform=ax4.transAxes,
                     fontsize=7.5, verticalalignment='top', fontfamily='monospace',
                     bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

            plt.tight_layout()

            self.fig_simulacion = fig
            self.datos_simulacion = {
                'n_sim': n_sim,
                'variabilidad': variabilidad * 100,
                'media': media,
                'mediana': mediana,
                'std': std,
                'cv': cv,
                'min': float(np.min(ahorros)),
                'max': float(np.max(ahorros)),
                'p5': float(np.percentile(ahorros, 5)),
                'p25': float(np.percentile(ahorros, 25)),
                'p75': float(np.percentile(ahorros, 75)),
                'p95': float(np.percentile(ahorros, 95)),
                'ic_lower': ic_boot_lower,
                'ic_upper': ic_boot_upper,
                'ic_metodo': 'Bootstrap Percentil',
                't_stat': t_stat,
                'p_value': p_valor_boot,
                'p_value_t': p_value_t,
                'decision': decision,
                'conf_nivel': conf_val * 100,
                'alfa': alfa_val,
                'n_bootstrap': n_boot,
                'ln_ajustado': ln_ajustado,
                'ln_shape': ln_shape if ln_ajustado else None,
                'ln_loc': ln_loc if ln_ajustado else None,
                'ln_scale': ln_scale if ln_ajustado else None,
                # Parametros de variabilidad matriz y correlacion
                'var_matriz_costos': var_matriz_costos,
                'pct_var_costos': pct_var_costos * 100,
                'var_matriz_ingresos': var_matriz_ingresos,
                'pct_var_ingresos': pct_var_ingresos * 100,
                'distribucion': distribucion,
                'rho': rho,
            }

            self.canvas_simulacion = FigureCanvasTkAgg(fig, self.frame_sim_container)
            self.canvas_simulacion.draw()
            self.canvas_simulacion.get_tk_widget().pack(fill='both', expand=True)

            messagebox.showinfo("Simulacion Completa",
                                f"Ahorro esperado: {media:,.0f}\n"
                                f"IC {conf_val*100:.0f}% Bootstrap: [{ic_boot_lower:,.0f}, {ic_boot_upper:,.0f}]\n"
                                f"P-valor (bootstrap): {p_valor_boot:.4e}\n"
                                f"Decision: {decision}")

        except Exception as e:
            messagebox.showerror("Error", f"Error en simulacion: {str(e)}")

    def _calcular_ahorro_grupo(self, satelites_info, tasa_general, tasa_especial,
                               limite_utilidad, limite_ingresos, utilidad_sin_estructura):
        """Recalcula el ahorro tributario total del grupo con par√°metros dados.

        Replica la l√≥gica exacta de calcular_margenes_optimos() para garantizar
        consistencia entre el c√°lculo base y el an√°lisis de sensibilidad.
        """
        total_impuesto_satelites = 0
        total_compras_matriz = 0
        total_costos_satelites = 0

        for sat in satelites_info:
            costo = sat['costo']
            gastos = sat['gastos']
            es_general = sat['es_general']

            if es_general:
                if gastos > 0:
                    margen = gastos / costo
                    impuesto = 0
                else:
                    margen = 0
                    impuesto = 0
            else:
                margen = (limite_utilidad + gastos) / costo
                precio_tentativo = costo * (1 + margen)

                if precio_tentativo > limite_ingresos:
                    margen = (limite_ingresos / costo) - 1

                if margen > 1.0:
                    # No califica, forzar general
                    if gastos > 0:
                        margen = gastos / costo
                    else:
                        margen = 0
                    impuesto = 0
                else:
                    utilidad_bruta = costo * margen
                    utilidad_neta = utilidad_bruta - gastos

                    if utilidad_neta <= limite_utilidad:
                        impuesto = utilidad_neta * tasa_especial
                    else:
                        impuesto = (limite_utilidad * tasa_especial +
                                    (utilidad_neta - limite_utilidad) * tasa_general)

            precio_venta = costo * (1 + margen)
            total_impuesto_satelites += impuesto
            total_compras_matriz += precio_venta
            total_costos_satelites += costo

        nueva_utilidad_matriz = utilidad_sin_estructura - total_compras_matriz + total_costos_satelites
        impuesto_matriz = nueva_utilidad_matriz * tasa_general

        impuesto_total_grupo = impuesto_matriz + total_impuesto_satelites
        impuesto_sin_estructura = utilidad_sin_estructura * tasa_general

        return impuesto_sin_estructura - impuesto_total_grupo

    def ejecutar_sensibilidad(self):
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Advertencia", "Primero debe calcular los m√°rgenes √≥ptimos")
                return
            
            variable = self.combo_variable.get()
            rango = float(self.entry_rango_sens.get()) / 100
            
            if self.canvas_sensibilidad:
                self.canvas_sensibilidad.get_tk_widget().destroy()
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
            fig.suptitle(f'An√°lisis de Sensibilidad - {variable}', fontsize=16, fontweight='bold')
            
            uit = self.resultados['parametros']['uit']
            tasa_gral = self.resultados['parametros']['tasa_general'] / 100
            tasa_esp = self.resultados['parametros']['tasa_especial'] / 100
            limite_util = self.resultados['parametros']['limite_utilidad']
            limite_ing = self.resultados['parametros']['limite_ingresos']
            util_sin_estructura = self.resultados['matriz']['utilidad_sin_estructura']

            # Datos base de sat√©lites para el helper
            sats_base = [
                {
                    'costo': s['costo'],
                    'gastos': s['gastos_operativos'],
                    'es_general': s['regimen'].startswith('GENERAL')
                }
                for s in self.resultados['satelites']
            ]
            ahorro_base = self.resultados['grupo']['ahorro_tributario']

            if variable == "Costos Sat√©lites":
                variaciones = np.linspace(-rango, rango, 50)
                ahorros_totales = []

                for var in variaciones:
                    sats_var = [
                        {
                            'costo': s['costo'] * (1 + var),
                            'gastos': s['gastos'] * (1 + var),
                            'es_general': s['es_general']
                        }
                        for s in sats_base
                    ]
                    ahorro = self._calcular_ahorro_grupo(
                        sats_var, tasa_gral, tasa_esp,
                        limite_util, limite_ing, util_sin_estructura
                    )
                    ahorros_totales.append(ahorro)

                ax1.plot(variaciones * 100, ahorros_totales, color='#e74c3c', linewidth=2.5, marker='o', markersize=3)
                ax1.axvline(0, color='gray', linestyle='--', alpha=0.5)
                ax1.axhline(ahorro_base, color='green', linestyle='--',
                           alpha=0.5, label=f'Ahorro Base ({ahorro_base:,.0f})')
                ax1.set_xlabel('Variaci√≥n en Costos (%)', fontsize=11)
                ax1.set_ylabel('Ahorro Tributario', fontsize=11)
                ax1.set_title('Impacto en Ahorro', fontsize=13, fontweight='bold')
                ax1.grid(True, alpha=0.3)
                ax1.legend()
                ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.2f}M' if x >= 1e6 else f'{x/1000:.0f}K'))

                # Elasticidad (variaci√≥n del ahorro por punto porcentual de cambio en costos)
                ahorros_arr = np.array(ahorros_totales)
                elasticidad = np.gradient(ahorros_arr, variaciones * 100)
                ax2.plot(variaciones * 100, elasticidad, color='#3498db', linewidth=2.5)
                ax2.axvline(0, color='gray', linestyle='--', alpha=0.5)
                ax2.axhline(0, color='black', linestyle='-', alpha=0.3)
                ax2.set_xlabel('Variaci√≥n en Costos (%)', fontsize=11)
                ax2.set_ylabel('dAhorro/dCosto (S/ por pp)', fontsize=11)
                ax2.set_title('Sensibilidad del Ahorro', fontsize=13, fontweight='bold')
                ax2.grid(True, alpha=0.3)
                ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.1f}K'))

            elif variable == "Tasa R√©gimen Especial":
                tasa_base = self.resultados['parametros']['tasa_especial']
                rango_abs = tasa_base * rango
                tasas_especial = np.linspace(max(1, tasa_base - rango_abs), tasa_base + rango_abs, 50)
                ahorros_totales = []

                for tasa in tasas_especial:
                    ahorro = self._calcular_ahorro_grupo(
                        sats_base, tasa_gral, tasa / 100,
                        limite_util, limite_ing, util_sin_estructura
                    )
                    ahorros_totales.append(ahorro)

                ax1.plot(tasas_especial, ahorros_totales, color='#9b59b6', linewidth=2.5, marker='o', markersize=3)
                ax1.axvline(tasa_base, color='red', linestyle='--', alpha=0.5, label=f'Tasa Actual ({tasa_base:.1f}%)')
                ax1.axhline(ahorro_base, color='green', linestyle='--',
                           alpha=0.5, label=f'Ahorro Base ({ahorro_base:,.0f})')
                ax1.set_xlabel('Tasa R√©gimen Especial (%)', fontsize=11)
                ax1.set_ylabel('Ahorro Tributario', fontsize=11)
                ax1.set_title('Sensibilidad a Tasa Especial', fontsize=13, fontweight='bold')
                ax1.grid(True, alpha=0.3)
                ax1.legend()
                ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.2f}M' if x >= 1e6 else f'{x/1000:.0f}K'))

                # Elasticidad: dAhorro/dTasa
                ahorros_arr = np.array(ahorros_totales)
                elasticidad = np.gradient(ahorros_arr, tasas_especial)
                ax2.plot(tasas_especial, elasticidad, color='#e67e22', linewidth=2.5)
                ax2.axvline(tasa_base, color='red', linestyle='--', alpha=0.5)
                ax2.axhline(0, color='black', linestyle='-', alpha=0.3)
                ax2.set_xlabel('Tasa R√©gimen Especial (%)', fontsize=11)
                ax2.set_ylabel('dAhorro/dTasa (S/ por pp)', fontsize=11)
                ax2.set_title('Elasticidad a Tasa Especial', fontsize=13, fontweight='bold')
                ax2.grid(True, alpha=0.3)
                ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))

            else:  # L√≠mite Utilidad
                limite_base = self.resultados['parametros']['limite_utilidad'] / uit
                rango_abs = limite_base * rango
                limites_uit = np.linspace(max(1, limite_base - rango_abs), limite_base + rango_abs, 50)
                ahorros_totales = []

                for limite_uit_val in limites_uit:
                    ahorro = self._calcular_ahorro_grupo(
                        sats_base, tasa_gral, tasa_esp,
                        limite_uit_val * uit, limite_ing, util_sin_estructura
                    )
                    ahorros_totales.append(ahorro)

                ax1.plot(limites_uit, ahorros_totales, color='#1abc9c', linewidth=2.5, marker='o', markersize=3)
                ax1.axvline(limite_base, color='red', linestyle='--', alpha=0.5, label=f'L√≠mite Actual ({limite_base:.1f} UIT)')
                ax1.axhline(ahorro_base, color='green', linestyle='--',
                           alpha=0.5, label=f'Ahorro Base ({ahorro_base:,.0f})')
                ax1.set_xlabel('L√≠mite Utilidad (UIT)', fontsize=11)
                ax1.set_ylabel('Ahorro Tributario', fontsize=11)
                ax1.set_title('Sensibilidad a L√≠mite', fontsize=13, fontweight='bold')
                ax1.grid(True, alpha=0.3)
                ax1.legend()
                ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.2f}M' if x >= 1e6 else f'{x/1000:.0f}K'))

                # Ahorro marginal por UIT adicional
                ahorros_arr = np.array(ahorros_totales)
                ahorro_marginal = np.gradient(ahorros_arr, limites_uit)
                ax2.plot(limites_uit, ahorro_marginal, color='#f39c12', linewidth=2.5)
                ax2.axvline(limite_base, color='red', linestyle='--', alpha=0.5)
                ax2.axhline(0, color='black', linestyle='-', alpha=0.3)
                ax2.set_xlabel('L√≠mite Utilidad (UIT)', fontsize=11)
                ax2.set_ylabel('Ahorro Marginal por UIT', fontsize=11)
                ax2.set_title('Beneficio Marginal', fontsize=13, fontweight='bold')
                ax2.grid(True, alpha=0.3)
                ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
            
            plt.tight_layout()

            self.fig_sensibilidad = fig  # Guardar para Excel

            self.canvas_sensibilidad = FigureCanvasTkAgg(fig, self.frame_sens_container)
            self.canvas_sensibilidad.draw()
            self.canvas_sensibilidad.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            messagebox.showerror("Error", f"Error en sensibilidad: {str(e)}")

    def generar_comparativo(self):
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Advertencia", "Primero debe calcular los m√°rgenes √≥ptimos")
                return
            
            if self.canvas_comparativo:
                self.canvas_comparativo.get_tk_widget().destroy()
            
            # Generar escenarios
            escenarios = []
            
            # Escenario actual
            escenarios.append({
                'nombre': 'Actual',
                'ahorro': self.resultados['grupo']['ahorro_tributario'],
                'tasa_efectiva': self.resultados['grupo']['impuesto_total'] / 
                               (self.resultados['matriz']['nueva_utilidad'] + self.resultados['grupo']['total_utilidad_satelites']) * 100,
                'satelites': len(self.resultados['satelites'])
            })
            
            # Escenario: L√≠mite +20%
            uit = self.resultados['parametros']['uit']
            limite_util_aumentado = self.resultados['parametros']['limite_utilidad'] * 1.2
            num_especial = sum(1 for s in self.resultados['satelites'] if not s['regimen'].startswith('GENERAL'))
            ahorro_esc2 = num_especial * limite_util_aumentado * \
                         (self.resultados['parametros']['tasa_general'] - self.resultados['parametros']['tasa_especial']) / 100
            
            escenarios.append({
                'nombre': 'L√≠mite +20%',
                'ahorro': ahorro_esc2,
                'tasa_efectiva': escenarios[0]['tasa_efectiva'] - 1.2,
                'satelites': len(self.resultados['satelites'])
            })
            
            # Escenario: Tasa especial -2pp
            diferencial = (self.resultados['parametros']['tasa_general'] - (self.resultados['parametros']['tasa_especial'] - 2))
            ahorro_esc3 = sum([s['utilidad_neta'] for s in self.resultados['satelites'] 
                              if not s['regimen'].startswith('GENERAL')]) * diferencial / 100
            
            escenarios.append({
                'nombre': 'Tasa Esp. -2pp',
                'ahorro': ahorro_esc3,
                'tasa_efectiva': escenarios[0]['tasa_efectiva'] - 0.8,
                'satelites': len(self.resultados['satelites'])
            })
            
            # Escenario: Duplicar sat√©lites especiales
            ahorro_esc4 = self.resultados['grupo']['ahorro_tributario'] * 1.5
            
            escenarios.append({
                'nombre': 'M√°s Sat√©lites',
                'ahorro': ahorro_esc4,
                'tasa_efectiva': escenarios[0]['tasa_efectiva'] - 2.5,
                'satelites': len(self.resultados['satelites']) * 2
            })
            
            # Crear gr√°ficos
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
            fig.suptitle('An√°lisis Comparativo Multi-Escenario', fontsize=16, fontweight='bold')
            
            nombres = [e['nombre'] for e in escenarios]
            ahorros = [e['ahorro'] for e in escenarios]
            tasas = [e['tasa_efectiva'] for e in escenarios]
            
            colores = ['#3498db', '#2ecc71', '#f39c12', '#9b59b6']
            
            # Gr√°fico 1: Ahorros
            bars = ax1.bar(nombres, ahorros, color=colores, alpha=0.7, edgecolor='black', linewidth=2)
            ax1.set_ylabel('Ahorro Tributario Anual', fontsize=11)
            ax1.set_title('Comparaci√≥n de Ahorro', fontsize=13, fontweight='bold')
            ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
            ax1.grid(axis='y', alpha=0.3)
            
            for i, bar in enumerate(bars):
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:,.0f}',
                        ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            # Gr√°fico 2: Tasa efectiva
            ax2.plot(nombres, tasas, marker='o', color='#e74c3c', linewidth=2.5, markersize=10)
            ax2.set_ylabel('Tasa Efectiva Grupo (%)', fontsize=11)
            ax2.set_title('Tasa Efectiva por Escenario', fontsize=13, fontweight='bold')
            ax2.grid(True, alpha=0.3)
            
            for i, (x, y) in enumerate(zip(nombres, tasas)):
                ax2.text(i, y + 0.3, f'{y:.2f}%', ha='center', fontsize=9, fontweight='bold')
            
            plt.tight_layout()

            self.fig_comparativo = fig  # Guardar para Excel
            self.datos_comparativo = escenarios  # Guardar datos

            self.canvas_comparativo = FigureCanvasTkAgg(fig, self.frame_comp_container)
            self.canvas_comparativo.draw()
            self.canvas_comparativo.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            messagebox.showerror("Error", f"Error en comparativo: {str(e)}")

        
    def _calcular_margen_equilibrio_exacto(self, costo, gastos, tasa_pc, tasa_especial,
                                             tasa_general, limite_utilidad, limite_ingresos, es_general):
        """Calcula el margen de equilibrio exacto donde IR = PaC (saldo = 0).

        Prueba 3 escenarios en orden:
        A) Regimen Especial Puro: Util * t_esp = Ingreso * t_pc
        B) Regimen MYPE Mixto: L*t_esp + (Util-L)*t_gral = Ingreso * t_pc
        C) Regimen General Puro: Util * t_gral = Ingreso * t_pc

        Returns: (margen, regimen_str) o (None, None) si no hay solucion factible.
        """
        if costo <= 0:
            return None, None

        # --- Escenario C: Regimen General Puro ---
        # Util * t_gral = Ingreso * t_pc
        # (C*m - G) * t_gral = C*(1+m) * t_pc
        # m = (C*t_pc + G*t_gral) / (C*(t_gral - t_pc))
        m_eq_gral = None
        if (tasa_general - tasa_pc) > 0:
            m_eq_gral = (tasa_pc * costo + tasa_general * gastos) / (costo * (tasa_general - tasa_pc))

        # Si esta forzado a General, solo Escenario C
        if es_general:
            if m_eq_gral is not None and m_eq_gral > 0:
                util_c = costo * m_eq_gral - gastos
                if util_c >= 0:
                    return m_eq_gral, "General"
            return None, None

        # --- Escenario A: Regimen Especial Puro ---
        # Util * t_esp = Ingreso * t_pc
        # (C*m - G) * t_esp = C*(1+m) * t_pc
        # m = (C*t_pc + G*t_esp) / (C*(t_esp - t_pc))
        if (tasa_especial - tasa_pc) > 0:
            m_eq_esp = (tasa_pc * costo + tasa_especial * gastos) / (costo * (tasa_especial - tasa_pc))
            if m_eq_esp > 0:
                util_a = costo * m_eq_esp - gastos
                precio_a = costo * (1 + m_eq_esp)
                if util_a >= 0 and util_a <= limite_utilidad and precio_a <= limite_ingresos:
                    return m_eq_esp, "Especial"

        # --- Escenario B: Regimen MYPE Mixto (LA CORRECCION CLAVE) ---
        # L*t_esp + (Util - L)*t_gral = Ingreso * t_pc
        # L*t_esp + (C*m - G - L)*t_gral = C*(1+m)*t_pc
        # C*m*t_gral - C*m*t_pc = C*t_pc + G*t_gral + L*t_gral - L*t_esp
        # m = [C*t_pc + G*t_gral + L*(t_gral - t_esp)] / [C*(t_gral - t_pc)]
        if (tasa_general - tasa_pc) > 0:
            m_eq_mixto = (tasa_pc * costo + tasa_general * gastos + limite_utilidad * (tasa_general - tasa_especial)) / (costo * (tasa_general - tasa_pc))
            if m_eq_mixto > 0:
                util_b = costo * m_eq_mixto - gastos
                precio_b = costo * (1 + m_eq_mixto)
                # Valido solo si utilidad > limite (si no, seria Escenario A)
                if util_b > limite_utilidad and precio_b <= limite_ingresos:
                    return m_eq_mixto, "Mixto"

        # --- Fallback a Escenario C si nada mas funciona ---
        if m_eq_gral is not None and m_eq_gral > 0:
            util_c = costo * m_eq_gral - gastos
            if util_c >= 0:
                return m_eq_gral, "General"

        return None, None

    def _calcular_ir_equilibrio(self, utilidad_neta, tasa_especial, tasa_general,
                                 limite_utilidad, regimen_equilibrio):
        """Calcula el IR en el punto de equilibrio segun el regimen determinado."""
        if regimen_equilibrio == "Especial":
            return utilidad_neta * tasa_especial
        elif regimen_equilibrio == "Mixto":
            return limite_utilidad * tasa_especial + (utilidad_neta - limite_utilidad) * tasa_general
        else:  # General
            return utilidad_neta * tasa_general

    
    def simular_equilibrio_financiero(self):
        """
        SIMULACI√ìN MONTE CARLO ‚Äì DISTRIBUCI√ìN DEL SALDO DE REGULARIZACI√ìN (PAGO MARZO).
        Compara el escenario √ìPTIMO (m√°ximo ahorro) vs EQUILIBRIO (saldo ~ 0).
        Incluye:
          - Bootstrap percentil para IC y p‚Äëvalor.
          - Gr√°ficos: densidad del grupo + violines por empresa.
          - Reporte de percentiles y reducci√≥n de riesgo.
        """
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Advertencia", "Primero debe calcular los m√°rgenes √≥ptimos")
                return
    
            # --------------------------------------------------------------
            # 1. CONFIGURACI√ìN DE LA SIMULACI√ìN
            # --------------------------------------------------------------
            n_sim = int(self.entry_n_sim_eq.get())
            variabilidad = float(self.entry_variabilidad.get()) / 100 if hasattr(self, 'entry_variabilidad') else 0.10
    
            # Par√°metros de incertidumbre de la matriz y correlaci√≥n
            var_matriz_costos = self.VAR_MATRIZ_COSTOS.get()
            pct_var_costos = self.PCT_VAR_MATRIZ_COSTOS.get() / 100.0
            var_matriz_ingresos = self.VAR_MATRIZ_INGRESOS.get()
            pct_var_ingresos = self.PCT_VAR_MATRIZ_INGRESOS.get() / 100.0
            distribucion = self.DISTRIBUCION_FACTORES.get()
            rho = self.CORRELACION_MATRIZ_SAT.get()
            if rho < -1.0 or rho > 1.0:
                raise ValueError("La correlaci√≥n (rho) debe estar entre -1 y 1.")
    
            # Par√°metros tributarios
            tasa_gral = self.resultados['parametros']['tasa_general'] / 100
            tasa_esp = self.resultados['parametros']['tasa_especial'] / 100
            tasa_pc = self.TASA_PAGO_CUENTA.get() / 100
            limite_util = self.resultados['parametros']['limite_utilidad']
            limite_ing = self.resultados['parametros']['limite_ingresos']
    
            # Datos base de la matriz
            ingresos_matriz_base = self.resultados['matriz']['ingresos']
            costos_matriz_base = self.resultados['matriz']['costos_externos']
            n_sats = len(self.resultados['satelites'])
    
            # --------------------------------------------------------------
            # 2. ESTRUCTURA PARA ALMACENAR HISTORIAL DE SALDOS
            # --------------------------------------------------------------
            # Cada clave guarda dos listas: 'opt' (escenario √≥ptimo) y 'eq' (equilibrio)
            historial_saldos = {
                'Matriz': {'opt': [], 'eq': []},
                'Grupo_Total': {'opt': [], 'eq': []}
            }
            for sat in self.resultados['satelites']:
                historial_saldos[sat['nombre']] = {'opt': [], 'eq': []}
    
            # --------------------------------------------------------------
            # 3. BUCLE PRINCIPAL DE MONTE CARLO
            # --------------------------------------------------------------
            for _ in range(n_sim):
                # A) Factores aleatorios correlacionados
                f_costos_m, f_ingresos_m, factores_sat = self._generar_factores_simulacion(
                    n_sats, variabilidad,
                    var_matriz_costos, pct_var_costos,
                    var_matriz_ingresos, pct_var_ingresos,
                    distribucion, rho
                )
    
                # B) Simular la matriz base
                ingresos_sim = ingresos_matriz_base * f_ingresos_m
                costos_m_sim = costos_matriz_base * f_costos_m
                util_base_sim = max(0, ingresos_sim - costos_m_sim)  # utilidad antes de compras a sat√©lites
                pac_matriz = tasa_pc * ingresos_sim   # pago a cuenta fijo (depende de ingresos)
    
                # Acumuladores para el grupo
                saldo_grupo_opt = 0.0
                saldo_grupo_eq = 0.0
                compras_opt_total = 0.0
                compras_eq_total = 0.0
                costos_sats_total = 0.0
    
                # C) Simular cada sat√©lite
                for idx, sat in enumerate(self.resultados['satelites']):
                    factor = factores_sat[idx]
                    costo_sim = sat['costo'] * factor
                    gastos_sim = sat['gastos_operativos'] * factor
                    es_gral = sat['regimen'].startswith('GENERAL')
                    costos_sats_total += costo_sim
    
                    # ========== ESCENARIO 1: MARGEN √ìPTIMO (maximiza ahorro) ==========
                    if costo_sim > 0:
                        if es_gral:
                            # R√âGIMEN GENERAL: margen de equilibrio operativo (utilidad neta cero)
                            m_opt = (gastos_sim / costo_sim) if gastos_sim > 0 else 0.0
                        else:
                            # R√âGIMEN ESPECIAL: f√≥rmula est√°ndar (maximiza utilidad hasta el l√≠mite)
                            m_opt = (limite_util + gastos_sim) / costo_sim
                            # Ajuste por l√≠mite de ingresos
                            if costo_sim * (1 + m_opt) > limite_ing:
                                m_opt = (limite_ing / costo_sim) - 1
                    else:
                        m_opt = 0.0
    
                    precio_opt = costo_sim * (1 + m_opt)
                    util_opt = max(0, precio_opt - costo_sim - gastos_sim)
    
                    # Impuesto seg√∫n r√©gimen (√≥ptimo)
                    if es_gral:
                        ir_opt = util_opt * tasa_gral
                    elif util_opt <= limite_util:
                        ir_opt = util_opt * tasa_esp
                    else:
                        ir_opt = limite_util * tasa_esp + (util_opt - limite_util) * tasa_gral
    
                    pac_opt = tasa_pc * precio_opt
                    saldo_opt = ir_opt - pac_opt
    
                    compras_opt_total += precio_opt
                    saldo_grupo_opt += saldo_opt
                    historial_saldos[sat['nombre']]['opt'].append(saldo_opt)
    
                    # ========== ESCENARIO 2: MARGEN DE EQUILIBRIO (IR = PaC) ==========
                    m_eq, reg_eq = self._calcular_margen_equilibrio_exacto(
                        costo_sim, gastos_sim, tasa_pc, tasa_esp, tasa_gral,
                        limite_util, limite_ing, es_gral
                    )
                    if m_eq is None:
                        m_eq = m_opt   # fallback al margen √≥ptimo si no hay equilibrio factible
    
                    precio_eq = costo_sim * (1 + m_eq)
                    util_eq = max(0, precio_eq - costo_sim - gastos_sim)
                    ir_eq = self._calcular_ir_equilibrio(
                        util_eq, tasa_esp, tasa_gral, limite_util, reg_eq or "General"
                    )
                    pac_eq = tasa_pc * precio_eq
                    saldo_eq = ir_eq - pac_eq
    
                    compras_eq_total += precio_eq
                    saldo_grupo_eq += saldo_eq
                    historial_saldos[sat['nombre']]['eq'].append(saldo_eq)
    
                # D) Calcular la matriz final para ambos escenarios
                # ---- Escenario √ìptimo ----
                util_matriz_opt = max(0, util_base_sim - compras_opt_total + costos_sats_total)
                imp_matriz_opt = util_matriz_opt * tasa_gral
                saldo_matriz_opt = imp_matriz_opt - pac_matriz
    
                # ---- Escenario Equilibrio ----
                util_matriz_eq = max(0, util_base_sim - compras_eq_total + costos_sats_total)
                imp_matriz_eq = util_matriz_eq * tasa_gral
                saldo_matriz_eq = imp_matriz_eq - pac_matriz
    
                # Guardar historiales de matriz y grupo
                historial_saldos['Matriz']['opt'].append(saldo_matriz_opt)
                historial_saldos['Matriz']['eq'].append(saldo_matriz_eq)
    
                historial_saldos['Grupo_Total']['opt'].append(saldo_matriz_opt + saldo_grupo_opt)
                historial_saldos['Grupo_Total']['eq'].append(saldo_matriz_eq + saldo_grupo_eq)
    
            # --------------------------------------------------------------
            # 4. ESTAD√çSTICAS AVANZADAS (BOOTSTRAP)
            # --------------------------------------------------------------
            saldos_g_opt = np.array(historial_saldos['Grupo_Total']['opt'])
            saldos_g_eq  = np.array(historial_saldos['Grupo_Total']['eq'])
    
            # Bootstrap percentil para la media del saldo en equilibrio
            conf_val = self.CONF_NIVEL.get() / 100
            n_boot = 5000
            rng = np.random.default_rng(42)
    
            boot_eq = np.array([
                np.mean(rng.choice(saldos_g_eq, size=len(saldos_g_eq), replace=True))
                for _ in range(n_boot)
            ])
            alpha = 1 - conf_val
            ic_eq_low = np.percentile(boot_eq, 100 * alpha / 2)
            ic_eq_high = np.percentile(boot_eq, 100 * (1 - alpha / 2))
    
            # Bootstrap para la diferencia de medias (√≥ptimo - equilibrio)
            boot_diff = np.array([
                np.mean(rng.choice(saldos_g_opt, size=len(saldos_g_opt), replace=True)) -
                np.mean(rng.choice(saldos_g_eq,   size=len(saldos_g_eq),   replace=True))
                for _ in range(n_boot)
            ])
            p_valor_diff = np.mean(boot_diff <= 0)   # unilateral: H1: media_opt > media_eq
            p_valor_diff = max(p_valor_diff, 1 / n_boot)
    
            # Estad√≠sticas descriptivas
            stats_eq = {
                'media': np.mean(saldos_g_eq),
                'mediana': np.median(saldos_g_eq),
                'std': np.std(saldos_g_eq),
                'p5': np.percentile(saldos_g_eq, 5),
                'p25': np.percentile(saldos_g_eq, 25),
                'p75': np.percentile(saldos_g_eq, 75),
                'p95': np.percentile(saldos_g_eq, 95),
                'ic_low': ic_eq_low,
                'ic_high': ic_eq_high
            }
            stats_opt = {
                'media': np.mean(saldos_g_opt),
                'mediana': np.median(saldos_g_opt),
                'p95': np.percentile(saldos_g_opt, 95)
            }
    
            # --------------------------------------------------------------
            # 5. VISUALIZACI√ìN AVANZADA (DENSIDAD + VIOLINES)
            # --------------------------------------------------------------
            if self.canvas_equilibrio_sim:
                self.canvas_equilibrio_sim.get_tk_widget().destroy()
    
            fig = plt.figure(figsize=(14, 10))
            gs = fig.add_gridspec(2, 1, height_ratios=[1, 2])
    
            # ----- PANEL SUPERIOR: DENSIDAD DEL GRUPO CONSOLIDADO -----
            ax1 = fig.add_subplot(gs[0])
    
            # Estimaci√≥n de densidad kernel (KDE) para suavizar
            from scipy.stats import gaussian_kde
            kde_opt = gaussian_kde(saldos_g_opt)
            kde_eq  = gaussian_kde(saldos_g_eq)
    
            x_min = min(saldos_g_opt.min(), saldos_g_eq.min())
            x_max = max(saldos_g_opt.max(), saldos_g_eq.max())
            x = np.linspace(x_min, x_max, 300)
    
            ax1.fill_between(x, kde_opt(x), alpha=0.4, color='#E74C3C', label='√ìptimo (riesgo)')
            ax1.fill_between(x, kde_eq(x),  alpha=0.5, color='#27AE60', label='Equilibrio (estable)')
    
            # L√≠neas de referencia
            ax1.axvline(stats_eq['media'], color='#1E8449', linestyle='--', linewidth=2,
                        label=f"Media Eq: {stats_eq['media']/1000:.0f}k")
            ax1.axvline(stats_eq['p95'],   color='#0B5345', linestyle=':', linewidth=1.5,
                        label=f"P95 Eq: {stats_eq['p95']/1000:.0f}k")
            ax1.axvline(stats_opt['p95'],  color='#943126', linestyle=':', linewidth=1.5,
                        label=f"P95 Opt: {stats_opt['p95']/1000:.0f}k")
    
            ax1.set_title(f'DISTRIBUCI√ìN DEL SALDO DE REGULARIZACI√ìN (GRUPO) ‚Äì {n_sim} escenarios',
                          fontsize=12, fontweight='bold')
            ax1.set_xlabel('Saldo a pagar en Marzo (S/)')
            ax1.set_ylabel('Densidad')
            ax1.yaxis.set_visible(False)   # solo queremos ver la forma
            ax1.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}k'))
            ax1.legend(loc='upper right', fontsize=9)
    
            # Recuadro con estad√≠sticas clave
            textstr = f"Equilibrio vs √ìptimo\n"
            textstr += f"Media:    {stats_eq['media']/1000:.0f}k vs {stats_opt['media']/1000:.0f}k\n"
            textstr += f"P95:      {stats_eq['p95']/1000:.0f}k vs {stats_opt['p95']/1000:.0f}k\n"
            textstr += f"Reducci√≥n P95: {100*(1 - stats_eq['p95']/stats_opt['p95']):.1f}%\n"
            textstr += f"IC 95% Eq: [{stats_eq['ic_low']/1000:.0f}k , {stats_eq['ic_high']/1000:.0f}k]\n"
            textstr += f"p-valor (media Opt > Eq): {p_valor_diff:.4f}"
    
            props = dict(boxstyle='round', facecolor='wheat', alpha=0.8)
            ax1.text(0.98, 0.95, textstr, transform=ax1.transAxes, fontsize=9,
                     verticalalignment='top', horizontalalignment='right',
                     bbox=props)
    
            # ----- PANEL INFERIOR: VIOLINPLOT POR EMPRESA -----
            ax2 = fig.add_subplot(gs[1])
    
            empresas = ['Matriz'] + [s['nombre'] for s in self.resultados['satelites']]
            data_opt = [historial_saldos[emp]['opt'] for emp in empresas]
            data_eq  = [historial_saldos[emp]['eq']  for emp in empresas]
    
            posiciones = np.arange(len(empresas)) * 2.0
            ancho = 0.7
    
            # Violinplots para escenario √≥ptimo (rojo)
            vp_opt = ax2.violinplot(data_opt, positions=posiciones - 0.4,
                                     widths=ancho, showmeans=False, showmedians=True)
            for pc in vp_opt['bodies']:
                pc.set_facecolor('#E74C3C')
                pc.set_alpha(0.5)
                pc.set_edgecolor('#C0392B')
            vp_opt['cmedians'].set_color('#C0392B')
            vp_opt['cmedians'].set_linewidth(1.5)
    
            # Violinplots para escenario equilibrio (verde)
            vp_eq = ax2.violinplot(data_eq, positions=posiciones + 0.4,
                                    widths=ancho, showmeans=False, showmedians=True)
            for pc in vp_eq['bodies']:
                pc.set_facecolor('#27AE60')
                pc.set_alpha(0.6)
                pc.set_edgecolor('#1E8449')
            vp_eq['cmedians'].set_color('#1E8449')
            vp_eq['cmedians'].set_linewidth(1.5)
    
            # L√≠nea horizontal en cero
            ax2.axhline(0, color='black', linewidth=0.8, linestyle='-', alpha=0.7)
            ax2.text(posiciones[-1] + 1.2, 0, 'Saldo 0', va='center', fontsize=8)
    
            ax2.set_title('RIESGO DE REGULARIZACI√ìN POR EMPRESA (√ìptimo vs Equilibrio)',
                          fontsize=12, fontweight='bold')
            ax2.set_xticks(posiciones)
            ax2.set_xticklabels(empresas, fontsize=9, fontweight='bold')
            ax2.set_ylabel('Saldo a Regularizar en Marzo (S/)')
            ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}k'))
            ax2.grid(axis='y', alpha=0.2, linestyle='--')
    
            # Leyenda personalizada
            from matplotlib.patches import Patch
            legend_elements = [
                Patch(facecolor='#E74C3C', alpha=0.6, edgecolor='#C0392B',
                      label='Escenario √ìptimo (m√°x. ahorro)'),
                Patch(facecolor='#27AE60', alpha=0.7, edgecolor='#1E8449',
                      label='Escenario Equilibrio (saldo ~0)')
            ]
            ax2.legend(handles=legend_elements, loc='upper right', fontsize=9)
    
            plt.tight_layout()
            self.fig_equilibrio_sim = fig
    
            # --------------------------------------------------------------
            # 6. GUARDAR RESULTADOS PARA EXPORTACI√ìN
            # --------------------------------------------------------------
            self.datos_equilibrio_sim = {
                'n_sim': n_sim,
                'variabilidad': variabilidad * 100,
                'config': {
                    'var_matriz_costos': var_matriz_costos,
                    'pct_var_costos': pct_var_costos * 100,
                    'var_matriz_ingresos': var_matriz_ingresos,
                    'pct_var_ingresos': pct_var_ingresos * 100,
                    'distribucion': distribucion,
                    'rho': rho
                },
                'grupo_eq': stats_eq,
                'grupo_opt': stats_opt,
                'p_valor_diferencia': p_valor_diff,
                'historial_saldos': historial_saldos,   # guardado para posible exportaci√≥n
                'empresas': empresas
            }
    
            self.canvas_equilibrio_sim = FigureCanvasTkAgg(fig, self.frame_eq_sim_graficos)
            self.canvas_equilibrio_sim.draw()
            self.canvas_equilibrio_sim.get_tk_widget().pack(fill='both', expand=True)
    
            # --------------------------------------------------------------
            # 7. MENSAJE DE FINALIZACI√ìN
            # --------------------------------------------------------------
            messagebox.showinfo(
                "Simulaci√≥n de Equilibrio Financiero",
                f"An√°lisis completado.\n\n"
                f"Saldo medio en equilibrio: {stats_eq['media']:,.0f}\n"
                f"IC 95%: [{stats_eq['ic_low']:,.0f} , {stats_eq['ic_high']:,.0f}]\n"
                f"Reducci√≥n del P95: {100*(1 - stats_eq['p95']/stats_opt['p95']):.1f}%\n"
                f"p-valor (√≥ptimo > equilibrio): {p_valor_diff:.4f}"
            )
    
        except Exception as e:
            messagebox.showerror("Error", f"Error en simulaci√≥n de equilibrio:\n{str(e)}")
            import traceback
            traceback.print_exc()   # para depuraci√≥n en consola



 
    
    
    
    def analizar_pagos_cuenta(self):
        """Analisis de punto de equilibrio financiero: donde IR = Pagos a Cuenta (saldo = 0, eficiencia = 100%).

        El margen de equilibrio es la metrica principal. El margen optimo (regimen) se muestra como referencia.
        Usa calcular_margen_equilibrio_exacto con 3 escenarios: Especial, Mixto MYPE, General.
        """
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Advertencia", "Primero debe calcular los margenes optimos")
                return

            tasa_pc = self.TASA_PAGO_CUENTA.get() / 100
            tasa_general = self.resultados['parametros']['tasa_general'] / 100
            tasa_especial = self.resultados['parametros']['tasa_especial'] / 100
            limite_utilidad = self.resultados['parametros']['limite_utilidad']
            limite_ingresos = self.resultados['parametros']['limite_ingresos']
            n_pasos = int(self.entry_pasos_sens_pc.get())

            resultados_pc = []

            for sat in self.resultados['satelites']:
                costo = sat['costo']
                gastos = sat['gastos_operativos']
                margen_actual = sat['margen_optimo'] / 100
                es_general = sat['regimen'].startswith('GENERAL')

                # === MARGEN DE EQUILIBRIO EXACTO (3 escenarios: A, B, C) ===
                margen_equilibrio, regimen_equilibrio = self._calcular_margen_equilibrio_exacto(
                    costo, gastos, tasa_pc, tasa_especial, tasa_general,
                    limite_utilidad, limite_ingresos, es_general
                )

                # === Metricas EN el punto de equilibrio ===
                if margen_equilibrio is not None:
                    precio_eq = costo * (1 + margen_equilibrio)
                    utilidad_neta_eq = max(0, costo * margen_equilibrio - gastos)
                    pago_cuenta_eq = tasa_pc * precio_eq
                    ir_eq = self._calcular_ir_equilibrio(
                        utilidad_neta_eq, tasa_especial, tasa_general,
                        limite_utilidad, regimen_equilibrio
                    )
                    saldo_eq = ir_eq - pago_cuenta_eq  # debe ser ~0
                    ahorro_eq = utilidad_neta_eq * tasa_general - ir_eq
                else:
                    precio_eq = 0
                    utilidad_neta_eq = 0
                    pago_cuenta_eq = 0
                    ir_eq = 0
                    saldo_eq = 0
                    ahorro_eq = 0

                # === Metricas EN el margen optimo (referencia) ===
                precio_opt = costo * (1 + margen_actual)
                pago_cuenta_opt = tasa_pc * precio_opt
                ir_opt = sat['impuesto']
                saldo_opt = ir_opt - pago_cuenta_opt
                eficiencia_opt = (ir_opt / pago_cuenta_opt * 100) if pago_cuenta_opt > 0 else 0

                # === Sensibilidad: variar margen centrado en el equilibrio ===
                centro = margen_equilibrio if margen_equilibrio is not None else margen_actual
                margen_min = max(0.001, centro * 0.2)
                margen_max_feasible = min(limite_ingresos / costo - 1, 1.0) if costo > 0 else 1.0
                margen_max = min(centro * 3.0, margen_max_feasible)
                if margen_max <= margen_min:
                    margen_max = margen_min + 0.10

                margenes_test = np.linspace(margen_min, margen_max, n_pasos)

                margenes_extras = [margen_actual]
                if margen_equilibrio is not None:
                    margenes_extras.append(margen_equilibrio)
                margenes_test = np.sort(np.unique(np.concatenate([margenes_test, margenes_extras])))

                sensibilidad = []
                for m in margenes_test:
                    precio = costo * (1 + m)
                    utilidad_neta = max(0, costo * m - gastos)

                    if es_general:
                        ir = utilidad_neta * tasa_general
                        regimen_m = "General"
                        ahorro_regimen = 0
                    elif utilidad_neta > limite_utilidad or precio > limite_ingresos:
                        if utilidad_neta > limite_utilidad:
                            ir = limite_utilidad * tasa_especial + (utilidad_neta - limite_utilidad) * tasa_general
                            regimen_m = "Mixto"
                        else:
                            ir = utilidad_neta * tasa_general
                            regimen_m = "General*"
                        ahorro_regimen = utilidad_neta * tasa_general - ir
                    else:
                        ir = utilidad_neta * tasa_especial
                        regimen_m = "Especial"
                        ahorro_regimen = utilidad_neta * (tasa_general - tasa_especial)

                    pago_cuenta = tasa_pc * precio
                    saldo = ir - pago_cuenta
                    efic = (ir / pago_cuenta * 100) if pago_cuenta > 0 else 0

                    sensibilidad.append({
                        'margen': m * 100,
                        'precio_venta': precio,
                        'utilidad_neta': utilidad_neta,
                        'regimen': regimen_m,
                        'ir': ir,
                        'pago_cuenta': pago_cuenta,
                        'saldo': saldo,
                        'ahorro_regimen': ahorro_regimen,
                        'eficiencia_pc': efic,
                        'es_optimo': abs(m - margen_actual) < 0.0001,
                        'es_equilibrio': margen_equilibrio is not None and abs(m - margen_equilibrio) < 0.0001,
                    })

                # Sensibilidad 2D: recalcular equilibrio exacto para cada tasa PaC
                tasas_pc_test = [0.5, 1.0, 1.5, 2.0, 2.5]
                sens_2d = []
                for t_pc in tasas_pc_test:
                    t_pc_dec = t_pc / 100
                    m_eq_2d, reg_2d = self._calcular_margen_equilibrio_exacto(
                        costo, gastos, t_pc_dec, tasa_especial, tasa_general,
                        limite_utilidad, limite_ingresos, es_general
                    )

                    if m_eq_2d is not None and m_eq_2d > 0:
                        precio_2d = costo * (1 + m_eq_2d)
                        util_2d = max(0, costo * m_eq_2d - gastos)
                        ir_2d = self._calcular_ir_equilibrio(
                            util_2d, tasa_especial, tasa_general, limite_utilidad, reg_2d
                        )
                        pc_2d = t_pc_dec * precio_2d
                        ahorro_2d = util_2d * tasa_general - ir_2d
                    else:
                        m_eq_2d = 0
                        precio_2d = 0
                        util_2d = 0
                        ir_2d = 0
                        pc_2d = 0
                        ahorro_2d = 0
                        reg_2d = "N/A"

                    sens_2d.append({
                        'tasa_pc': t_pc,
                        'margen_equilibrio': m_eq_2d * 100 if m_eq_2d else 0,
                        'precio_venta': precio_2d,
                        'utilidad_neta': util_2d,
                        'pago_cuenta': pc_2d,
                        'ir': ir_2d,
                        'saldo': ir_2d - pc_2d,
                        'ahorro_regimen': ahorro_2d,
                        'regimen': reg_2d,
                    })

                resultados_pc.append({
                    'nombre': sat['nombre'],
                    'costo': costo,
                    'gastos': gastos,
                    'es_general': es_general,
                    # Equilibrio (METRICA PRINCIPAL)
                    'margen_equilibrio': margen_equilibrio * 100 if margen_equilibrio is not None else None,
                    'regimen_equilibrio': regimen_equilibrio,
                    'precio_equilibrio': precio_eq,
                    'utilidad_neta_eq': utilidad_neta_eq,
                    'ir_equilibrio': ir_eq,
                    'pago_cuenta_eq': pago_cuenta_eq,
                    'saldo_equilibrio': saldo_eq,
                    'ahorro_equilibrio': ahorro_eq,
                    # Margen optimo (REFERENCIA)
                    'margen_optimo_regimen': margen_actual * 100,
                    'precio_optimo': precio_opt,
                    'ir_optimo': ir_opt,
                    'pago_cuenta_optimo': pago_cuenta_opt,
                    'saldo_optimo': saldo_opt,
                    'eficiencia_optimo': eficiencia_opt,
                    # Sensibilidad
                    'sensibilidad': sensibilidad,
                    'sensibilidad_2d': sens_2d,
                })

            # --- ANALISIS GLOBAL: SIN ESTRUCTURA vs CON ESTRUCTURA (enfoque equilibrio financiero) ---
            r = self.resultados
            impuesto_sin_estructura = r['matriz']['impuesto_sin_estructura']
            utilidad_sin_estructura = r['matriz']['utilidad_sin_estructura']

            # Datos base
            ingresos_matriz = r['matriz']['ingresos']
            total_costos_satelites = sum(rpc['costo'] for rpc in resultados_pc)

            # Pagos a cuenta SIN estructura
            pc_matriz_sin = tasa_pc * ingresos_matriz
            saldo_matriz_sin = impuesto_sin_estructura - pc_matriz_sin

            # === CONSOLIDADO EN PUNTO DE EQUILIBRIO (CORRECCION: recalcular impuesto matriz) ===
            total_compras_eq = sum(rpc['precio_equilibrio'] for rpc in resultados_pc)
            total_ir_eq_sats = sum(rpc['ir_equilibrio'] for rpc in resultados_pc)
            total_pc_eq_sats = sum(rpc['pago_cuenta_eq'] for rpc in resultados_pc)
            total_saldo_eq_sats = sum(rpc['saldo_equilibrio'] for rpc in resultados_pc)

            # Recalcular utilidad e impuesto de la matriz EN EQUILIBRIO
            # La matriz compra a los satelites a precio_equilibrio en lugar de a costo
            utilidad_matriz_eq = utilidad_sin_estructura - total_compras_eq + total_costos_satelites
            utilidad_matriz_eq = max(0, utilidad_matriz_eq)
            impuesto_matriz_eq = utilidad_matriz_eq * tasa_general
            pc_matriz_eq = tasa_pc * ingresos_matriz
            saldo_matriz_eq = impuesto_matriz_eq - pc_matriz_eq

            # Grupo en equilibrio
            ir_grupo_eq = impuesto_matriz_eq + total_ir_eq_sats
            pc_grupo_eq = pc_matriz_eq + total_pc_eq_sats
            saldo_grupo_eq = ir_grupo_eq - pc_grupo_eq
            eficiencia_eq = (ir_grupo_eq / pc_grupo_eq * 100) if pc_grupo_eq > 0 else 0
            utilidad_eq_sats = sum(rpc['utilidad_neta_eq'] for rpc in resultados_pc)
            utilidad_total_eq = utilidad_matriz_eq + utilidad_eq_sats
            tasa_efectiva_eq = (ir_grupo_eq / utilidad_total_eq * 100) if utilidad_total_eq > 0 else 0
            ahorro_eq_total = impuesto_sin_estructura - ir_grupo_eq
            ahorro_eq_pct = (ahorro_eq_total / impuesto_sin_estructura * 100) if impuesto_sin_estructura > 0 else 0

            # === CONSOLIDADO EN MARGEN OPTIMO (REFERENCIA - misma correccion) ===
            total_compras_opt = sum(rpc['precio_optimo'] for rpc in resultados_pc)
            total_ir_opt_sats = sum(rpc['ir_optimo'] for rpc in resultados_pc)
            total_pc_opt_sats = sum(rpc['pago_cuenta_optimo'] for rpc in resultados_pc)
            total_saldo_opt_sats = sum(rpc['saldo_optimo'] for rpc in resultados_pc)

            utilidad_matriz_opt = utilidad_sin_estructura - total_compras_opt + total_costos_satelites
            utilidad_matriz_opt = max(0, utilidad_matriz_opt)
            impuesto_matriz_opt = utilidad_matriz_opt * tasa_general
            pc_matriz_opt = tasa_pc * ingresos_matriz
            saldo_matriz_opt = impuesto_matriz_opt - pc_matriz_opt

            ir_grupo_opt = impuesto_matriz_opt + total_ir_opt_sats
            pc_grupo_opt = pc_matriz_opt + total_pc_opt_sats
            saldo_grupo_opt = ir_grupo_opt - pc_grupo_opt
            eficiencia_opt = (ir_grupo_opt / pc_grupo_opt * 100) if pc_grupo_opt > 0 else 0
            utilidad_opt_sats = r['grupo']['total_utilidad_satelites']
            utilidad_total_opt = utilidad_matriz_opt + utilidad_opt_sats
            tasa_efectiva_opt = (ir_grupo_opt / utilidad_total_opt * 100) if utilidad_total_opt > 0 else 0
            ahorro_opt_total = impuesto_sin_estructura - ir_grupo_opt
            ahorro_opt_pct = (ahorro_opt_total / impuesto_sin_estructura * 100) if impuesto_sin_estructura > 0 else 0

            eficiencia_sin = (impuesto_sin_estructura / pc_matriz_sin * 100) if pc_matriz_sin > 0 else 0
            tasa_efectiva_sin = (impuesto_sin_estructura / utilidad_sin_estructura * 100) if utilidad_sin_estructura > 0 else 0

            self.datos_pagos_cuenta = {
                'tasa_pago_cuenta': tasa_pc * 100,
                'satelites': resultados_pc,
                'global': {
                    'nombre_matriz': r['matriz']['nombre'],
                    'ingresos_matriz': r['matriz']['ingresos'],
                    'impuesto_sin_estructura': impuesto_sin_estructura,
                    'utilidad_sin_estructura': utilidad_sin_estructura,
                    'pc_matriz_sin': pc_matriz_sin,
                    'saldo_matriz_sin': saldo_matriz_sin,
                    'tasa_efectiva_sin': tasa_efectiva_sin,
                    'eficiencia_sin': eficiencia_sin,
                    # Matriz en equilibrio
                    'utilidad_matriz_eq': utilidad_matriz_eq,
                    'impuesto_matriz_eq': impuesto_matriz_eq,
                    'pc_matriz_eq': pc_matriz_eq,
                    'saldo_matriz_eq': saldo_matriz_eq,
                    # Equilibrio (principal)
                    'ir_grupo_eq': ir_grupo_eq,
                    'pc_grupo_eq': pc_grupo_eq,
                    'saldo_grupo_eq': saldo_grupo_eq,
                    'eficiencia_eq': eficiencia_eq,
                    'tasa_efectiva_eq': tasa_efectiva_eq,
                    'ahorro_eq_total': ahorro_eq_total,
                    'ahorro_eq_pct': ahorro_eq_pct,
                    'total_ir_eq_sats': total_ir_eq_sats,
                    'total_pc_eq_sats': total_pc_eq_sats,
                    'total_saldo_eq_sats': total_saldo_eq_sats,
                    # Matriz en margen optimo
                    'utilidad_matriz_opt': utilidad_matriz_opt,
                    'impuesto_matriz_opt': impuesto_matriz_opt,
                    'pc_matriz_opt': pc_matriz_opt,
                    'saldo_matriz_opt': saldo_matriz_opt,
                    # Margen optimo (referencia)
                    'ir_grupo_opt': ir_grupo_opt,
                    'pc_grupo_opt': pc_grupo_opt,
                    'saldo_grupo_opt': saldo_grupo_opt,
                    'eficiencia_opt': eficiencia_opt,
                    'tasa_efectiva_opt': tasa_efectiva_opt,
                    'ahorro_opt_total': ahorro_opt_total,
                    'ahorro_opt_pct': ahorro_opt_pct,
                    'total_ir_opt_sats': total_ir_opt_sats,
                    'total_pc_opt_sats': total_pc_opt_sats,
                    'total_saldo_opt_sats': total_saldo_opt_sats,
                },
            }

            self._mostrar_resultados_pagos_cuenta()
            self._generar_graficos_pagos_cuenta()

            n_eq = sum(1 for rpc in resultados_pc if rpc['margen_equilibrio'] is not None)
            messagebox.showinfo("Analisis de Equilibrio Financiero",
                                f"Punto de equilibrio calculado para {n_eq}/{len(resultados_pc)} satelites.\n"
                                f"Eficiencia PaC en Equilibrio: {eficiencia_eq:.1f}% (ideal=100%)\n"
                                f"Saldo Grupo en Equilibrio: {saldo_grupo_eq:,.2f}")

        except Exception as e:
            messagebox.showerror("Error", f"Error en analisis de pagos a cuenta: {str(e)}")

    def _mostrar_resultados_pagos_cuenta(self):
        """Muestra resultados priorizando el punto de equilibrio financiero (IR = PaC, saldo = 0)."""
        self.text_pagos_cuenta.delete(1.0, tk.END)
        d = self.datos_pagos_cuenta
        g = d['global']

        texto = f"""{'='*140}
  PUNTO DE EQUILIBRIO FINANCIERO - Donde IR = Pagos a Cuenta (Saldo = 0, Eficiencia = 100%)
  Tasa Pago a Cuenta: {d['tasa_pago_cuenta']:.1f}%  |  Grupo: {g['nombre_matriz']}
{'='*140}

  CONCEPTO: El punto de equilibrio financiero es el margen donde los pagos a cuenta adelantados
  al fisco ({d['tasa_pago_cuenta']:.1f}% de ingresos) igualan EXACTAMENTE al IR anual definitivo.
  En este punto: saldo de regularizacion = 0, eficiencia = 100%, flujo de caja optimizado.

  Formulas segun escenario:
  A) Especial Puro: m = (C*t_pc + G*t_esp) / (C*(t_esp - t_pc))       [si Util <= 15 UIT]
  B) Mixto MYPE:    m = (C*t_pc + G*t_gral + L*(t_gral-t_esp)) / (C*(t_gral-t_pc))  [si Util > 15 UIT e Ingreso <= 1700 UIT]
  C) General Puro:  m = (C*t_pc + G*t_gral) / (C*(t_gral - t_pc))     [si Ingreso > 1700 UIT]

{'='*140}
  I. IMPUESTO GLOBAL SIN ESTRUCTURA SATELITAL
{'='*140}
  Empresa:                        {g['nombre_matriz']}
  Ingresos Totales:               {g['ingresos_matriz']:>15,.2f}
  Utilidad Gravable:              {g['utilidad_sin_estructura']:>15,.2f}
  Tasa Aplicable:                 {'29.5% (Regimen General)':>30}
  Impuesto a la Renta:            {g['impuesto_sin_estructura']:>15,.2f}
  {'-'*70}
  Pagos a Cuenta ({d['tasa_pago_cuenta']:.1f}%):        {g['pc_matriz_sin']:>15,.2f}
  Saldo Regularizacion:           {g['saldo_matriz_sin']:>15,.2f}  {'(SALDO A FAVOR)' if g['saldo_matriz_sin'] < 0 else '(SALDO POR PAGAR)'}
  Eficiencia PaC:                 {g['eficiencia_sin']:>10.1f}%  (ideal=100%)
  Tasa Efectiva:                  {g['tasa_efectiva_sin']:>10.1f}%

{'='*140}
  II. IMPUESTO GLOBAL CON ESTRUCTURA EN PUNTO DE EQUILIBRIO (IR = PaC)
{'='*140}
  A) EMPRESA MATRIZ
  {'-'*70}
  Utilidad Matriz (ajustada):     {g['utilidad_matriz_eq']:>15,.2f}
  Impuesto Matriz:                {g['impuesto_matriz_eq']:>15,.2f}
  Pagos a Cuenta Matriz:          {g['pc_matriz_eq']:>15,.2f}
  Saldo Matriz:                   {g['saldo_matriz_eq']:>15,.2f}  {'(FAVOR)' if g['saldo_matriz_eq'] < 0 else '(PAGAR)'}

  B) SATELITES EN PUNTO DE EQUILIBRIO (saldo = 0 por satelite)
  {'-'*70}
  IR Satelites (en equilibrio):   {g['total_ir_eq_sats']:>15,.2f}
  PaC Satelites (en equilibrio):  {g['total_pc_eq_sats']:>15,.2f}
  Saldo Satelites:                {g['total_saldo_eq_sats']:>15,.2f}  (debe ser ~0)

  C) GRUPO CONSOLIDADO EN EQUILIBRIO
  {'-'*70}
  IR Total Grupo:                 {g['ir_grupo_eq']:>15,.2f}
  PaC Total Grupo:                {g['pc_grupo_eq']:>15,.2f}
  Saldo Grupo:                    {g['saldo_grupo_eq']:>15,.2f}  {'(~0 = EQUILIBRIO PERFECTO)' if abs(g['saldo_grupo_eq']) < 100 else '(FAVOR)' if g['saldo_grupo_eq'] < 0 else '(PAGAR)'}
  Eficiencia PaC:                 {g['eficiencia_eq']:>10.1f}%  (ideal=100%)
  Tasa Efectiva:                  {g['tasa_efectiva_eq']:>10.1f}%
  Ahorro vs Sin Estructura:       {g['ahorro_eq_total']:>15,.2f} ({g['ahorro_eq_pct']:.1f}%)

{'='*140}
  III. COMPARATIVO: SIN ESTRUCTURA vs EQUILIBRIO vs MARGEN OPTIMO
{'='*140}
  {'Concepto':<35} | {'SIN Estructura':>16} | {'EQUILIBRIO':>16} | {'Margen Optimo':>16} | {'Mejor Escenario':>16}
  {'-'*110}
  {'IR Total Grupo':<35} | {g['impuesto_sin_estructura']:>16,.0f} | {g['ir_grupo_eq']:>16,.0f} | {g['ir_grupo_opt']:>16,.0f} | {'Equilibrio' if g['ir_grupo_eq'] < g['ir_grupo_opt'] else 'M. Optimo':>16}
  {'PaC Total Grupo':<35} | {g['pc_matriz_sin']:>16,.0f} | {g['pc_grupo_eq']:>16,.0f} | {g['pc_grupo_opt']:>16,.0f} | {'-':>16}
  {'Saldo Regularizacion':<35} | {g['saldo_matriz_sin']:>16,.0f} | {g['saldo_grupo_eq']:>16,.0f} | {g['saldo_grupo_opt']:>16,.0f} | {'Equilibrio' if abs(g['saldo_grupo_eq']) < abs(g['saldo_grupo_opt']) else 'M. Optimo':>16}
  {'Eficiencia PaC (%)':<35} | {g['eficiencia_sin']:>15.1f}% | {g['eficiencia_eq']:>15.1f}% | {g['eficiencia_opt']:>15.1f}% | {'Equilibrio' if abs(g['eficiencia_eq'] - 100) < abs(g['eficiencia_opt'] - 100) else 'M. Optimo':>16}
  {'Tasa Efectiva (%)':<35} | {g['tasa_efectiva_sin']:>15.1f}% | {g['tasa_efectiva_eq']:>15.1f}% | {g['tasa_efectiva_opt']:>15.1f}% | {'Equilibrio' if g['tasa_efectiva_eq'] < g['tasa_efectiva_opt'] else 'M. Optimo':>16}
  {'Ahorro Tributario':<35} | {'-':>16} | {g['ahorro_eq_total']:>16,.0f} | {g['ahorro_opt_total']:>16,.0f} | {'Equilibrio' if g['ahorro_eq_total'] > g['ahorro_opt_total'] else 'M. Optimo':>16}
  {'Ahorro %':<35} | {'-':>16} | {g['ahorro_eq_pct']:>15.1f}% | {g['ahorro_opt_pct']:>15.1f}% | {'-':>16}
  {'='*110}

{'='*140}
  IV. DETALLE POR SATELITE - PUNTO DE EQUILIBRIO FINANCIERO
{'='*140}
"""

        for i, sat in enumerate(d['satelites'], 1):
            tiene_eq = sat['margen_equilibrio'] is not None
            eq_str = f"{sat['margen_equilibrio']:.4f}%" if tiene_eq else "N/A"
            texto += f"""
  {'-'*140}
  {i}. {sat['nombre']} | Costo: {sat['costo']:,.0f} | Gastos: {sat['gastos']:,.0f}
  {'-'*140}
  >>> PUNTO DE EQUILIBRIO (IR = PaC, Saldo = 0) <<<
  Margen Equilibrio:              {eq_str:>20}  {'   [Regimen: ' + sat['regimen_equilibrio'] + ']' if tiene_eq else '   [Sin equilibrio factible]'}
  Precio Venta en Equilibrio:     {sat['precio_equilibrio']:>15,.2f}
  Utilidad Neta en Equilibrio:    {sat['utilidad_neta_eq']:>15,.2f}
  IR en Equilibrio:               {sat['ir_equilibrio']:>15,.2f}
  PaC en Equilibrio:              {sat['pago_cuenta_eq']:>15,.2f}
  Saldo en Equilibrio:            {sat['saldo_equilibrio']:>15,.2f}  {'(~0 PERFECTO)' if abs(sat['saldo_equilibrio']) < 1 else ''}
  Ahorro vs Reg. General:         {sat['ahorro_equilibrio']:>15,.2f}

  --- Referencia: Margen Optimo (maximiza ahorro fiscal) ---
  Margen Optimo:                  {sat['margen_optimo_regimen']:>10.4f}%
  IR en Margen Optimo:            {sat['ir_optimo']:>15,.2f}
  PaC en Margen Optimo:           {sat['pago_cuenta_optimo']:>15,.2f}
  Saldo en Margen Optimo:         {sat['saldo_optimo']:>15,.2f}  {'(FAVOR)' if sat['saldo_optimo'] < 0 else '(PAGAR)'}
  Eficiencia PaC en Optimo:       {sat['eficiencia_optimo']:>10.1f}%
"""
            # Tabla de sensibilidad
            texto += f"""
  {'SENSIBILIDAD - Como varia el saldo al mover el margen':^140}
  El punto EQUILIBRIO tiene saldo=0 (eficiencia 100%). El punto OPTIMO maximiza el ahorro fiscal.
  {'-'*140}
  {'Margen%':>8} | {'Precio Venta':>14} | {'Util. Neta':>12} | {'Regimen':>10} | {'IR':>12} | {'Pago Cta':>12} | {'Saldo':>12} | {'Ahorro Reg':>12} | {'Efic%':>7} | {'Nota':>14}
  {'-'*140}
"""
            for s in sat['sensibilidad']:
                nota = ""
                if s['es_equilibrio']:
                    nota = "<<< EQUILIBRIO"
                elif s['es_optimo']:
                    nota = "(ref: optimo)"
                texto += f"  {s['margen']:>8.4f} | {s['precio_venta']:>14,.0f} | {s['utilidad_neta']:>12,.0f} | {s['regimen']:>10} | {s['ir']:>12,.0f} | {s['pago_cuenta']:>12,.0f} | {s['saldo']:>12,.0f} | {s['ahorro_regimen']:>12,.0f} | {s['eficiencia_pc']:>6.1f}% | {nota:>14}\n"

            # Tabla sensibilidad 2D
            texto += f"""
  {'SENSIBILIDAD - Punto de Equilibrio ante cambios en la Tasa de Pago a Cuenta':^140}
  Si SUNAT cambiara la tasa PaC, cual seria el nuevo margen de equilibrio para mantener saldo = 0.
  {'-'*120}
  {'Tasa PaC%':>10} | {'M. Equil%':>10} | {'Precio Eq.':>14} | {'Util. Neta':>12} | {'IR':>12} | {'PaC':>12} | {'Saldo':>12} | {'Ahorro Reg':>12} | {'Regimen':>10}
  {'-'*120}
"""
            for s2 in sat['sensibilidad_2d']:
                texto += f"  {s2['tasa_pc']:>10.1f} | {s2['margen_equilibrio']:>9.4f}% | {s2['precio_venta']:>14,.0f} | {s2['utilidad_neta']:>12,.0f} | {s2['ir']:>12,.0f} | {s2['pago_cuenta']:>12,.0f} | {s2['saldo']:>12,.0f} | {s2['ahorro_regimen']:>12,.0f} | {s2['regimen']:>10}\n"

        texto += f"""
{'='*140}
  V. INTERPRETACION ESTRATEGICA
{'='*140}

  PUNTO DE EQUILIBRIO FINANCIERO (concepto central):
  Cada satelite tiene un margen donde IR = PaC, lo que significa:
    - Saldo de regularizacion = 0 (no hay dinero inmovilizado ni falta liquidez)
    - Eficiencia del pago a cuenta = 100%
    - El flujo de caja es optimo: cada sol adelantado se consume exactamente

  EQUILIBRIO vs MARGEN OPTIMO:
  - El MARGEN OPTIMO maximiza el ahorro fiscal (diferencial de tasas).
  - El MARGEN DE EQUILIBRIO maximiza la eficiencia financiera (saldo = 0).
  - Si ambos coinciden: situacion ideal.
  - Si difieren: la gerencia debe decidir que priorizar:
      * Equilibrio: mejor flujo de caja, cero saldos pendientes
      * Optimo: mayor ahorro fiscal, pero posible saldo a favor/contra

  RESULTADO DEL GRUPO EN EQUILIBRIO:
  - IR Total: {g['ir_grupo_eq']:,.0f} | Ahorro: {g['ahorro_eq_total']:,.0f} ({g['ahorro_eq_pct']:.1f}%)
  - Eficiencia PaC: {g['eficiencia_eq']:.1f}% | Tasa Efectiva: {g['tasa_efectiva_eq']:.1f}%
{'='*140}
"""
        self.text_pagos_cuenta.insert(1.0, texto)

    def _generar_graficos_pagos_cuenta(self):
        """Genera graficos priorizando el punto de equilibrio financiero."""
        if self.canvas_pagos_cuenta:
            self.canvas_pagos_cuenta.get_tk_widget().destroy()

        d = self.datos_pagos_cuenta
        g = d['global']
        n_sats = len(d['satelites'])

        # Layout: primera fila = 2 paneles globales, luego satelites
        n_cols = 2
        n_rows_sats = (n_sats + 1) // 2
        n_rows = 1 + n_rows_sats

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(14, 5 * n_rows), squeeze=False)
        fig.suptitle('PUNTO DE EQUILIBRIO FINANCIERO - IR = Pagos a Cuenta (Saldo = 0)',
                     fontsize=14, fontweight='bold', color='#1F4E79')

        # --- PANEL GLOBAL 1: Comparativo 3 escenarios ---
        ax_g1 = axes[0][0]
        categorias = ['IR\nTotal', 'Pagos a\nCuenta', 'Saldo\nRegulariz.']
        vals_sin = [g['impuesto_sin_estructura'], g['pc_matriz_sin'], g['saldo_matriz_sin']]
        vals_eq = [g['ir_grupo_eq'], g['pc_grupo_eq'], g['saldo_grupo_eq']]
        vals_opt = [g['ir_grupo_opt'], g['pc_grupo_opt'], g['saldo_grupo_opt']]
        x = np.arange(len(categorias))
        width = 0.25
        bars1 = ax_g1.bar(x - width, vals_sin, width, label='Sin Estructura', color='#E74C3C', alpha=0.85, edgecolor='white')
        bars2 = ax_g1.bar(x, vals_eq, width, label='Equilibrio (IR=PaC)', color='#3498DB', alpha=0.85, edgecolor='white')
        bars3 = ax_g1.bar(x + width, vals_opt, width, label='Margen Optimo (ref)', color='#95A5A6', alpha=0.7, edgecolor='white')
        ax_g1.set_title('Impuesto Global: Sin Estructura vs Equilibrio vs Optimo', fontsize=10, fontweight='bold')
        ax_g1.set_xticks(x)
        ax_g1.set_xticklabels(categorias, fontsize=8)
        ax_g1.legend(fontsize=7)
        ax_g1.grid(True, alpha=0.3, axis='y')
        ax_g1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:,.0f}K'))
        for bar in bars1:
            h = bar.get_height()
            ax_g1.text(bar.get_x() + bar.get_width()/2., h, f'{h/1000:,.0f}K',
                      ha='center', va='bottom', fontsize=6, fontweight='bold', color='#C0392B')
        for bar in bars2:
            h = bar.get_height()
            ax_g1.text(bar.get_x() + bar.get_width()/2., h, f'{h/1000:,.0f}K',
                      ha='center', va='bottom', fontsize=6, fontweight='bold', color='#2980B9')
        ax_g1.annotate(f'Ahorro Equilibrio:\nS/ {g["ahorro_eq_total"]:,.0f} ({g["ahorro_eq_pct"]:.1f}%)',
                       xy=(0.98, 0.95), xycoords='axes fraction', ha='right', va='top',
                       fontsize=8, fontweight='bold', color='#2980B9',
                       bbox=dict(boxstyle='round,pad=0.4', facecolor='#E8F0FE', edgecolor='#2980B9', alpha=0.9))

        # --- PANEL GLOBAL 2: Eficiencia (3 escenarios) ---
        ax_g2 = axes[0][1]
        metricas = ['Tasa\nEfectiva %', 'Eficiencia\nPaC %']
        vals_sin_pct = [g['tasa_efectiva_sin'], g['eficiencia_sin']]
        vals_eq_pct = [g['tasa_efectiva_eq'], g['eficiencia_eq']]
        vals_opt_pct = [g['tasa_efectiva_opt'], g['eficiencia_opt']]
        x2 = np.arange(len(metricas))
        bars4 = ax_g2.bar(x2 - width, vals_sin_pct, width, label='Sin Estructura', color='#E74C3C', alpha=0.85, edgecolor='white')
        bars5 = ax_g2.bar(x2, vals_eq_pct, width, label='Equilibrio', color='#3498DB', alpha=0.85, edgecolor='white')
        bars6 = ax_g2.bar(x2 + width, vals_opt_pct, width, label='Margen Optimo (ref)', color='#95A5A6', alpha=0.7, edgecolor='white')
        ax_g2.set_title('Eficiencia Financiera: 3 Escenarios', fontsize=10, fontweight='bold')
        ax_g2.set_xticks(x2)
        ax_g2.set_xticklabels(metricas, fontsize=9)
        ax_g2.set_ylabel('%', fontsize=10)
        ax_g2.legend(fontsize=7)
        ax_g2.grid(True, alpha=0.3, axis='y')
        ax_g2.axhline(y=100, color='#3498DB', linestyle='--', alpha=0.5, linewidth=1.5)
        ax_g2.text(1.35, 100, 'Equilibrio\nperfecto', fontsize=7, color='#3498DB', alpha=0.7, va='center')
        for bar in bars4:
            h = bar.get_height()
            ax_g2.text(bar.get_x() + bar.get_width()/2., h, f'{h:.1f}%',
                      ha='center', va='bottom', fontsize=7, fontweight='bold', color='#C0392B')
        for bar in bars5:
            h = bar.get_height()
            ax_g2.text(bar.get_x() + bar.get_width()/2., h, f'{h:.1f}%',
                      ha='center', va='bottom', fontsize=7, fontweight='bold', color='#2980B9')

        # --- PANELES POR SATELITE: IR vs PaC con equilibrio destacado ---
        for idx, sat in enumerate(d['satelites']):
            row = 1 + idx // n_cols
            col = idx % n_cols
            ax = axes[row][col]

            sens = sat['sensibilidad']
            margenes = [s['margen'] for s in sens]
            irs = [s['ir'] for s in sens]
            pacs = [s['pago_cuenta'] for s in sens]

            ax.plot(margenes, irs, 'b-', linewidth=2, label='IR Anual', marker='o', markersize=3)
            ax.plot(margenes, pacs, 'r--', linewidth=2, label=f'Pago a Cuenta ({d["tasa_pago_cuenta"]:.1f}%)', marker='s', markersize=3)
            ax.fill_between(margenes, irs, pacs, where=[i < p for i, p in zip(irs, pacs)],
                           alpha=0.15, color='red', label='Saldo a Favor')
            ax.fill_between(margenes, irs, pacs, where=[i >= p for i, p in zip(irs, pacs)],
                           alpha=0.15, color='green', label='Saldo por Pagar')

            # EQUILIBRIO: linea principal, gruesa, destacada
            if sat['margen_equilibrio'] is not None:
                ax.axvline(sat['margen_equilibrio'], color='#3498DB', linestyle='-', linewidth=2.5,
                          label=f'EQUILIBRIO ({sat["margen_equilibrio"]:.2f}%)', zorder=5)
                # Punto de cruce
                ax.plot(sat['margen_equilibrio'], sat['ir_equilibrio'], 'D', color='#3498DB',
                       markersize=10, zorder=6, markeredgecolor='white', markeredgewidth=1.5)

            # Margen optimo: linea secundaria, punteada, mas discreta
            ax.axvline(sat['margen_optimo_regimen'], color='gray', linestyle=':', linewidth=1,
                      label=f'M. Optimo ref ({sat["margen_optimo_regimen"]:.2f}%)', alpha=0.7)

            ax.set_xlabel('Margen (%)', fontsize=9)
            ax.set_ylabel('Monto (S/)', fontsize=9)
            ax.set_title(f'{sat["nombre"]} - Punto de Equilibrio', fontsize=11, fontweight='bold')
            ax.legend(fontsize=6, loc='upper left')
            ax.grid(True, alpha=0.3)
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))

        # Ocultar subplots vacios
        for idx in range(n_sats, n_rows_sats * n_cols):
            row = 1 + idx // n_cols
            col = idx % n_cols
            axes[row][col].set_visible(False)

        plt.tight_layout()

        self.fig_pagos_cuenta = fig

        self.canvas_pagos_cuenta = FigureCanvasTkAgg(fig, self.frame_pagos_graficos)
        self.canvas_pagos_cuenta.draw()
        self.canvas_pagos_cuenta.get_tk_widget().pack(fill='both', expand=True)

    def _estilo_header(self):
        """Retorna estilos para headers de tabla."""
        return {
            'font': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                bottom=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
        }

    def _estilo_celda(self, es_alterno=False):
        """Retorna estilos para celdas de datos."""
        fill_color = 'D6E4F0' if es_alterno else 'FFFFFF'
        return {
            'font': Font(name='Calibri', size=10),
            'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                bottom=Side(style='thin', color='C0C0C0'),
                top=Side(style='thin', color='C0C0C0'),
                left=Side(style='thin', color='C0C0C0'),
                right=Side(style='thin', color='C0C0C0')
            )
        }

    def _estilo_titulo(self):
        """Retorna estilos para t√≠tulos de secci√≥n."""
        return {
            'font': Font(name='Calibri', bold=True, color='1F4E79', size=14),
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    def _estilo_subtitulo(self):
        """Retorna estilos para subt√≠tulos."""
        return {
            'font': Font(name='Calibri', bold=True, color='2E75B6', size=11),
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    def _estilo_kpi_valor(self):
        """Retorna estilos para valores KPI destacados."""
        return {
            'font': Font(name='Calibri', bold=True, color='FFFFFF', size=13),
            'fill': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                bottom=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79')
            )
        }

    def _estilo_kpi_label(self):
        """Retorna estilos para etiquetas KPI."""
        return {
            'font': Font(name='Calibri', bold=True, color='1F4E79', size=10),
            'fill': PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                bottom=Side(style='thin', color='1F4E79'),
                top=Side(style='thin', color='1F4E79'),
                left=Side(style='thin', color='1F4E79'),
                right=Side(style='thin', color='1F4E79')
            )
        }

    def _aplicar_estilo(self, celda, estilo_dict):
        """Aplica un diccionario de estilos a una celda."""
        for attr, value in estilo_dict.items():
            setattr(celda, attr, value)

    def _guardar_figura_temp(self, fig, nombre, tmpdir):
        """Guarda una figura matplotlib como PNG de alta resoluci√≥n."""
        if fig is None:
            return None
        path = os.path.join(tmpdir, f"{nombre}.png")
        fig.savefig(path, dpi=180, bbox_inches='tight', facecolor='white', edgecolor='none')
        return path

    def _insertar_imagen(self, ws, img_path, celda, ancho_cm=28, alto_cm=16):
        """Inserta imagen en hoja Excel con tama√±o especificado."""
        if img_path and os.path.exists(img_path):
            img = XlImage(img_path)
            img.width = int(ancho_cm * 37.8)   # cm a pixels aprox
            img.height = int(alto_cm * 37.8)
            ws.add_image(img, celda)

    def exportar_excel(self):
        try:
            if not hasattr(self, 'resultados'):
                messagebox.showwarning("Atenci√≥n", "Ejecute el c√°lculo primero")
                return

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Reporte_Gerencial_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            if not path:
                return

            r = self.resultados
            utilidad_total = r['matriz']['nueva_utilidad'] + r['grupo']['total_utilidad_satelites']
            tasa_efectiva = (r['grupo']['impuesto_total'] / utilidad_total * 100) if utilidad_total > 0 else 0
            roi = (r['grupo']['ahorro_tributario'] / r['matriz']['total_compras'] * 100) if r['matriz']['total_compras'] > 0 else 0

            wb = Workbook()

            # Directorio temporal para im√°genes
            with tempfile.TemporaryDirectory() as tmpdir:
                # Guardar todas las figuras disponibles
                img_resultados = self._guardar_figura_temp(self.fig_resultados, 'resultados', tmpdir)
                img_simulacion = self._guardar_figura_temp(self.fig_simulacion, 'simulacion', tmpdir)
                img_sensibilidad = self._guardar_figura_temp(self.fig_sensibilidad, 'sensibilidad', tmpdir)
                img_comparativo = self._guardar_figura_temp(self.fig_comparativo, 'comparativo', tmpdir)
                img_pagos_cuenta = self._guardar_figura_temp(self.fig_pagos_cuenta, 'pagos_cuenta', tmpdir)
                img_equilibrio_sim = self._guardar_figura_temp(self.fig_equilibrio_sim, 'equilibrio_sim', tmpdir)

                # ========================================================
                # HOJA 1: RESUMEN EJECUTIVO
                # ========================================================
                ws1 = wb.active
                ws1.title = "RESUMEN EJECUTIVO"
                ws1.sheet_properties.tabColor = "1F4E79"

                # T√≠tulo principal
                ws1.merge_cells('B2:H2')
                c = ws1['B2']
                c.value = "REPORTE DE OPTIMIZACION TRIBUTARIA - PRESENTACION GERENCIAL"
                self._aplicar_estilo(c, {
                    'font': Font(name='Calibri', bold=True, color='1F4E79', size=18),
                    'alignment': Alignment(horizontal='center', vertical='center')
                })
                ws1.row_dimensions[2].height = 40

                # Subt√≠tulo
                ws1.merge_cells('B3:H3')
                c = ws1['B3']
                c.value = f"Grupo {r['matriz']['nombre']} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                self._aplicar_estilo(c, {
                    'font': Font(name='Calibri', italic=True, color='666666', size=11),
                    'alignment': Alignment(horizontal='center', vertical='center')
                })

                # L√≠nea separadora
                for col in range(2, 9):
                    c = ws1.cell(row=4, column=col)
                    c.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                ws1.row_dimensions[4].height = 4

                # KPIs principales (fila 6-7)
                kpis = [
                    ("AHORRO TRIBUTARIO", f"S/ {r['grupo']['ahorro_tributario']:,.0f}"),
                    ("AHORRO %", f"{r['grupo']['ahorro_porcentual']:.1f}%"),
                    ("TASA EFECTIVA", f"{tasa_efectiva:.1f}%"),
                    ("ROI ESTRUCTURA", f"{roi:.1f}%"),
                ]

                col_start = 2
                for i, (label, valor) in enumerate(kpis):
                    col = col_start + (i * 2)
                    # Label
                    c = ws1.cell(row=6, column=col)
                    c.value = label
                    self._aplicar_estilo(c, self._estilo_kpi_label())
                    ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col)
                    ws1.column_dimensions[get_column_letter(col)].width = 22
                    # Valor
                    c = ws1.cell(row=7, column=col)
                    c.value = valor
                    self._aplicar_estilo(c, self._estilo_kpi_valor())
                ws1.row_dimensions[6].height = 30
                ws1.row_dimensions[7].height = 35

                # Secci√≥n: Escenario Sin Estructura
                fila = 9
                ws1.merge_cells(f'B{fila}:E{fila}')
                c = ws1[f'B{fila}']
                c.value = "ESCENARIO SIN ESTRUCTURA SATELITAL"
                self._aplicar_estilo(c, self._estilo_subtitulo())
                fila += 1

                datos_sin = [
                    ("Ingresos Totales", f"S/ {r['matriz']['ingresos']:,.2f}"),
                    ("Costos Externos", f"S/ {r['matriz']['costos_externos']:,.2f}"),
                    ("Utilidad Gravable", f"S/ {r['matriz']['utilidad_sin_estructura']:,.2f}"),
                    ("Impuesto (29.5%)", f"S/ {r['matriz']['impuesto_sin_estructura']:,.2f}"),
                ]
                for j, (lab, val) in enumerate(datos_sin):
                    c = ws1.cell(row=fila + j, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws1.cell(row=fila + j, column=4)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))

                # Secci√≥n: Con Estructura
                fila += len(datos_sin) + 1
                ws1.merge_cells(f'B{fila}:E{fila}')
                c = ws1[f'B{fila}']
                c.value = "ESCENARIO CON ESTRUCTURA OPTIMIZADA"
                self._aplicar_estilo(c, self._estilo_subtitulo())
                fila += 1

                datos_con = [
                    ("Impuesto Matriz (con estructura)", f"S/ {r['matriz']['impuesto_con_estructura']:,.2f}"),
                    ("Impuesto Total Satelites", f"S/ {r['grupo']['total_impuesto_satelites']:,.2f}"),
                    ("Impuesto Total Grupo", f"S/ {r['grupo']['impuesto_total']:,.2f}"),
                    ("AHORRO NETO", f"S/ {r['grupo']['ahorro_tributario']:,.2f}"),
                ]
                for j, (lab, val) in enumerate(datos_con):
                    c = ws1.cell(row=fila + j, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws1.cell(row=fila + j, column=4)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                    if j == len(datos_con) - 1:
                        ws1.cell(row=fila + j, column=2).font = Font(name='Calibri', bold=True, color='006600', size=11)
                        ws1.cell(row=fila + j, column=4).font = Font(name='Calibri', bold=True, color='006600', size=11)

                # Seccion: Equilibrio Financiero (si se calculo)
                if self.datos_pagos_cuenta:
                    gpc = self.datos_pagos_cuenta['global']
                    fila += len(datos_con) + 1
                    ws1.merge_cells(f'B{fila}:E{fila}')
                    c = ws1[f'B{fila}']
                    c.value = "EQUILIBRIO FINANCIERO (IR = Pagos a Cuenta, Saldo = 0)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
                    fila += 1

                    datos_eq_res = [
                        ("IR Grupo (Equilibrio)", f"S/ {gpc['ir_grupo_eq']:,.2f}"),
                        ("PaC Grupo (Equilibrio)", f"S/ {gpc['pc_grupo_eq']:,.2f}"),
                        ("Saldo Regularizacion", f"S/ {gpc['saldo_grupo_eq']:,.2f}"),
                        ("Eficiencia PaC", f"{gpc['eficiencia_eq']:.1f}% (ideal=100%)"),
                        ("Tasa Efectiva Grupo", f"{gpc['tasa_efectiva_eq']:.1f}%"),
                        ("AHORRO vs Sin Estructura", f"S/ {gpc['ahorro_eq_total']:,.2f} ({gpc['ahorro_eq_pct']:.1f}%)"),
                    ]
                    for j, (lab, val) in enumerate(datos_eq_res):
                        c = ws1.cell(row=fila + j, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws1.cell(row=fila + j, column=4)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                        if j == len(datos_eq_res) - 1:
                            ws1.cell(row=fila + j, column=2).font = Font(name='Calibri', bold=True, color='2E75B6', size=11)
                            ws1.cell(row=fila + j, column=4).font = Font(name='Calibri', bold=True, color='2E75B6', size=11)

                # ========================================================
                # HOJA 2: DETALLE SATELITES
                # ========================================================
                ws2 = wb.create_sheet("DETALLE SATELITES")
                ws2.sheet_properties.tabColor = "2E75B6"

                ws2.merge_cells('B2:L2')
                c = ws2['B2']
                c.value = "DETALLE POR EMPRESA SATELITE"
                self._aplicar_estilo(c, self._estilo_titulo())
                ws2.row_dimensions[2].height = 30

                # Headers
                headers = ["N", "Empresa", "Tipo", "Regimen", "Costos", "Gastos Op.",
                           "Margen %", "Precio Venta", "Utilidad Neta", "Impuesto",
                           "Ahorro Ind.", "Tasa Efect."]
                for j, h in enumerate(headers, 2):
                    c = ws2.cell(row=4, column=j)
                    c.value = h
                    self._aplicar_estilo(c, self._estilo_header())
                ws2.row_dimensions[4].height = 28

                # Datos
                for i, sat in enumerate(r['satelites']):
                    fila = 5 + i
                    alt = i % 2 == 0
                    datos_fila = [
                        i + 1,
                        sat['nombre'],
                        sat['tipo'],
                        sat['regimen'],
                        sat['costo'],
                        sat['gastos_operativos'],
                        sat['margen_optimo'],
                        sat['precio_venta'],
                        sat['utilidad_neta'],
                        sat['impuesto'],
                        sat['ahorro_individual'],
                        sat['tasa_efectiva']
                    ]
                    for j, val in enumerate(datos_fila, 2):
                        c = ws2.cell(row=fila, column=j)
                        if isinstance(val, (int, float)) and j >= 6:
                            if j == 8:  # Margen %
                                c.value = val
                                c.number_format = '0.00"%"'
                            elif j == 13:  # Tasa efectiva
                                c.value = val
                                c.number_format = '0.00"%"'
                            else:
                                c.value = val
                                c.number_format = '#,##0.00'
                        else:
                            c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(alt))

                # Totales
                fila_total = 5 + len(r['satelites'])
                ws2.cell(row=fila_total, column=2).value = "TOTAL"
                ws2.cell(row=fila_total, column=2).font = Font(bold=True, size=11, color='1F4E79')
                for col_idx, val in [
                    (6, sum(s['costo'] for s in r['satelites'])),
                    (7, sum(s['gastos_operativos'] for s in r['satelites'])),
                    (10, sum(s['utilidad_neta'] for s in r['satelites'])),
                    (11, sum(s['impuesto'] for s in r['satelites'])),
                    (12, sum(s['ahorro_individual'] for s in r['satelites']))
                ]:
                    c = ws2.cell(row=fila_total, column=col_idx)
                    c.value = val
                    c.number_format = '#,##0.00'
                    c.font = Font(bold=True, size=10, color='1F4E79')
                    c.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

                # --- METRICAS DE EQUILIBRIO FINANCIERO POR SATELITE ---
                if self.datos_pagos_cuenta and self.datos_pagos_cuenta.get('satelites'):
                    fila_eq = fila_total + 2
                    ws2.merge_cells(f'B{fila_eq}:Q{fila_eq}')
                    c = ws2[f'B{fila_eq}']
                    c.value = "METRICAS DE EQUILIBRIO FINANCIERO POR SATELITE (IR = PaC, Saldo = 0)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
                    fila_eq += 1

                    headers_eq = ["N", "Empresa", "Regimen Eq.", "Margen Eq.%",
                                  "Precio Eq.", "Utilidad Eq.", "IR Eq.", "PaC Eq.",
                                  "Saldo Eq.", "Ahorro Eq.", "Margen Opt.%", "IR Opt.",
                                  "Saldo Opt.", "Efic.Opt.%"]
                    for j, h in enumerate(headers_eq, 2):
                        c = ws2.cell(row=fila_eq, column=j)
                        c.value = h
                        self._aplicar_estilo(c, self._estilo_header())
                        if "Eq." in h:
                            c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                            c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
                    ws2.row_dimensions[fila_eq].height = 28
                    fila_eq += 1

                    for i, sat_pc in enumerate(self.datos_pagos_cuenta['satelites']):
                        alt = i % 2 == 0
                        tiene_eq = sat_pc['margen_equilibrio'] is not None
                        datos_eq_fila = [
                            i + 1,
                            sat_pc['nombre'],
                            sat_pc.get('regimen_equilibrio', 'N/A') if tiene_eq else 'N/A',
                            sat_pc['margen_equilibrio'] if tiene_eq else 'N/A',
                            sat_pc['precio_equilibrio'],
                            sat_pc['utilidad_neta_eq'],
                            sat_pc['ir_equilibrio'],
                            sat_pc['pago_cuenta_eq'],
                            sat_pc['saldo_equilibrio'],
                            sat_pc['ahorro_equilibrio'],
                            sat_pc['margen_optimo_regimen'],
                            sat_pc['ir_optimo'],
                            sat_pc['saldo_optimo'],
                            sat_pc['eficiencia_optimo'],
                        ]
                        for j, val in enumerate(datos_eq_fila, 2):
                            c = ws2.cell(row=fila_eq, column=j)
                            if isinstance(val, (int, float)):
                                c.value = val
                                if j in (5, 15):  # Margen %
                                    c.number_format = '0.0000"%"'
                                elif j == 15:  # Eficiencia %
                                    c.number_format = '0.00"%"'
                                else:
                                    c.number_format = '#,##0.00'
                            else:
                                c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(alt))
                        fila_eq += 1

                # Ajustar anchos
                anchos = [4, 16, 12, 26, 14, 14, 10, 16, 14, 14, 14, 12]
                for j, w in enumerate(anchos, 2):
                    ws2.column_dimensions[get_column_letter(j)].width = w
                for j in range(14, 17):
                    ws2.column_dimensions[get_column_letter(j)].width = 14

                # ========================================================
                # HOJA 3: GRAFICOS RESULTADOS
                # ========================================================
                ws3 = wb.create_sheet("GRAFICOS RESULTADOS")
                ws3.sheet_properties.tabColor = "2ECC71"

                ws3.merge_cells('B2:H2')
                c = ws3['B2']
                c.value = "VISUALIZACION ESTRATEGICA - ANALISIS TRIBUTARIO"
                self._aplicar_estilo(c, self._estilo_titulo())

                if img_resultados:
                    self._insertar_imagen(ws3, img_resultados, 'B4', ancho_cm=36, alto_cm=22)
                else:
                    ws3['B4'] = "Ejecute 'Calcular Margenes Optimos' para generar los graficos."
                    ws3['B4'].font = Font(italic=True, color='999999', size=11)

                # ========================================================
                # HOJA 4: SIMULACION MONTE CARLO
                # ========================================================
                ws4 = wb.create_sheet("SIMULACION MONTE CARLO")
                ws4.sheet_properties.tabColor = "E74C3C"

                ws4.merge_cells('B2:H2')
                c = ws4['B2']
                c.value = "SIMULACION MONTE CARLO - ANALISIS DE RIESGO"
                self._aplicar_estilo(c, self._estilo_titulo())

                if self.datos_simulacion:
                    ds = self.datos_simulacion

                    # Tabla de estadisticas
                    ws4.merge_cells('B4:C4')
                    c = ws4['B4']
                    c.value = "ESTADISTICAS DE SIMULACION"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    stats_data = [
                        ("Iteraciones", f"{ds['n_sim']:,}"),
                        ("Variabilidad", f"+/-{ds['variabilidad']:.1f}%"),
                        ("Ahorro Medio", f"S/ {ds['media']:,.0f}"),
                        ("Mediana", f"S/ {ds['mediana']:,.0f}"),
                        ("Desv. Estandar", f"S/ {ds['std']:,.0f}"),
                        ("Coef. Variacion", f"{ds['cv']:.2f}%"),
                        ("Minimo", f"S/ {ds['min']:,.0f}"),
                        ("Maximo", f"S/ {ds['max']:,.0f}"),
                        ("Percentil 5", f"S/ {ds['p5']:,.0f}"),
                        ("Percentil 25", f"S/ {ds['p25']:,.0f}"),
                        ("Percentil 75", f"S/ {ds['p75']:,.0f}"),
                        ("Percentil 95", f"S/ {ds['p95']:,.0f}"),
                    ]
                    for i, (lab, val) in enumerate(stats_data):
                        fila = 5 + i
                        c = ws4.cell(row=fila, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws4.cell(row=fila, column=3)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    ws4.column_dimensions['B'].width = 24
                    ws4.column_dimensions['C'].width = 28

                    # Bootstrap IC
                    fila_b = 5 + len(stats_data) + 1
                    ws4.merge_cells(f'B{fila_b}:C{fila_b}')
                    c = ws4[f'B{fila_b}']
                    c.value = f"INTERVALO DE CONFIANZA - {ds.get('ic_metodo', 'Bootstrap Percentil')}"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    ic_data = [
                        ("Metodo", ds.get('ic_metodo', 'Bootstrap Percentil')),
                        ("N Bootstrap", f"{ds.get('n_bootstrap', 5000):,}"),
                        (f"IC {ds['conf_nivel']:.0f}% Inferior", f"S/ {ds['ic_lower']:,.0f}"),
                        (f"IC {ds['conf_nivel']:.0f}% Superior", f"S/ {ds['ic_upper']:,.0f}"),
                    ]
                    for i, (lab, val) in enumerate(ic_data):
                        f = fila_b + 1 + i
                        c = ws4.cell(row=f, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws4.cell(row=f, column=3)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Prueba de hipotesis
                    fila_h = fila_b + 1 + len(ic_data) + 1
                    ws4.merge_cells(f'B{fila_h}:C{fila_h}')
                    c = ws4[f'B{fila_h}']
                    c.value = "PRUEBA DE HIPOTESIS (No Parametrica Bootstrap)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    hipotesis_data = [
                        ("H0", "Ahorro medio = 0"),
                        ("H1", "Ahorro medio > 0"),
                        ("P-Valor (Bootstrap)", f"{ds['p_value']:.2e}"),
                        ("t-stat (referencia)", f"{ds['t_stat']:.4f}"),
                        ("P-Valor t (referencia)", f"{ds.get('p_value_t', ds['p_value']):.2e}"),
                        ("Alfa", f"{ds['alfa']}"),
                        ("Nivel Confianza", f"{ds['conf_nivel']:.1f}%"),
                        ("DECISION", ds['decision']),
                    ]
                    for i, (lab, val) in enumerate(hipotesis_data):
                        f = fila_h + 1 + i
                        c = ws4.cell(row=f, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws4.cell(row=f, column=3)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        if lab == "DECISION":
                            c.font = Font(name='Calibri', bold=True, color='006600', size=11)

                    # Ajuste Log-Normal
                    if ds.get('ln_ajustado'):
                        fila_ln = fila_h + 1 + len(hipotesis_data) + 1
                        ws4.merge_cells(f'B{fila_ln}:C{fila_ln}')
                        c = ws4[f'B{fila_ln}']
                        c.value = "AJUSTE DISTRIBUCION LOG-NORMAL"
                        self._aplicar_estilo(c, self._estilo_subtitulo())
                        ln_data = [
                            ("Distribucion", "Log-Normal"),
                            ("Shape (s)", f"{ds['ln_shape']:.4f}"),
                            ("Location", f"{ds['ln_loc']:.4f}"),
                            ("Scale", f"S/ {ds['ln_scale']:,.0f}"),
                        ]
                        for i, (lab, val) in enumerate(ln_data):
                            f = fila_ln + 1 + i
                            c = ws4.cell(row=f, column=2)
                            c.value = lab
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            c.alignment = Alignment(horizontal='left')
                            c = ws4.cell(row=f, column=3)
                            c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Notas de configuracion matriz
                    if ds.get('var_matriz_costos') or ds.get('var_matriz_ingresos'):
                        if ds.get('ln_ajustado'):
                            fila_nota = fila_ln + 1 + len(ln_data) + 1
                        else:
                            fila_nota = fila_h + 1 + len(hipotesis_data) + 1
                        ws4.merge_cells(f'B{fila_nota}:C{fila_nota}')
                        c = ws4[f'B{fila_nota}']
                        c.value = "CONFIGURACION VARIABILIDAD MATRIZ"
                        self._aplicar_estilo(c, self._estilo_subtitulo())
                        mat_notas = []
                        if ds.get('var_matriz_costos'):
                            mat_notas.append(("Var. Costos Matriz", f"+/-{ds['pct_var_costos']:.1f}%"))
                        if ds.get('var_matriz_ingresos'):
                            mat_notas.append(("Var. Ingresos Matriz", f"+/-{ds['pct_var_ingresos']:.1f}%"))
                        mat_notas.append(("Distribucion", ds.get('distribucion', 'Log-normal')))
                        mat_notas.append(("Correlacion (rho)", f"{ds.get('rho', 0):.2f}"))
                        for i, (lab, val) in enumerate(mat_notas):
                            f = fila_nota + 1 + i
                            c = ws4.cell(row=f, column=2)
                            c.value = lab
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            c.alignment = Alignment(horizontal='left')
                            c = ws4.cell(row=f, column=3)
                            c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Grafico
                    if img_simulacion:
                        self._insertar_imagen(ws4, img_simulacion, 'E4', ancho_cm=36, alto_cm=22)
                else:
                    ws4['B4'] = "Ejecute la Simulacion Monte Carlo para generar datos y graficos."
                    ws4['B4'].font = Font(italic=True, color='999999', size=11)

                # ========================================================
                # HOJA 5: ANALISIS SENSIBILIDAD
                # ========================================================
                ws5 = wb.create_sheet("SENSIBILIDAD")
                ws5.sheet_properties.tabColor = "F39C12"

                ws5.merge_cells('B2:H2')
                c = ws5['B2']
                c.value = "ANALISIS DE SENSIBILIDAD"
                self._aplicar_estilo(c, self._estilo_titulo())

                ws5.merge_cells('B4:H4')
                c = ws5['B4']
                c.value = "Impacto de variaciones en parametros clave sobre el ahorro tributario"
                self._aplicar_estilo(c, {
                    'font': Font(name='Calibri', italic=True, color='666666', size=10),
                    'alignment': Alignment(horizontal='left')
                })

                if img_sensibilidad:
                    self._insertar_imagen(ws5, img_sensibilidad, 'B6', ancho_cm=36, alto_cm=16)
                else:
                    ws5['B6'] = "Ejecute el Analisis de Sensibilidad para generar graficos."
                    ws5['B6'].font = Font(italic=True, color='999999', size=11)

                # ========================================================
                # HOJA 6: ANALISIS COMPARATIVO
                # ========================================================
                ws6 = wb.create_sheet("COMPARATIVO")
                ws6.sheet_properties.tabColor = "9B59B6"

                ws6.merge_cells('B2:H2')
                c = ws6['B2']
                c.value = "ANALISIS COMPARATIVO MULTI-ESCENARIO"
                self._aplicar_estilo(c, self._estilo_titulo())

                if hasattr(self, 'datos_comparativo') and self.datos_comparativo:
                    # Tabla comparativa
                    headers_comp = ["Escenario", "Ahorro Tributario", "Tasa Efectiva %", "N Satelites"]
                    for j, h in enumerate(headers_comp, 2):
                        c = ws6.cell(row=4, column=j)
                        c.value = h
                        self._aplicar_estilo(c, self._estilo_header())
                    ws6.row_dimensions[4].height = 28
                    ws6.column_dimensions['B'].width = 20
                    ws6.column_dimensions['C'].width = 22
                    ws6.column_dimensions['D'].width = 18
                    ws6.column_dimensions['E'].width = 14

                    for i, esc in enumerate(self.datos_comparativo):
                        fila = 5 + i
                        alt = i % 2 == 0
                        vals = [
                            esc['nombre'],
                            esc['ahorro'],
                            esc['tasa_efectiva'],
                            esc['satelites']
                        ]
                        for j, val in enumerate(vals, 2):
                            c = ws6.cell(row=fila, column=j)
                            if j == 3:
                                c.value = val
                                c.number_format = '#,##0.00'
                            elif j == 4:
                                c.value = val
                                c.number_format = '0.00"%"'
                            else:
                                c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(alt))

                if img_comparativo:
                    fila_img = 5 + (len(self.datos_comparativo) if hasattr(self, 'datos_comparativo') and self.datos_comparativo else 0) + 2
                    self._insertar_imagen(ws6, img_comparativo, f'B{fila_img}', ancho_cm=36, alto_cm=16)
                else:
                    ws6['B10'] = "Ejecute el Analisis Comparativo para generar graficos."
                    ws6['B10'].font = Font(italic=True, color='999999', size=11)

                # ========================================================
                # HOJA 7: PAGOS A CUENTA Y SENSIBILIDAD
                # ========================================================
                ws7 = wb.create_sheet("PAGOS A CUENTA")
                ws7.sheet_properties.tabColor = "E67E22"

                ws7.merge_cells('B2:K2')
                c = ws7['B2']
                c.value = "ANALISIS DE PAGOS A CUENTA Y EQUILIBRIO FINANCIERO"
                self._aplicar_estilo(c, self._estilo_titulo())
                ws7.row_dimensions[2].height = 30

                if self.datos_pagos_cuenta:
                    dpc = self.datos_pagos_cuenta
                    gpc = dpc['global']

                    # --- EXPLICACION GERENCIAL DEL CONCEPTO ---
                    ws7.merge_cells('B4:K4')
                    c = ws7['B4']
                    c.value = "CONCEPTO: PUNTO DE EQUILIBRIO FINANCIERO (IR = Pagos a Cuenta, Saldo = 0)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    ws7.merge_cells('B5:K5')
                    c = ws7['B5']
                    c.value = ("El punto de equilibrio financiero es el margen donde los pagos a cuenta mensuales (adelantos obligatorios del IR al "
                              f"{dpc['tasa_pago_cuenta']:.1f}% de los ingresos) igualan EXACTAMENTE al Impuesto a la Renta anual. "
                              "En este punto: saldo de regularizacion = 0, eficiencia = 100%, flujo de caja optimizado. "
                              "No hay dinero inmovilizado en SUNAT ni necesidad de liquidez adicional para regularizacion.")
                    self._aplicar_estilo(c, {
                        'font': Font(name='Calibri', italic=True, color='444444', size=10),
                        'alignment': Alignment(horizontal='left', wrap_text=True)
                    })
                    ws7.row_dimensions[5].height = 50

                    # --- SECCION I: ESCENARIO SIN ESTRUCTURA ---
                    fila = 7
                    ws7.merge_cells(f'B{fila}:F{fila}')
                    c = ws7[f'B{fila}']
                    c.value = f"I. ESCENARIO SIN ESTRUCTURA SATELITAL - {gpc['nombre_matriz']}"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila += 1

                    datos_sin_est = [
                        ("Ingresos Totales", gpc['ingresos_matriz']),
                        ("Utilidad Gravable", gpc['utilidad_sin_estructura']),
                        ("Tasa Aplicable", "29.5% (Regimen General)"),
                        ("Impuesto a la Renta", gpc['impuesto_sin_estructura']),
                        (f"Pagos a Cuenta ({dpc['tasa_pago_cuenta']:.1f}%)", gpc['pc_matriz_sin']),
                        ("Saldo Regularizacion", gpc['saldo_matriz_sin']),
                        ("Eficiencia PaC", f"{gpc['eficiencia_sin']:.1f}%"),
                        ("Tasa Efectiva", f"{gpc['tasa_efectiva_sin']:.1f}%"),
                    ]
                    for i, (lab, val) in enumerate(datos_sin_est):
                        c = ws7.cell(row=fila + i, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws7.cell(row=fila + i, column=3)
                        if isinstance(val, (int, float)):
                            c.value = val
                            c.number_format = '#,##0.00'
                        else:
                            c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        if lab == "Saldo Regularizacion":
                            color_s = 'CC0000' if gpc['saldo_matriz_sin'] < 0 else '006600'
                            c.font = Font(name='Calibri', bold=True, color=color_s, size=10)
                            c2 = ws7.cell(row=fila + i, column=4)
                            c2.value = "(FAVOR)" if gpc['saldo_matriz_sin'] < 0 else "(POR PAGAR)"
                            c2.font = Font(name='Calibri', bold=True, color=color_s, size=9)

                    fila += len(datos_sin_est) + 1

                    # --- SECCION II: GRUPO EN PUNTO DE EQUILIBRIO ---
                    ws7.merge_cells(f'B{fila}:F{fila}')
                    c = ws7[f'B{fila}']
                    c.value = "II. GRUPO CONSOLIDADO EN PUNTO DE EQUILIBRIO (IR = PaC, Saldo = 0)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila += 1

                    datos_eq = [
                        ("--- MATRIZ ---", ""),
                        ("Utilidad Matriz (ajustada)", gpc['utilidad_matriz_eq']),
                        ("Impuesto Matriz", gpc['impuesto_matriz_eq']),
                        ("PaC Matriz", gpc['pc_matriz_eq']),
                        ("Saldo Matriz", gpc['saldo_matriz_eq']),
                        ("--- SATELITES EN EQUILIBRIO ---", ""),
                        ("IR Satelites (equilibrio)", gpc['total_ir_eq_sats']),
                        ("PaC Satelites (equilibrio)", gpc['total_pc_eq_sats']),
                        ("Saldo Satelites", gpc['total_saldo_eq_sats']),
                        ("--- GRUPO CONSOLIDADO ---", ""),
                        ("IR Total Grupo", gpc['ir_grupo_eq']),
                        ("PaC Total Grupo", gpc['pc_grupo_eq']),
                        ("Saldo Grupo", gpc['saldo_grupo_eq']),
                        ("Eficiencia PaC", f"{gpc['eficiencia_eq']:.1f}%"),
                        ("Tasa Efectiva", f"{gpc['tasa_efectiva_eq']:.1f}%"),
                        ("Ahorro vs Sin Estructura", gpc['ahorro_eq_total']),
                        ("Ahorro %", f"{gpc['ahorro_eq_pct']:.1f}%"),
                    ]
                    for i, (lab, val) in enumerate(datos_eq):
                        c = ws7.cell(row=fila + i, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        if lab.startswith("---"):
                            c.font = Font(name='Calibri', bold=True, color='1F4E79', size=10)
                        c = ws7.cell(row=fila + i, column=3)
                        if isinstance(val, (int, float)):
                            c.value = val
                            c.number_format = '#,##0.00'
                        else:
                            c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        if "Saldo" in lab and isinstance(val, (int, float)):
                            color_s = '2E75B6' if abs(val) < 100 else ('CC0000' if val < 0 else '006600')
                            c.font = Font(name='Calibri', bold=True, color=color_s, size=10)
                            c2 = ws7.cell(row=fila + i, column=4)
                            c2.value = "(EQUILIBRIO)" if abs(val) < 100 else ("(FAVOR)" if val < 0 else "(PAGAR)")
                            c2.font = Font(name='Calibri', bold=True, color=color_s, size=9)
                        if lab == "Ahorro vs Sin Estructura":
                            c.font = Font(name='Calibri', bold=True, color='006600', size=11)

                    fila += len(datos_eq) + 1

                    # --- SECCION III: COMPARATIVO 3 ESCENARIOS ---
                    ws7.merge_cells(f'B{fila}:H{fila}')
                    c = ws7[f'B{fila}']
                    c.value = "III. COMPARATIVO: SIN ESTRUCTURA vs EQUILIBRIO vs MARGEN OPTIMO (ref)"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila += 1

                    headers_comp_eq = ["Concepto", "Sin Estructura", "EQUILIBRIO", "M. Optimo (ref)", "Mejor"]
                    for j, h in enumerate(headers_comp_eq, 2):
                        c = ws7.cell(row=fila, column=j)
                        c.value = h
                        self._aplicar_estilo(c, self._estilo_header())
                        if h == "EQUILIBRIO":
                            c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                    ws7.row_dimensions[fila].height = 28
                    fila += 1

                    filas_comp = [
                        ("IR Total Grupo", gpc['impuesto_sin_estructura'], gpc['ir_grupo_eq'], gpc['ir_grupo_opt'],
                         "Equilibrio" if gpc['ir_grupo_eq'] <= gpc['ir_grupo_opt'] else "M. Optimo"),
                        ("PaC Total Grupo", gpc['pc_matriz_sin'], gpc['pc_grupo_eq'], gpc['pc_grupo_opt'], "-"),
                        ("Saldo Regularizacion", gpc['saldo_matriz_sin'], gpc['saldo_grupo_eq'], gpc['saldo_grupo_opt'],
                         "Equilibrio" if abs(gpc['saldo_grupo_eq']) < abs(gpc['saldo_grupo_opt']) else "M. Optimo"),
                        ("Ahorro Tributario", 0, gpc['ahorro_eq_total'], gpc['ahorro_opt_total'],
                         "Equilibrio" if gpc['ahorro_eq_total'] >= gpc['ahorro_opt_total'] else "M. Optimo"),
                    ]
                    for i, (concepto, v_sin, v_eq, v_opt, mejor) in enumerate(filas_comp):
                        alt = i % 2 == 0
                        c = ws7.cell(row=fila, column=2)
                        c.value = concepto
                        self._aplicar_estilo(c, self._estilo_celda(alt))
                        c.alignment = Alignment(horizontal='left')
                        for j, val in enumerate([v_sin, v_eq, v_opt], 3):
                            c = ws7.cell(row=fila, column=j)
                            c.value = val
                            c.number_format = '#,##0.00'
                            self._aplicar_estilo(c, self._estilo_celda(alt))
                            if j == 4:  # Columna equilibrio destacada
                                c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                        c = ws7.cell(row=fila, column=6)
                        c.value = mejor
                        self._aplicar_estilo(c, self._estilo_celda(alt))
                        if "Equilibrio" in mejor:
                            c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                        fila += 1

                    # Filas de porcentaje
                    for i, (concepto, v_sin, v_eq, v_opt, mejor) in enumerate([
                        ("Eficiencia PaC (%)", gpc['eficiencia_sin'], gpc['eficiencia_eq'], gpc['eficiencia_opt'],
                         "Equilibrio" if abs(gpc['eficiencia_eq'] - 100) <= abs(gpc['eficiencia_opt'] - 100) else "M. Optimo"),
                        ("Tasa Efectiva (%)", gpc['tasa_efectiva_sin'], gpc['tasa_efectiva_eq'], gpc['tasa_efectiva_opt'],
                         "Equilibrio" if gpc['tasa_efectiva_eq'] <= gpc['tasa_efectiva_opt'] else "M. Optimo"),
                    ]):
                        alt = (len(filas_comp) + i) % 2 == 0
                        c = ws7.cell(row=fila, column=2)
                        c.value = concepto
                        self._aplicar_estilo(c, self._estilo_celda(alt))
                        c.alignment = Alignment(horizontal='left')
                        for j, val in enumerate([v_sin, v_eq, v_opt], 3):
                            c = ws7.cell(row=fila, column=j)
                            c.value = val
                            c.number_format = '0.0"%"'
                            self._aplicar_estilo(c, self._estilo_celda(alt))
                            if j == 4:
                                c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                        c = ws7.cell(row=fila, column=6)
                        c.value = mejor
                        self._aplicar_estilo(c, self._estilo_celda(alt))
                        if "Equilibrio" in mejor:
                            c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                        fila += 1

                    # KPI
                    fila += 1
                    kpis_eq = [
                        ("AHORRO EN EQUILIBRIO", f"S/ {gpc['ahorro_eq_total']:,.0f} ({gpc['ahorro_eq_pct']:.1f}%)"),
                        ("EFICIENCIA PaC", f"{gpc['eficiencia_eq']:.1f}%"),
                    ]
                    for i, (lab, val) in enumerate(kpis_eq):
                        col_k = 2 + i * 3
                        c = ws7.cell(row=fila, column=col_k)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_kpi_label())
                        c = ws7.cell(row=fila + 1, column=col_k)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_kpi_valor())
                    ws7.row_dimensions[fila].height = 30
                    ws7.row_dimensions[fila + 1].height = 35
                    fila += 3

                    # --- SECCION IV: INTERPRETACION GERENCIAL ---
                    ws7.merge_cells(f'B{fila}:K{fila}')
                    c = ws7[f'B{fila}']
                    c.value = "IV. INTERPRETACION GERENCIAL - ESTRATEGIA DE EQUILIBRIO FINANCIERO"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila += 1

                    interp_lines = [
                        f"PUNTO DE EQUILIBRIO: Cada satelite opera al margen donde IR = Pagos a Cuenta. "
                        f"El saldo de regularizacion es cero: no hay dinero inmovilizado en SUNAT ni falta liquidez.",

                        f"SIN ESTRUCTURA: {gpc['nombre_matriz']} paga IR de S/ {gpc['impuesto_sin_estructura']:,.0f} (29.5%). "
                        f"Eficiencia PaC: {gpc['eficiencia_sin']:.1f}%. "
                        + ("Saldo a favor inmovilizado." if gpc['saldo_matriz_sin'] < 0 else "Debe regularizar saldo."),

                        f"EN EQUILIBRIO: El grupo paga IR de S/ {gpc['ir_grupo_eq']:,.0f} (tasa efectiva {gpc['tasa_efectiva_eq']:.1f}%). "
                        f"Eficiencia PaC: {gpc['eficiencia_eq']:.1f}%. Ahorro: S/ {gpc['ahorro_eq_total']:,.0f} ({gpc['ahorro_eq_pct']:.1f}%).",

                        f"REFERENCIA (Margen Optimo): Maximiza ahorro fiscal a S/ {gpc['ahorro_opt_total']:,.0f} ({gpc['ahorro_opt_pct']:.1f}%), "
                        f"pero con eficiencia PaC de {gpc['eficiencia_opt']:.1f}% y saldo de regularizacion de S/ {gpc['saldo_grupo_opt']:,.0f}.",

                        "CONCLUSION: El enfoque de equilibrio financiero prioriza que IR = PaC en cada satelite, "
                        "logrando saldo cero y flujo de caja optimo. La gerencia puede elegir entre maximizar el ahorro "
                        "fiscal (margen optimo) o maximizar la eficiencia financiera (punto de equilibrio)."
                    ]
                    for line in interp_lines:
                        ws7.merge_cells(f'B{fila}:K{fila}')
                        c = ws7[f'B{fila}']
                        c.value = line
                        c.font = Font(name='Calibri', size=10, color='333333')
                        c.alignment = Alignment(horizontal='left', wrap_text=True)
                        ws7.row_dimensions[fila].height = 40
                        fila += 1
                    ws7[f'B{fila - 1}'].font = Font(name='Calibri', size=10, color='1F4E79', bold=True)

                    fila += 1

                    ws7.column_dimensions['B'].width = 30
                    ws7.column_dimensions['C'].width = 22
                    ws7.column_dimensions['D'].width = 22
                    ws7.column_dimensions['E'].width = 22
                    ws7.column_dimensions['F'].width = 18

                    # --- SECCION V: DETALLE POR SATELITE (equilibrio primero) ---
                    ws7.merge_cells(f'B{fila}:K{fila}')
                    c = ws7[f'B{fila}']
                    c.value = "V. DETALLE POR SATELITE - PUNTO DE EQUILIBRIO FINANCIERO"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila += 1

                    fila_base = fila
                    for idx_sat, sat_pc in enumerate(dpc['satelites']):
                        ws7.merge_cells(f'B{fila_base}:K{fila_base}')
                        c = ws7[f'B{fila_base}']
                        c.value = f"{idx_sat + 1}. {sat_pc['nombre']} | Costo: {sat_pc['costo']:,.0f} | Gastos: {sat_pc['gastos']:,.0f}"
                        self._aplicar_estilo(c, self._estilo_subtitulo())
                        fila_base += 1

                        tiene_eq = sat_pc['margen_equilibrio'] is not None
                        eq_str = f"{sat_pc['margen_equilibrio']:.4f}%" if tiene_eq else "N/A"
                        reg_eq_str = f" [{sat_pc['regimen_equilibrio']}]" if tiene_eq else ""

                        info_sat = [
                            (">>> PUNTO DE EQUILIBRIO (IR=PaC) <<<", ""),
                            ("Margen Equilibrio", f"{eq_str}{reg_eq_str}"),
                            ("Precio Venta en Equilibrio", sat_pc['precio_equilibrio']),
                            ("Utilidad Neta en Equilibrio", sat_pc['utilidad_neta_eq']),
                            ("IR = PaC en Equilibrio", sat_pc['ir_equilibrio']),
                            ("Saldo en Equilibrio", sat_pc['saldo_equilibrio']),
                            ("Ahorro vs Reg. General", sat_pc['ahorro_equilibrio']),
                            ("--- Referencia: Margen Optimo ---", ""),
                            ("Margen Optimo", f"{sat_pc['margen_optimo_regimen']:.4f}%"),
                            ("IR en Optimo", sat_pc['ir_optimo']),
                            ("PaC en Optimo", sat_pc['pago_cuenta_optimo']),
                            ("Saldo en Optimo", sat_pc['saldo_optimo']),
                            ("Eficiencia PaC en Optimo", f"{sat_pc['eficiencia_optimo']:.1f}%"),
                        ]
                        for i, (lab, val) in enumerate(info_sat):
                            c = ws7.cell(row=fila_base + i, column=2)
                            c.value = lab
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            c.alignment = Alignment(horizontal='left')
                            if lab.startswith(">>>"):
                                c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                            elif lab.startswith("---"):
                                c.font = Font(name='Calibri', bold=True, color='888888', size=9)
                            c = ws7.cell(row=fila_base + i, column=3)
                            if isinstance(val, (int, float)):
                                c.value = val
                                c.number_format = '#,##0.00'
                            else:
                                c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            if lab == "Saldo en Equilibrio" and isinstance(val, (int, float)):
                                c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                                c2 = ws7.cell(row=fila_base + i, column=4)
                                c2.value = "(~0 EQUILIBRIO)" if abs(val) < 1 else ""
                                c2.font = Font(name='Calibri', bold=True, color='2E75B6', size=9)

                        fila_base += len(info_sat) + 1

                        # Tabla de sensibilidad
                        ws7.merge_cells(f'B{fila_base}:K{fila_base}')
                        c = ws7[f'B{fila_base}']
                        c.value = "Sensibilidad del Margen - EQUILIBRIO es donde Saldo = 0"
                        self._aplicar_estilo(c, {
                            'font': Font(name='Calibri', bold=True, color='2E75B6', size=10),
                            'alignment': Alignment(horizontal='left')
                        })
                        fila_base += 1

                        headers_sens = ["Margen%", "Precio Venta", "Util. Neta", "Regimen",
                                       "IR", "Pago Cta", "Saldo", "Ahorro Reg.", "Efic%", "Nota"]
                        for j, h in enumerate(headers_sens, 2):
                            c = ws7.cell(row=fila_base, column=j)
                            c.value = h
                            self._aplicar_estilo(c, self._estilo_header())
                        ws7.row_dimensions[fila_base].height = 24
                        fila_base += 1

                        for i, s in enumerate(sat_pc['sensibilidad']):
                            nota = "<<< EQUILIBRIO" if s['es_equilibrio'] else ("(ref: optimo)" if s['es_optimo'] else "")
                            datos_fila = [
                                s['margen'], s['precio_venta'], s['utilidad_neta'],
                                s['regimen'], s['ir'], s['pago_cuenta'],
                                s['saldo'], s['ahorro_regimen'], s['eficiencia_pc'], nota
                            ]
                            for j, val in enumerate(datos_fila, 2):
                                c = ws7.cell(row=fila_base, column=j)
                                if isinstance(val, (int, float)):
                                    c.value = val
                                    if j == 2:
                                        c.number_format = '0.0000"%"'
                                    elif j == 10:
                                        c.number_format = '0.0"%"'
                                    else:
                                        c.number_format = '#,##0.00'
                                else:
                                    c.value = val
                                estilo = self._estilo_celda(i % 2 == 0)
                                self._aplicar_estilo(c, estilo)
                                # Equilibrio = linea principal destacada en azul
                                if s['es_equilibrio']:
                                    c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                                    c.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                                elif s['es_optimo']:
                                    c.font = Font(name='Calibri', color='888888', size=9)
                            fila_base += 1

                        # Sensibilidad 2D: equilibrio ante cambios en tasa PaC
                        fila_base += 1
                        ws7.merge_cells(f'B{fila_base}:J{fila_base}')
                        c = ws7[f'B{fila_base}']
                        c.value = "Punto de Equilibrio ante cambios en la Tasa de Pago a Cuenta"
                        self._aplicar_estilo(c, {
                            'font': Font(name='Calibri', bold=True, color='2E75B6', size=10),
                            'alignment': Alignment(horizontal='left')
                        })
                        fila_base += 1

                        headers_2d = ["Tasa PaC%", "M. Equil%", "Precio Eq.", "Util. Neta",
                                     "IR = PaC", "Saldo", "Ahorro Reg.", "Regimen"]
                        for j, h in enumerate(headers_2d, 2):
                            c = ws7.cell(row=fila_base, column=j)
                            c.value = h
                            self._aplicar_estilo(c, self._estilo_header())
                            if h in ("M. Equil%", "IR = PaC", "Saldo"):
                                c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                        fila_base += 1

                        for i, s2 in enumerate(sat_pc['sensibilidad_2d']):
                            datos_2d = [s2['tasa_pc'], s2['margen_equilibrio'], s2['precio_venta'],
                                       s2['utilidad_neta'], s2['ir'], s2['saldo'],
                                       s2['ahorro_regimen'], s2['regimen']]
                            for j, val in enumerate(datos_2d, 2):
                                c = ws7.cell(row=fila_base, column=j)
                                if isinstance(val, (int, float)):
                                    c.value = val
                                    if j in (2, 3):
                                        c.number_format = '0.0000"%"' if j == 3 else '0.0"%"'
                                    else:
                                        c.number_format = '#,##0.00'
                                else:
                                    c.value = val
                                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            fila_base += 1

                        fila_base += 2

                    # Ajustar anchos para las columnas de sensibilidad
                    for col_idx in range(2, 12):
                        ws7.column_dimensions[get_column_letter(col_idx)].width = max(
                            ws7.column_dimensions[get_column_letter(col_idx)].width or 0, 14)

                    # Grafico pagos a cuenta
                    if img_pagos_cuenta:
                        self._insertar_imagen(ws7, img_pagos_cuenta, f'B{fila_base}', ancho_cm=36, alto_cm=20)
                else:
                    ws7['B4'] = "Ejecute el Analisis de Pagos a Cuenta para generar datos y graficos."
                    ws7['B4'].font = Font(italic=True, color='999999', size=11)

                # ========================================================
                # HOJA 8: SENSIBILIDAD MARGENES POR SATELITE
                # ========================================================
                ws8 = wb.create_sheet("SENSIBILIDAD MARGENES")
                ws8.sheet_properties.tabColor = "1ABC9C"

                ws8.merge_cells('B2:K2')
                c = ws8['B2']
                c.value = "ANALISIS DE SENSIBILIDAD - MARGENES POR SATELITE"
                self._aplicar_estilo(c, self._estilo_titulo())
                ws8.row_dimensions[2].height = 30

                ws8.merge_cells('B3:K3')
                c = ws8['B3']
                c.value = ("Muestra como se diluye el ahorro fiscal al alejarse del margen optimo. "
                          "Desde la perspectiva de teoria de juegos, el punto optimo es el equilibrio de Nash "
                          "donde ningun satelite puede mejorar su posicion unilateralmente.")
                c.font = Font(name='Calibri', italic=True, color='666666', size=9)
                c.alignment = Alignment(wrap_text=True)
                ws8.row_dimensions[3].height = 30

                fila_sens = 5

                for idx_sat, sat in enumerate(r['satelites']):
                    costo_s = sat['costo']
                    gastos_s = sat['gastos_operativos']
                    margen_opt = sat['margen_optimo'] / 100
                    es_general_s = sat['regimen'].startswith('GENERAL')

                    ws8.merge_cells(f'B{fila_sens}:K{fila_sens}')
                    c = ws8[f'B{fila_sens}']
                    c.value = f"{idx_sat + 1}. {sat['nombre']} - Margen Optimo: {sat['margen_optimo']:.4f}%"
                    self._aplicar_estilo(c, self._estilo_subtitulo())
                    fila_sens += 1

                    # Generar variaciones de margen
                    tasa_general_dec = r['parametros']['tasa_general'] / 100
                    tasa_especial_dec = r['parametros']['tasa_especial'] / 100
                    limite_util_val = r['parametros']['limite_utilidad']
                    limite_ing_val = r['parametros']['limite_ingresos']
                    tasa_pc_dec = self.TASA_PAGO_CUENTA.get() / 100

                    m_min = max(0.001, margen_opt * 0.2)
                    m_max_f = min(limite_ing_val / costo_s - 1, 1.0) if costo_s > 0 else 1.0
                    m_max = min(margen_opt * 3.0, m_max_f)
                    if m_max <= m_min:
                        m_max = m_min + 0.10

                    margenes_v = np.linspace(m_min, m_max, 15)
                    if margen_opt not in margenes_v:
                        margenes_v = np.sort(np.append(margenes_v, margen_opt))

                    headers_s = ["Margen%", "Precio", "Util. Neta", "Regimen", "IR",
                                "Pago Cta", "Saldo Reg.", "Ahorro Reg.", "Ahorro vs Base", "Nota"]
                    for j, h in enumerate(headers_s, 2):
                        c = ws8.cell(row=fila_sens, column=j)
                        c.value = h
                        self._aplicar_estilo(c, self._estilo_header())
                    ws8.row_dimensions[fila_sens].height = 24
                    fila_sens += 1

                    # Ahorro base para comparar
                    ahorro_base_sat = sat['ahorro_individual']

                    for i, m in enumerate(margenes_v):
                        precio_v = costo_s * (1 + m)
                        util_b = costo_s * m
                        util_n = max(0, util_b - gastos_s)

                        if es_general_s:
                            ir_v = util_n * tasa_general_dec
                            reg_v = "General"
                            ahorro_reg = 0
                        elif util_n > limite_util_val or precio_v > limite_ing_val:
                            if util_n > limite_util_val:
                                ir_v = limite_util_val * tasa_especial_dec + (util_n - limite_util_val) * tasa_general_dec
                                reg_v = "Mixto"
                            else:
                                ir_v = util_n * tasa_general_dec
                                reg_v = "General*"
                            ahorro_reg = util_n * tasa_general_dec - ir_v
                        else:
                            ir_v = util_n * tasa_especial_dec
                            reg_v = "Especial"
                            ahorro_reg = util_n * (tasa_general_dec - tasa_especial_dec)

                        pac_v = tasa_pc_dec * precio_v
                        saldo_v = ir_v - pac_v
                        diff_vs_base = ahorro_reg - ahorro_base_sat
                        es_opt = abs(m - margen_opt) < 0.0001
                        nota_v = "OPTIMO" if es_opt else ""

                        datos_v = [m * 100, precio_v, util_n, reg_v, ir_v, pac_v,
                                  saldo_v, ahorro_reg, diff_vs_base, nota_v]
                        for j, val in enumerate(datos_v, 2):
                            c = ws8.cell(row=fila_sens, column=j)
                            if isinstance(val, (int, float)):
                                c.value = val
                                if j == 2:
                                    c.number_format = '0.0000"%"'
                                else:
                                    c.number_format = '#,##0.00'
                            else:
                                c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            if es_opt:
                                c.font = Font(name='Calibri', bold=True, color='006600', size=10)
                            if j == 10 and isinstance(val, (int, float)) and val < 0:
                                c.font = Font(name='Calibri', color='CC0000', size=10)
                        fila_sens += 1

                    fila_sens += 2

                for col_idx in range(2, 12):
                    ws8.column_dimensions[get_column_letter(col_idx)].width = max(
                        ws8.column_dimensions[get_column_letter(col_idx)].width or 0, 14)

                # ========================================================
                # HOJA 9: PARAMETROS
                # ========================================================
                ws9 = wb.create_sheet("PARAMETROS")
                ws9.sheet_properties.tabColor = "95A5A6"

                ws9.merge_cells('B2:D2')
                c = ws9['B2']
                c.value = "PARAMETROS TRIBUTARIOS UTILIZADOS"
                self._aplicar_estilo(c, self._estilo_titulo())

                params = [
                    ("UIT", f"S/ {r['parametros']['uit']:,.2f}"),
                    ("Tasa Regimen General", f"{r['parametros']['tasa_general']:.2f}%"),
                    ("Tasa Regimen Especial", f"{r['parametros']['tasa_especial']:.2f}%"),
                    ("Diferencial de Tasas", f"{r['parametros']['tasa_general'] - r['parametros']['tasa_especial']:.2f} pp"),
                    ("Tasa Pago a Cuenta IR", f"{self.TASA_PAGO_CUENTA.get():.2f}%"),
                    ("Limite Utilidad Especial", f"S/ {r['parametros']['limite_utilidad']:,.2f}"),
                    ("Limite Ingresos Especial", f"S/ {r['parametros']['limite_ingresos']:,.2f}"),
                    ("N Empresas Satelites", f"{len(r['satelites'])}"),
                    ("Empresa Matriz", r['matriz']['nombre']),
                    ("", ""),
                    ("--- SIMULACION: VARIABILIDAD MATRIZ ---", ""),
                    ("Variabilidad costos matriz (+/-%)", f"{self.PCT_VAR_MATRIZ_COSTOS.get():.1f}%" if self.VAR_MATRIZ_COSTOS.get() else "Desactivada"),
                    ("Variabilidad ingresos matriz (+/-%)", f"{self.PCT_VAR_MATRIZ_INGRESOS.get():.1f}%" if self.VAR_MATRIZ_INGRESOS.get() else "Desactivada"),
                    ("Distribucion factores", self.DISTRIBUCION_FACTORES.get()),
                    ("Correlacion matriz-satelites (rho)", f"{self.CORRELACION_MATRIZ_SAT.get():.2f}"),
                ]

                headers_p = ["Parametro", "Valor"]
                for j, h in enumerate(headers_p, 2):
                    c = ws9.cell(row=4, column=j)
                    c.value = h
                    self._aplicar_estilo(c, self._estilo_header())
                ws9.column_dimensions['B'].width = 30
                ws9.column_dimensions['C'].width = 25

                for i, (lab, val) in enumerate(params):
                    fila = 5 + i
                    c = ws9.cell(row=fila, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws9.cell(row=fila, column=3)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                # Formulas
                fila_f = 5 + len(params) + 2
                ws9.merge_cells(f'B{fila_f}:D{fila_f}')
                c = ws9[f'B{fila_f}']
                c.value = "FORMULAS APLICADAS"
                self._aplicar_estilo(c, self._estilo_subtitulo())

                ws9.merge_cells(f'B{fila_f+1}:D{fila_f+1}')
                c = ws9[f'B{fila_f+1}']
                c.value = "Margen Optimo: m* = (Limite_Utilidad + Gastos_Operativos) / Costos"
                c.font = Font(name='Consolas', size=11, bold=True, color='1F4E79')

                ws9.merge_cells(f'B{fila_f+2}:D{fila_f+2}')
                c = ws9[f'B{fila_f+2}']
                c.value = "Equilibrio A (Especial): m = (C*t_pc + G*t_esp) / (C*(t_esp - t_pc))"
                c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

                ws9.merge_cells(f'B{fila_f+3}:D{fila_f+3}')
                c = ws9[f'B{fila_f+3}']
                c.value = "Equilibrio B (Mixto MYPE): m = (C*t_pc + G*t_gral + L*(t_gral-t_esp)) / (C*(t_gral-t_pc))"
                c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

                ws9.merge_cells(f'B{fila_f+4}:D{fila_f+4}')
                c = ws9[f'B{fila_f+4}']
                c.value = "Equilibrio C (General): m = (C*t_pc + G*t_gral) / (C*(t_gral - t_pc))"
                c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

                ws9.merge_cells(f'B{fila_f+6}:D{fila_f+6}')
                c = ws9[f'B{fila_f+6}']
                c.value = ("La formula de margen optimo garantiza que la utilidad neta alcance exactamente "
                          "el limite del regimen especial. Las 3 formulas de equilibrio determinan el punto "
                          "donde IR = PaC segun el regimen resultante: Especial (util<=15UIT), "
                          "Mixto MYPE (util>15UIT, ing<=1700UIT), o General (ing>1700UIT).")
                c.font = Font(name='Calibri', italic=True, color='666666', size=9)
                c.alignment = Alignment(wrap_text=True)

                # ========================================================
                # HOJA 10: SIMULACION EQUILIBRIO FINANCIERO
                # ========================================================
                ws10 = wb.create_sheet("SIM. EQUILIBRIO")
                ws10.sheet_properties.tabColor = "8E44AD"

                ws10.merge_cells('B2:H2')
                c = ws10['B2']
                c.value = "SIMULACION MONTE CARLO - EQUILIBRIO FINANCIERO"
                self._aplicar_estilo(c, self._estilo_titulo())

                if self.datos_equilibrio_sim:
                    deq = self.datos_equilibrio_sim

                    ws10.merge_cells('B4:C4')
                    c = ws10['B4']
                    c.value = "ESTADISTICAS MARGEN DE EQUILIBRIO"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    stats_eq_data = [
                        ("Iteraciones", f"{deq['n_sim']:,}"),
                        ("Variabilidad", f"+/-{deq['variabilidad']:.1f}%"),
                        ("", ""),
                        ("--- MARGEN EQUILIBRIO ---", ""),
                        ("Media", f"{deq['margen_media']:.4f}%"),
                        ("Mediana", f"{deq['margen_mediana']:.4f}%"),
                        ("Desv. Estandar", f"{deq['margen_std']:.4f}%"),
                        ("Percentil 5", f"{deq['margen_p5']:.4f}%"),
                        ("Percentil 95", f"{deq['margen_p95']:.4f}%"),
                        (f"IC {deq['conf_nivel']:.0f}% Inferior", f"{deq['ic_margen_lower']:.4f}%"),
                        (f"IC {deq['conf_nivel']:.0f}% Superior", f"{deq['ic_margen_upper']:.4f}%"),
                        ("", ""),
                        ("--- AHORRO EN EQUILIBRIO ---", ""),
                        ("Media", f"S/ {deq['ahorro_media']:,.0f}"),
                        ("Mediana", f"S/ {deq['ahorro_mediana']:,.0f}"),
                        ("Desv. Estandar", f"S/ {deq['ahorro_std']:,.0f}"),
                        ("Percentil 5", f"S/ {deq['ahorro_p5']:,.0f}"),
                        ("Percentil 95", f"S/ {deq['ahorro_p95']:,.0f}"),
                        (f"IC {deq['conf_nivel']:.0f}% Inferior", f"S/ {deq['ic_ahorro_lower']:,.0f}"),
                        (f"IC {deq['conf_nivel']:.0f}% Superior", f"S/ {deq['ic_ahorro_upper']:,.0f}"),
                        ("P-Valor (Bootstrap)", f"{deq['p_valor_boot']:.4e}"),
                    ]
                    for i, (lab, val) in enumerate(stats_eq_data):
                        fila = 5 + i
                        c = ws10.cell(row=fila, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        if lab.startswith("---"):
                            c.font = Font(name='Calibri', bold=True, color='1F4E79', size=10)
                        c = ws10.cell(row=fila, column=3)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Distribucion de regimenes
                    fila_r = 5 + len(stats_eq_data) + 1
                    ws10.merge_cells(f'B{fila_r}:C{fila_r}')
                    c = ws10[f'B{fila_r}']
                    c.value = "DISTRIBUCION DE REGIMENES EN SIMULACION"
                    self._aplicar_estilo(c, self._estilo_subtitulo())

                    reg = deq['regimenes']
                    total_reg = sum(reg.values())
                    reg_data = [
                        ("Especial Puro", f"{reg['Especial']:,} ({reg['Especial']/max(1,total_reg)*100:.1f}%)"),
                        ("Mixto MYPE", f"{reg['Mixto']:,} ({reg['Mixto']/max(1,total_reg)*100:.1f}%)"),
                        ("General Puro", f"{reg['General']:,} ({reg['General']/max(1,total_reg)*100:.1f}%)"),
                        ("Sin solucion", f"{reg['Sin_solucion']:,} ({reg['Sin_solucion']/max(1,total_reg)*100:.1f}%)"),
                    ]
                    for i, (lab, val) in enumerate(reg_data):
                        f = fila_r + 1 + i
                        c = ws10.cell(row=f, column=2)
                        c.value = lab
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                        c.alignment = Alignment(horizontal='left')
                        c = ws10.cell(row=f, column=3)
                        c.value = val
                        self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Log-Normal fit
                    if deq.get('ln_ajustado'):
                        fila_ln = fila_r + 1 + len(reg_data) + 1
                        ws10.merge_cells(f'B{fila_ln}:C{fila_ln}')
                        c = ws10[f'B{fila_ln}']
                        c.value = "AJUSTE LOG-NORMAL (Margen Equilibrio)"
                        self._aplicar_estilo(c, self._estilo_subtitulo())
                        ln_data = [
                            ("Shape (s)", f"{deq['ln_shape']:.4f}"),
                            ("Scale", f"{deq['ln_scale']:.4f}"),
                        ]
                        for i, (lab, val) in enumerate(ln_data):
                            f = fila_ln + 1 + i
                            c = ws10.cell(row=f, column=2)
                            c.value = lab
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            c = ws10.cell(row=f, column=3)
                            c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    # Notas de configuracion matriz
                    if deq.get('var_matriz_costos') or deq.get('var_matriz_ingresos'):
                        fila_mat = fila_r + 1 + len(reg_data) + 1
                        if deq.get('ln_ajustado'):
                            fila_mat = fila_ln + 1 + 2 + 1
                        ws10.merge_cells(f'B{fila_mat}:C{fila_mat}')
                        c = ws10[f'B{fila_mat}']
                        c.value = "CONFIGURACION VARIABILIDAD MATRIZ"
                        self._aplicar_estilo(c, self._estilo_subtitulo())
                        mat_notas = []
                        if deq.get('var_matriz_costos'):
                            mat_notas.append(("Var. Costos Matriz", f"+/-{deq['pct_var_costos']:.1f}%"))
                        if deq.get('var_matriz_ingresos'):
                            mat_notas.append(("Var. Ingresos Matriz", f"+/-{deq['pct_var_ingresos']:.1f}%"))
                        mat_notas.append(("Distribucion", deq.get('distribucion', 'Log-normal')))
                        mat_notas.append(("Correlacion (rho)", f"{deq.get('rho', 0):.2f}"))
                        for i, (lab, val) in enumerate(mat_notas):
                            f = fila_mat + 1 + i
                            c = ws10.cell(row=f, column=2)
                            c.value = lab
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                            c.alignment = Alignment(horizontal='left')
                            c = ws10.cell(row=f, column=3)
                            c.value = val
                            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

                    ws10.column_dimensions['B'].width = 26
                    ws10.column_dimensions['C'].width = 28

                    if img_equilibrio_sim:
                        self._insertar_imagen(ws10, img_equilibrio_sim, 'E4', ancho_cm=36, alto_cm=22)
                else:
                    ws10['B4'] = "Ejecute 'Simular Equilibrio (Monte Carlo)' en la pestana Pagos a Cuenta."
                    ws10['B4'].font = Font(italic=True, color='999999', size=11)

                # Guardar
                wb.save(path)

            messagebox.showinfo(
                "Exportacion Completa",
                f"Reporte Gerencial exportado exitosamente.\n\n"
                f"Hojas generadas:\n"
                f"  1. Resumen Ejecutivo (+ Equilibrio Financiero)\n"
                f"  2. Detalle Satelites (+ Metricas Equilibrio)\n"
                f"  3. Graficos Resultados{'  [OK]' if img_resultados else '  [Pendiente]'}\n"
                f"  4. Simulacion Monte Carlo (Bootstrap+LogNormal){'  [OK]' if self.datos_simulacion else '  [Pendiente]'}\n"
                f"  5. Sensibilidad{'  [OK]' if img_sensibilidad else '  [Pendiente]'}\n"
                f"  6. Comparativo{'  [OK]' if img_comparativo else '  [Pendiente]'}\n"
                f"  7. Pagos a Cuenta y Equilibrio{'  [OK]' if self.datos_pagos_cuenta else '  [Pendiente]'}\n"
                f"  8. Sensibilidad Margenes\n"
                f"  9. Parametros\n"
                f" 10. Sim. Equilibrio (Monte Carlo){'  [OK]' if self.datos_equilibrio_sim else '  [Pendiente]'}\n\n"
                f"Archivo: {path}"
            )
        except Exception as e:
            messagebox.showerror("Error de Exportacion", f"No se pudo generar el Excel:\n{str(e)}")
    
    def copiar_resultados(self):
        try:
            if hasattr(self, 'resultados'):
                self.root.clipboard_clear()
                self.root.clipboard_append(self.text_resultados.get(1.0, tk.END))
                messagebox.showinfo("√âxito", "Resultados copiados")
        except Exception as e:
            messagebox.showerror("Error", f"Error: {str(e)}")
    
    def limpiar_datos(self):
        self.entry_nombre_matriz.delete(0, tk.END)
        self.entry_ingresos_matriz.delete(0, tk.END)
        self.entry_costos_matriz.delete(0, tk.END)
        
        for sat in self.satelites_entries:
            sat['entry_nombre'].delete(0, tk.END)
            sat['entry_costo'].delete(0, tk.END)
            sat['entry_gastos'].delete(0, tk.END)
            sat['var_regimen_general'].set(False)
        
        self.text_resultados.delete(1.0, tk.END)
    
    def cargar_ejemplo(self):
        self.entry_nombre_matriz.delete(0, tk.END)
        self.entry_nombre_matriz.insert(0, "KALLPA")
        
        self.entry_ingresos_matriz.delete(0, tk.END)
        self.entry_ingresos_matriz.insert(0, "36184868.55")
        
        self.entry_costos_matriz.delete(0, tk.END)
        self.entry_costos_matriz.insert(0, "31674202.9240117")
        
        nombres = ["IG PAME", "SUMAQ WARMY", "KUSKA KUSISJA", "MISKY"]
        costos = ["1714276", "7000000", "4500000", "2550000"]
        gastos = ["0", "0", "0", "500000"]
        
        for i, sat in enumerate(self.satelites_entries):
            if i < len(nombres):
                sat['entry_nombre'].delete(0, tk.END)
                sat['entry_nombre'].insert(0, nombres[i])
                
                sat['entry_costo'].delete(0, tk.END)
                sat['entry_costo'].insert(0, costos[i])
                
                sat['entry_gastos'].delete(0, tk.END)
                sat['entry_gastos'].insert(0, gastos[i])
                
                sat['var_regimen_general'].set(False)
        
        messagebox.showinfo("Ejemplo", "Datos cargados")

if __name__ == "__main__":
    root = tk.Tk()
    app = OptimizadorPro(root)
    root.mainloop()