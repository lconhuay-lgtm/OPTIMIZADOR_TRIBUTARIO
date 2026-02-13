import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWs7Mixin:
    def _exportar_ws7(self, wb, r, imgs):
        # HOJA 7: PAGOS A CUENTA Y SENSIBILIDAD
        # ========================================================
        ws7 = wb.create_sheet("PAGOS A CUENTA")
        ws7.sheet_properties.tabColor = "E67E22"

        ws7.merge_cells('B2:K2')
        c = ws7['B2']
        c.value = "ANALISIS DE PAGOS A CUENTA Y EQUILIBRIO FINANCIERO"
        self._aplicar_estilo(c, self._estilo_titulo())
        ws7.row_dimensions[2].height = 30

        if self.datos_pagos_cuenta:
            dpc = self.datos_pagos_cuenta
            gpc = dpc['global']
            img_pagos_cuenta = imgs.get('pagos_cuenta')

            # --- EXPLICACION GERENCIAL DEL CONCEPTO ---
            ws7.merge_cells('B4:K4')
            c = ws7['B4']
            c.value = "CONCEPTO: PUNTO DE EQUILIBRIO FINANCIERO (IR = Pagos a Cuenta, Saldo = 0)"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            ws7.merge_cells('B5:K5')
            c = ws7['B5']
            c.value = ("El punto de equilibrio financiero es el margen donde los pagos a cuenta mensuales (adelantos obligatorios del IR al "
                      f"{dpc['tasa_pago_cuenta']:.1f}% de los ingresos) igualan EXACTAMENTE al Impuesto a la Renta anual. "
                      "En este punto: saldo de regularizacion = 0, eficiencia = 100%, flujo de caja optimizado. "
                      "No hay dinero inmovilizado en SUNAT ni necesidad de liquidez adicional para regularizacion.")
            self._aplicar_estilo(c, {
                'font': Font(name='Calibri', italic=True, color='444444', size=10),
                'alignment': Alignment(horizontal='left', wrap_text=True)
            })
            ws7.row_dimensions[5].height = 50

            # --- SECCION I: ESCENARIO SIN ESTRUCTURA ---
            fila = 7
            ws7.merge_cells(f'B{fila}:F{fila}')
            c = ws7[f'B{fila}']
            c.value = f"I. ESCENARIO SIN ESTRUCTURA SATELITAL - {gpc['nombre_matriz']}"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            fila += 1

            datos_sin_est = [
                ("Ingresos Totales", gpc['ingresos_matriz']),
                ("Utilidad Gravable", gpc['utilidad_sin_estructura']),
                ("Tasa Aplicable", "29.5% (Regimen General)"),
                ("Impuesto a la Renta", gpc['impuesto_sin_estructura']),
                (f"Pagos a Cuenta ({dpc['tasa_pago_cuenta']:.1f}%)", gpc['pc_matriz_sin']),
                ("Saldo Regularizacion", gpc['saldo_matriz_sin']),
                ("Eficiencia PaC", f"{gpc['eficiencia_sin']:.1f}%"),
                ("Tasa Efectiva", f"{gpc['tasa_efectiva_sin']:.1f}%"),
            ]
            for i, (lab, val) in enumerate(datos_sin_est):
                c = ws7.cell(row=fila + i, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                c = ws7.cell(row=fila + i, column=3)
                if isinstance(val, (int, float)):
                    c.value = val
                    c.number_format = '#,##0.00'
                else:
                    c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                if lab == "Saldo Regularizacion":
                    color_s = 'CC0000' if gpc['saldo_matriz_sin'] < 0 else '006600'
                    c.font = Font(name='Calibri', bold=True, color=color_s, size=10)
                    c2 = ws7.cell(row=fila + i, column=4)
                    c2.value = "(FAVOR)" if gpc['saldo_matriz_sin'] < 0 else "(POR PAGAR)"
                    c2.font = Font(name='Calibri', bold=True, color=color_s, size=9)

            fila += len(datos_sin_est) + 1

            # --- SECCION II: GRUPO EN PUNTO DE EQUILIBRIO ---
            ws7.merge_cells(f'B{fila}:F{fila}')
            c = ws7[f'B{fila}']
            c.value = "II. GRUPO CONSOLIDADO EN PUNTO DE EQUILIBRIO (IR = PaC, Saldo = 0)"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            fila += 1

            datos_eq = [
                ("--- MATRIZ ---", ""),
                ("Utilidad Matriz (ajustada)", gpc['utilidad_matriz_eq']),
                ("Impuesto Matriz", gpc['impuesto_matriz_eq']),
                ("PaC Matriz", gpc['pc_matriz_eq']),
                ("Saldo Matriz", gpc['saldo_matriz_eq']),
                ("--- SATELITES EN EQUILIBRIO ---", ""),
                ("IR Satelites (equilibrio)", gpc['total_ir_eq_sats']),
                ("PaC Satelites (equilibrio)", gpc['total_pc_eq_sats']),
                ("Saldo Satelites", gpc['total_saldo_eq_sats']),
                ("--- GRUPO CONSOLIDADO ---", ""),
                ("IR Total Grupo", gpc['ir_grupo_eq']),
                ("PaC Total Grupo", gpc['pc_grupo_eq']),
                ("Saldo Grupo", gpc['saldo_grupo_eq']),
                ("Eficiencia PaC", f"{gpc['eficiencia_eq']:.1f}%"),
                ("Tasa Efectiva", f"{gpc['tasa_efectiva_eq']:.1f}%"),
                ("Ahorro vs Sin Estructura", gpc['ahorro_eq_total']),
                ("Ahorro %", f"{gpc['ahorro_eq_pct']:.1f}%"),
            ]
            for i, (lab, val) in enumerate(datos_eq):
                c = ws7.cell(row=fila + i, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                if lab.startswith("---"):
                    c.font = Font(name='Calibri', bold=True, color='1F4E79', size=10)
                c = ws7.cell(row=fila + i, column=3)
                if isinstance(val, (int, float)):
                    c.value = val
                    c.number_format = '#,##0.00'
                else:
                    c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                if "Saldo" in lab and isinstance(val, (int, float)):
                    color_s = '2E75B6' if abs(val) < 100 else ('CC0000' if val < 0 else '006600')
                    c.font = Font(name='Calibri', bold=True, color=color_s, size=10)
                    c2 = ws7.cell(row=fila + i, column=4)
                    c2.value = "(EQUILIBRIO)" if abs(val) < 100 else ("(FAVOR)" if val < 0 else "(PAGAR)")
                    c2.font = Font(name='Calibri', bold=True, color=color_s, size=9)
                if lab == "Ahorro vs Sin Estructura":
                    c.font = Font(name='Calibri', bold=True, color='006600', size=11)

            fila += len(datos_eq) + 1

            # --- SECCION III: COMPARATIVO 3 ESCENARIOS ---
            ws7.merge_cells(f'B{fila}:H{fila}')
            c = ws7[f'B{fila}']
            c.value = "III. COMPARATIVO: SIN ESTRUCTURA vs EQUILIBRIO vs MARGEN OPTIMO (ref)"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            fila += 1

            headers_comp_eq = ["Concepto", "Sin Estructura", "EQUILIBRIO", "M. Optimo (ref)", "Mejor"]
            for j, h in enumerate(headers_comp_eq, 2):
                c = ws7.cell(row=fila, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
                if h == "EQUILIBRIO":
                    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            ws7.row_dimensions[fila].height = 28
            fila += 1

            filas_comp = [
                ("IR Total Grupo", gpc['impuesto_sin_estructura'], gpc['ir_grupo_eq'], gpc['ir_grupo_opt'],
                 "Equilibrio" if gpc['ir_grupo_eq'] <= gpc['ir_grupo_opt'] else "M. Optimo"),
                ("PaC Total Grupo", gpc['pc_matriz_sin'], gpc['pc_grupo_eq'], gpc['pc_grupo_opt'], "-"),
                ("Saldo Regularizacion", gpc['saldo_matriz_sin'], gpc['saldo_grupo_eq'], gpc['saldo_grupo_opt'],
                 "Equilibrio" if abs(gpc['saldo_grupo_eq']) < abs(gpc['saldo_grupo_opt']) else "M. Optimo"),
                ("Ahorro Tributario", 0, gpc['ahorro_eq_total'], gpc['ahorro_opt_total'],
                 "Equilibrio" if gpc['ahorro_eq_total'] >= gpc['ahorro_opt_total'] else "M. Optimo"),
            ]
            for i, (concepto, v_sin, v_eq, v_opt, mejor) in enumerate(filas_comp):
                alt = i % 2 == 0
                c = ws7.cell(row=fila, column=2)
                c.value = concepto
                self._aplicar_estilo(c, self._estilo_celda(alt))
                c.alignment = Alignment(horizontal='left')
                for j, val in enumerate([v_sin, v_eq, v_opt], 3):
                    c = ws7.cell(row=fila, column=j)
                    c.value = val
                    c.number_format = '#,##0.00'
                    self._aplicar_estilo(c, self._estilo_celda(alt))
                    if j == 4:  # Columna equilibrio destacada
                        c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                c = ws7.cell(row=fila, column=6)
                c.value = mejor
                self._aplicar_estilo(c, self._estilo_celda(alt))
                if "Equilibrio" in mejor:
                    c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                fila += 1

            # Filas de porcentaje
            for i, (concepto, v_sin, v_eq, v_opt, mejor) in enumerate([
                ("Eficiencia PaC (%)", gpc['eficiencia_sin'], gpc['eficiencia_eq'], gpc['eficiencia_opt'],
                 "Equilibrio" if abs(gpc['eficiencia_eq'] - 100) <= abs(gpc['eficiencia_opt'] - 100) else "M. Optimo"),
                ("Tasa Efectiva (%)", gpc['tasa_efectiva_sin'], gpc['tasa_efectiva_eq'], gpc['tasa_efectiva_opt'],
                 "Equilibrio" if gpc['tasa_efectiva_eq'] <= gpc['tasa_efectiva_opt'] else "M. Optimo"),
            ]):
                alt = (len(filas_comp) + i) % 2 == 0
                c = ws7.cell(row=fila, column=2)
                c.value = concepto
                self._aplicar_estilo(c, self._estilo_celda(alt))
                c.alignment = Alignment(horizontal='left')
                for j, val in enumerate([v_sin, v_eq, v_opt], 3):
                    c = ws7.cell(row=fila, column=j)
                    c.value = val
                    c.number_format = '0.0"%"'
                    self._aplicar_estilo(c, self._estilo_celda(alt))
                    if j == 4:
                        c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                c = ws7.cell(row=fila, column=6)
                c.value = mejor
                self._aplicar_estilo(c, self._estilo_celda(alt))
                if "Equilibrio" in mejor:
                    c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                fila += 1

            # KPI
            fila += 1
            kpis_eq = [
                ("AHORRO EN EQUILIBRIO", f"S/ {gpc['ahorro_eq_total']:,.0f} ({gpc['ahorro_eq_pct']:.1f}%)"),
                ("EFICIENCIA PaC", f"{gpc['eficiencia_eq']:.1f}%"),
            ]
            for i, (lab, val) in enumerate(kpis_eq):
                col_k = 2 + i * 3
                c = ws7.cell(row=fila, column=col_k)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_kpi_label())
                c = ws7.cell(row=fila + 1, column=col_k)
                c.value = val
                self._aplicar_estilo(c, self._estilo_kpi_valor())
            ws7.row_dimensions[fila].height = 30
            ws7.row_dimensions[fila + 1].height = 35
            fila += 3

            # Continua en _exportar_ws7_detalle
            self._exportar_ws7_detalle(ws7, dpc, gpc, fila, img_pagos_cuenta)
        else:
            ws7['B4'] = "Ejecute el Analisis de Pagos a Cuenta para generar datos y graficos."
            ws7['B4'].font = Font(italic=True, color='999999', size=11)
