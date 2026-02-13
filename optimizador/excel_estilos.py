import os
import tempfile
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelEstilosMixin:
    def _estilo_header(self):
        """Retorna estilos para headers de tabla."""
        return {
            'font': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                bottom=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
            )
        }

    def _estilo_celda(self, es_alterno=False):
        """Retorna estilos para celdas de datos."""
        fill_color = 'D6E4F0' if es_alterno else 'FFFFFF'
        return {
            'font': Font(name='Calibri', size=10),
            'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                bottom=Side(style='thin', color='C0C0C0'),
                top=Side(style='thin', color='C0C0C0'),
                left=Side(style='thin', color='C0C0C0'),
                right=Side(style='thin', color='C0C0C0')
            )
        }

    def _estilo_titulo(self):
        """Retorna estilos para títulos de sección."""
        return {
            'font': Font(name='Calibri', bold=True, color='1F4E79', size=14),
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    def _estilo_subtitulo(self):
        """Retorna estilos para subtítulos."""
        return {
            'font': Font(name='Calibri', bold=True, color='2E75B6', size=11),
            'alignment': Alignment(horizontal='left', vertical='center')
        }

    def _estilo_kpi_valor(self):
        """Retorna estilos para valores KPI destacados."""
        return {
            'font': Font(name='Calibri', bold=True, color='FFFFFF', size=13),
            'fill': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                bottom=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79')
            )
        }

    def _estilo_kpi_label(self):
        """Retorna estilos para etiquetas KPI."""
        return {
            'font': Font(name='Calibri', bold=True, color='1F4E79', size=10),
            'fill': PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                bottom=Side(style='thin', color='1F4E79'),
                top=Side(style='thin', color='1F4E79'),
                left=Side(style='thin', color='1F4E79'),
                right=Side(style='thin', color='1F4E79')
            )
        }

    def _aplicar_estilo(self, celda, estilo_dict):
        """Aplica un diccionario de estilos a una celda."""
        for attr, value in estilo_dict.items():
            setattr(celda, attr, value)

    def _guardar_figura_temp(self, fig, nombre, tmpdir):
        """Guarda una figura matplotlib como PNG de alta resolución."""
        if fig is None:
            return None
        path = os.path.join(tmpdir, f"{nombre}.png")
        fig.savefig(path, dpi=180, bbox_inches='tight', facecolor='white', edgecolor='none')
        return path

    def _insertar_imagen(self, ws, img_path, celda, ancho_cm=28, alto_cm=16):
        """Inserta imagen en hoja Excel con tamaño especificado."""
        if img_path and os.path.exists(img_path):
            img = XlImage(img_path)
            img.width = int(ancho_cm * 37.8)   # cm a pixels aprox
            img.height = int(alto_cm * 37.8)
            ws.add_image(img, celda)

