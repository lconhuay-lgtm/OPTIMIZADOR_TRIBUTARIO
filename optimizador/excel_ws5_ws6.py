from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelWs5Ws6Mixin:
    def _exportar_ws5(self, wb, imgs):
        # HOJA 5: ANALISIS SENSIBILIDAD
        # ========================================================
        ws5 = wb.create_sheet("SENSIBILIDAD")
        ws5.sheet_properties.tabColor = "F39C12"

        ws5.merge_cells('B2:H2')
        c = ws5['B2']
        c.value = "ANALISIS DE SENSIBILIDAD"
        self._aplicar_estilo(c, self._estilo_titulo())

        ws5.merge_cells('B4:H4')
        c = ws5['B4']
        c.value = "Impacto de variaciones en parametros clave sobre el ahorro tributario"
        self._aplicar_estilo(c, {
            'font': Font(name='Calibri', italic=True, color='666666', size=10),
            'alignment': Alignment(horizontal='left')
        })

        if img_sensibilidad:
            self._insertar_imagen(ws5, img_sensibilidad, 'B6', ancho_cm=36, alto_cm=16)
        else:
            ws5['B6'] = "Ejecute el Analisis de Sensibilidad para generar graficos."
            ws5['B6'].font = Font(italic=True, color='999999', size=11)

        # Tabla de datos N-Satelites si existe
        if self.datos_sensibilidad_nsats:
            dns = self.datos_sensibilidad_nsats
            fila_ns = 28 if img_sensibilidad else 8
            ws5.merge_cells(f'B{fila_ns}:F{fila_ns}')
            c = ws5[f'B{fila_ns}']
            c.value = "SENSIBILIDAD: NUMERO DE SATELITES"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            ws5.merge_cells(f'B{fila_ns+1}:F{fila_ns+1}')
            c = ws5[f'B{fila_ns+1}']
            c.value = (f"Satelite referencia: Costo={dns['costo_ref']:,.0f}, "
                      f"Gastos={dns['gastos_ref']:,.0f} | N actual={dns['n_actual']}")
            c.font = Font(name='Calibri', italic=True, color='666666', size=9)

            ns_headers = ["N Satelites", "Ahorro Tributario (S/)", "Tasa Efectiva (%)",
                          "IR Grupo (S/)", "Ahorro Marginal (S/)"]
            for j, h in enumerate(ns_headers, 2):
                c = ws5.cell(row=fila_ns + 2, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())

            ahorros_arr_ns = dns['ahorros']
            for i, n in enumerate(dns['ns']):
                f = fila_ns + 3 + i
                ahorro_marginal = (ahorros_arr_ns[i] - ahorros_arr_ns[i-1]) if i > 0 else ahorros_arr_ns[0]
                datos_fila = [n, ahorros_arr_ns[i], dns['tasa_efectiva'][i],
                              dns['ir_grupo'][i], ahorro_marginal]
                for j, val in enumerate(datos_fila, 2):
                    c = ws5.cell(row=f, column=j)
                    c.value = val
                    if j == 2:
                        c.number_format = '0'
                    elif j == 4:
                        c.number_format = '0.00"%"'
                    else:
                        c.number_format = '#,##0'
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    if n == dns['n_actual']:
                        c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)

            for col_idx in range(2, 7):
                ws5.column_dimensions[get_column_letter(col_idx)].width = 22

        # ========================================================

    def _exportar_ws6(self, wb, imgs):
        # HOJA 6: ANALISIS COMPARATIVO
        # ========================================================
        ws6 = wb.create_sheet("COMPARATIVO")
        ws6.sheet_properties.tabColor = "9B59B6"

        ws6.merge_cells('B2:H2')
        c = ws6['B2']
        c.value = "ANALISIS COMPARATIVO MULTI-ESCENARIO"
        self._aplicar_estilo(c, self._estilo_titulo())

        if hasattr(self, 'datos_comparativo') and self.datos_comparativo:
            # Tabla comparativa
            headers_comp = ["Escenario", "Ahorro Tributario", "Tasa Efectiva %", "N Satelites"]
            for j, h in enumerate(headers_comp, 2):
                c = ws6.cell(row=4, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
            ws6.row_dimensions[4].height = 28
            ws6.column_dimensions['B'].width = 20
            ws6.column_dimensions['C'].width = 22
            ws6.column_dimensions['D'].width = 18
            ws6.column_dimensions['E'].width = 14

            for i, esc in enumerate(self.datos_comparativo):
                fila = 5 + i
                alt = i % 2 == 0
                vals = [
                    esc['nombre'],
                    esc['ahorro'],
                    esc['tasa_efectiva'],
                    esc['satelites']
                ]
                for j, val in enumerate(vals, 2):
                    c = ws6.cell(row=fila, column=j)
                    if j == 3:
                        c.value = val
                        c.number_format = '#,##0.00'
                    elif j == 4:
                        c.value = val
                        c.number_format = '0.00"%"'
                    else:
                        c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(alt))

        if img_comparativo:
            fila_img = 5 + (len(self.datos_comparativo) if hasattr(self, 'datos_comparativo') and self.datos_comparativo else 0) + 2
            self._insertar_imagen(ws6, img_comparativo, f'B{fila_img}', ancho_cm=36, alto_cm=16)
        else:
            ws6['B10'] = "Ejecute el Analisis Comparativo para generar graficos."
            ws6['B10'].font = Font(italic=True, color='999999', size=11)

        # ========================================================

