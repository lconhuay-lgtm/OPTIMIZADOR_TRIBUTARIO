import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelWs8Mixin:
    def _exportar_ws8(self, wb, r):
        # HOJA 8: SENSIBILIDAD MARGENES POR SATELITE
        # ========================================================
        ws8 = wb.create_sheet("SENSIBILIDAD MARGENES")
        ws8.sheet_properties.tabColor = "1ABC9C"

        ws8.merge_cells('B2:K2')
        c = ws8['B2']
        c.value = "ANALISIS DE SENSIBILIDAD - MARGENES POR SATELITE"
        self._aplicar_estilo(c, self._estilo_titulo())
        ws8.row_dimensions[2].height = 30

        ws8.merge_cells('B3:K3')
        c = ws8['B3']
        c.value = ("Muestra como se diluye el ahorro fiscal al alejarse del margen optimo. "
                  "Desde la perspectiva de teoria de juegos, el punto optimo es el equilibrio de Nash "
                  "donde ningun satelite puede mejorar su posicion unilateralmente.")
        c.font = Font(name='Calibri', italic=True, color='666666', size=9)
        c.alignment = Alignment(wrap_text=True)
        ws8.row_dimensions[3].height = 30

        fila_sens = 5

        for idx_sat, sat in enumerate(r['satelites']):
            costo_s = sat['costo']
            gastos_s = sat['gastos_operativos']
            margen_opt = sat['margen_optimo'] / 100
            es_general_s = sat['regimen'].startswith('GENERAL')

            ws8.merge_cells(f'B{fila_sens}:K{fila_sens}')
            c = ws8[f'B{fila_sens}']
            c.value = f"{idx_sat + 1}. {sat['nombre']} - Margen Optimo: {sat['margen_optimo']:.4f}%"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            fila_sens += 1

            # Generar variaciones de margen
            tasa_general_dec = r['parametros']['tasa_general'] / 100
            tasa_especial_dec = r['parametros']['tasa_especial'] / 100
            limite_util_val = r['parametros']['limite_utilidad']
            limite_ing_val = r['parametros']['limite_ingresos']
            tasa_pc_dec = self.TASA_PAGO_CUENTA.get() / 100

            m_min = max(0.001, margen_opt * 0.2)
            m_max_f = min(limite_ing_val / costo_s - 1, 1.0) if costo_s > 0 else 1.0
            m_max = min(margen_opt * 3.0, m_max_f)
            if m_max <= m_min:
                m_max = m_min + 0.10

            margenes_v = np.linspace(m_min, m_max, 15)
            if margen_opt not in margenes_v:
                margenes_v = np.sort(np.append(margenes_v, margen_opt))

            headers_s = ["Margen%", "Precio", "Util. Neta", "Regimen", "IR",
                        "Pago Cta", "Saldo Reg.", "Ahorro Reg.", "Ahorro vs Base", "Nota"]
            for j, h in enumerate(headers_s, 2):
                c = ws8.cell(row=fila_sens, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
            ws8.row_dimensions[fila_sens].height = 24
            fila_sens += 1

            # Ahorro base para comparar
            ahorro_base_sat = sat['ahorro_individual']

            for i, m in enumerate(margenes_v):
                precio_v = costo_s * (1 + m)
                util_b = costo_s * m
                util_n = max(0, util_b - gastos_s)

                if es_general_s:
                    ir_v = util_n * tasa_general_dec
                    reg_v = "General"
                    ahorro_reg = 0
                elif util_n > limite_util_val or precio_v > limite_ing_val:
                    if util_n > limite_util_val:
                        ir_v = limite_util_val * tasa_especial_dec + (util_n - limite_util_val) * tasa_general_dec
                        reg_v = "Mixto"
                    else:
                        ir_v = util_n * tasa_general_dec
                        reg_v = "General*"
                    ahorro_reg = util_n * tasa_general_dec - ir_v
                else:
                    ir_v = util_n * tasa_especial_dec
                    reg_v = "Especial"
                    ahorro_reg = util_n * (tasa_general_dec - tasa_especial_dec)

                pac_v = tasa_pc_dec * precio_v
                saldo_v = ir_v - pac_v
                diff_vs_base = ahorro_reg - ahorro_base_sat
                es_opt = abs(m - margen_opt) < 0.0001
                nota_v = "OPTIMO" if es_opt else ""

                datos_v = [m * 100, precio_v, util_n, reg_v, ir_v, pac_v,
                          saldo_v, ahorro_reg, diff_vs_base, nota_v]
                for j, val in enumerate(datos_v, 2):
                    c = ws8.cell(row=fila_sens, column=j)
                    if isinstance(val, (int, float)):
                        c.value = val
                        if j == 2:
                            c.number_format = '0.0000"%"'
                        else:
                            c.number_format = '#,##0.00'
                    else:
                        c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    if es_opt:
                        c.font = Font(name='Calibri', bold=True, color='006600', size=10)
                    if j == 10 and isinstance(val, (int, float)) and val < 0:
                        c.font = Font(name='Calibri', color='CC0000', size=10)
                fila_sens += 1

            fila_sens += 2

        for col_idx in range(2, 12):
            ws8.column_dimensions[get_column_letter(col_idx)].width = max(
                ws8.column_dimensions[get_column_letter(col_idx)].width or 0, 14)

        # ========================================================

