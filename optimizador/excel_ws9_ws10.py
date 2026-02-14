import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelWs9Ws10Mixin:
    def _exportar_ws9(self, wb, r):
        # HOJA 9: PARAMETROS
        # ========================================================
        ws9 = wb.create_sheet("PARAMETROS")
        ws9.sheet_properties.tabColor = "95A5A6"

        ws9.merge_cells('B2:D2')
        c = ws9['B2']
        c.value = "PARAMETROS TRIBUTARIOS UTILIZADOS"
        self._aplicar_estilo(c, self._estilo_titulo())

        params = [
            ("UIT", f"S/ {r['parametros']['uit']:,.2f}"),
            ("Tasa Regimen General", f"{r['parametros']['tasa_general']:.2f}%"),
            ("Tasa Regimen Especial", f"{r['parametros']['tasa_especial']:.2f}%"),
            ("Diferencial de Tasas", f"{r['parametros']['tasa_general'] - r['parametros']['tasa_especial']:.2f} pp"),
            ("Tasa Pago a Cuenta IR", f"{self.TASA_PAGO_CUENTA.get():.2f}%"),
            ("Limite Utilidad Especial", f"S/ {r['parametros']['limite_utilidad']:,.2f}"),
            ("Limite Ingresos Especial", f"S/ {r['parametros']['limite_ingresos']:,.2f}"),
            ("N Empresas Satelites", f"{len(r['satelites'])}"),
            ("Empresa Matriz", r['matriz']['nombre']),
            ("", ""),
            ("--- SIMULACION: VARIABILIDAD MATRIZ ---", ""),
            ("Variabilidad costos matriz (+/-%)", f"{self.PCT_VAR_MATRIZ_COSTOS.get():.1f}%" if self.VAR_MATRIZ_COSTOS.get() else "Desactivada"),
            ("Variabilidad ingresos matriz (+/-%)", f"{self.PCT_VAR_MATRIZ_INGRESOS.get():.1f}%" if self.VAR_MATRIZ_INGRESOS.get() else "Desactivada"),
            ("Distribucion factores", self.DISTRIBUCION_FACTORES.get()),
            ("Correlacion matriz-satelites (rho)", f"{self.CORRELACION_MATRIZ_SAT.get():.2f}"),
        ]

        headers_p = ["Parametro", "Valor"]
        for j, h in enumerate(headers_p, 2):
            c = ws9.cell(row=4, column=j)
            c.value = h
            self._aplicar_estilo(c, self._estilo_header())
        ws9.column_dimensions['B'].width = 30
        ws9.column_dimensions['C'].width = 25

        for i, (lab, val) in enumerate(params):
            fila = 5 + i
            c = ws9.cell(row=fila, column=2)
            c.value = lab
            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
            c.alignment = Alignment(horizontal='left')
            c = ws9.cell(row=fila, column=3)
            c.value = val
            self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

        # Formulas
        fila_f = 5 + len(params) + 2
        ws9.merge_cells(f'B{fila_f}:D{fila_f}')
        c = ws9[f'B{fila_f}']
        c.value = "FORMULAS APLICADAS"
        self._aplicar_estilo(c, self._estilo_subtitulo())

        ws9.merge_cells(f'B{fila_f+1}:D{fila_f+1}')
        c = ws9[f'B{fila_f+1}']
        c.value = "Margen Optimo: m* = (Limite_Utilidad + Gastos_Operativos) / Costos"
        c.font = Font(name='Consolas', size=11, bold=True, color='1F4E79')

        ws9.merge_cells(f'B{fila_f+2}:D{fila_f+2}')
        c = ws9[f'B{fila_f+2}']
        c.value = "Equilibrio A (Especial): m = (C*t_pc + G*t_esp) / (C*(t_esp - t_pc))"
        c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

        ws9.merge_cells(f'B{fila_f+3}:D{fila_f+3}')
        c = ws9[f'B{fila_f+3}']
        c.value = "Equilibrio B (Mixto MYPE): m = (C*t_pc + G*t_gral + L*(t_gral-t_esp)) / (C*(t_gral-t_pc))"
        c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

        ws9.merge_cells(f'B{fila_f+4}:D{fila_f+4}')
        c = ws9[f'B{fila_f+4}']
        c.value = "Equilibrio C (General): m = (C*t_pc + G*t_gral) / (C*(t_gral - t_pc))"
        c.font = Font(name='Consolas', size=10, bold=True, color='E67E22')

        ws9.merge_cells(f'B{fila_f+6}:D{fila_f+6}')
        c = ws9[f'B{fila_f+6}']
        c.value = ("La formula de margen optimo garantiza que la utilidad neta alcance exactamente "
                  "el limite del regimen especial. Las 3 formulas de equilibrio determinan el punto "
                  "donde IR = PaC segun el regimen resultante: Especial (util<=15UIT), "
                  "Mixto MYPE (util>15UIT, ing<=1700UIT), o General (ing>1700UIT).")
        c.font = Font(name='Calibri', italic=True, color='666666', size=9)
        c.alignment = Alignment(wrap_text=True)

        # ========================================================

    def _exportar_ws10(self, wb, imgs):
        # HOJA 10: SIMULACION EQUILIBRIO FINANCIERO
        # ========================================================
        ws10 = wb.create_sheet("SIM. EQUILIBRIO")
        ws10.sheet_properties.tabColor = "8E44AD"

        ws10.merge_cells('B2:H2')
        c = ws10['B2']
        c.value = "SIMULACION MONTE CARLO - EQUILIBRIO FINANCIERO"
        self._aplicar_estilo(c, self._estilo_titulo())

        if self.datos_equilibrio_sim and isinstance(self.datos_equilibrio_sim, dict) and 'grupo_eq' in self.datos_equilibrio_sim:
            deq = self.datos_equilibrio_sim
            seq = deq.get('grupo_eq', {})
            sopt = deq.get('grupo_opt', {})
            cfg = deq.get('config', {})

            ws10.merge_cells('B4:C4')
            c = ws10['B4']
            c.value = "SALDO DE REGULARIZACION - EQUILIBRIO vs OPTIMO"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            stats_eq_data = [
                ("Iteraciones", f"{deq.get('n_sim', 0):,}"),
                ("Variabilidad Satelites", f"+/-{deq.get('variabilidad', 0):.1f}%"),
                ("", ""),
                ("--- ESCENARIO EQUILIBRIO (Grupo) ---", ""),
                ("Media Saldo", f"S/ {seq.get('media', 0):,.0f}"),
                ("Mediana Saldo", f"S/ {seq.get('mediana', 0):,.0f}"),
                ("Desv. Estandar", f"S/ {seq.get('std', 0):,.0f}"),
                ("Percentil 5", f"S/ {seq.get('p5', 0):,.0f}"),
                ("Percentil 25", f"S/ {seq.get('p25', 0):,.0f}"),
                ("Percentil 75", f"S/ {seq.get('p75', 0):,.0f}"),
                ("Percentil 95", f"S/ {seq.get('p95', 0):,.0f}"),
                ("IC 95% Inferior", f"S/ {seq.get('ic_low', 0):,.0f}"),
                ("IC 95% Superior", f"S/ {seq.get('ic_high', 0):,.0f}"),
                ("", ""),
                ("--- ESCENARIO OPTIMO (Grupo) ---", ""),
                ("Media Saldo", f"S/ {sopt.get('media', 0):,.0f}"),
                ("Mediana Saldo", f"S/ {sopt.get('mediana', 0):,.0f}"),
                ("Percentil 95", f"S/ {sopt.get('p95', 0):,.0f}"),
                ("", ""),
                ("--- COMPARACION ---", ""),
                ("Reduccion P95", f"{100*(1 - seq.get('p95', 0)/sopt.get('p95', 1)):.1f}%" if sopt.get('p95', 0) != 0 else "N/A"),
                ("P-Valor (Opt > Eq)", f"{deq.get('p_valor_diferencia', 1):.4e}"),
            ]
            for i, (lab, val) in enumerate(stats_eq_data):
                fila = 5 + i
                c = ws10.cell(row=fila, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                if lab.startswith("---"):
                    c.font = Font(name='Calibri', bold=True, color='1F4E79', size=10)
                c = ws10.cell(row=fila, column=3)
                c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            # Detalle por empresa
            fila_emp = 5 + len(stats_eq_data) + 1
            ws10.merge_cells(f'B{fila_emp}:H{fila_emp}')
            c = ws10[f'B{fila_emp}']
            c.value = "DETALLE POR EMPRESA - MEDIANA SALDO (S/)"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            emp_headers = ["Empresa", "Mediana Opt", "P5 Opt", "P95 Opt",
                           "Mediana Eq", "P5 Eq", "P95 Eq"]
            for j, h in enumerate(emp_headers, 2):
                c = ws10.cell(row=fila_emp + 1, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())

            hist = deq.get('historial_saldos', {})
            empresas_exp = deq.get('empresas', [])
            for i, emp in enumerate(empresas_exp):
                f = fila_emp + 2 + i
                emp_hist = hist.get(emp, {'opt': [0], 'eq': [0]})
                opt_arr = np.array(emp_hist['opt']) if emp_hist['opt'] else np.array([0])
                eq_arr = np.array(emp_hist['eq']) if emp_hist['eq'] else np.array([0])
                vals = [emp,
                        float(np.median(opt_arr)), float(np.percentile(opt_arr, 5)),
                        float(np.percentile(opt_arr, 95)),
                        float(np.median(eq_arr)), float(np.percentile(eq_arr, 5)),
                        float(np.percentile(eq_arr, 95))]
                for j, val in enumerate(vals, 2):
                    c = ws10.cell(row=f, column=j)
                    if isinstance(val, float):
                        c.value = val
                        c.number_format = '#,##0'
                    else:
                        c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            fila_cfg = fila_emp + 2 + len(empresas_exp) + 1
            # Notas de configuracion matriz
            if cfg.get('var_matriz_costos') or cfg.get('var_matriz_ingresos'):
                ws10.merge_cells(f'B{fila_cfg}:C{fila_cfg}')
                c = ws10[f'B{fila_cfg}']
                c.value = "CONFIGURACION VARIABILIDAD MATRIZ"
                self._aplicar_estilo(c, self._estilo_subtitulo())
                mat_notas = []
                if cfg.get('var_matriz_costos'):
                    mat_notas.append(("Var. Costos Matriz", f"+/-{cfg.get('pct_var_costos', 0):.1f}%"))
                if cfg.get('var_matriz_ingresos'):
                    mat_notas.append(("Var. Ingresos Matriz", f"+/-{cfg.get('pct_var_ingresos', 0):.1f}%"))
                mat_notas.append(("Distribucion", cfg.get('distribucion', 'Log-normal')))
                mat_notas.append(("Correlacion (rho)", f"{cfg.get('rho', 0):.2f}"))
                for i, (lab, val) in enumerate(mat_notas):
                    f = fila_cfg + 1 + i
                    c = ws10.cell(row=f, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws10.cell(row=f, column=3)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            ws10.column_dimensions['B'].width = 22
            ws10.column_dimensions['C'].width = 18
            for col_idx in range(4, 9):
                ws10.column_dimensions[get_column_letter(col_idx)].width = 16

            if imgs.get('equilibrio_sim'):
                self._insertar_imagen(ws10, imgs['equilibrio_sim'], 'J4', ancho_cm=36, alto_cm=28)
        else:
            ws10['B4'] = "Ejecute 'Simular Equilibrio (Monte Carlo)' en la pestana Pagos a Cuenta."
            ws10['B4'].font = Font(italic=True, color='999999', size=11)

        # Guardar

