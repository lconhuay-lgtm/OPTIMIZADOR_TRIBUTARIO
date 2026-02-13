import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelWs4Mixin:
    def _exportar_ws4(self, wb, imgs):
        # HOJA 4: SIMULACION MONTE CARLO
        # ========================================================
        ws4 = wb.create_sheet("SIMULACION MONTE CARLO")
        ws4.sheet_properties.tabColor = "E74C3C"

        ws4.merge_cells('B2:H2')
        c = ws4['B2']
        c.value = "SIMULACION MONTE CARLO - ANALISIS DE RIESGO"
        self._aplicar_estilo(c, self._estilo_titulo())

        if self.datos_simulacion:
            ds = self.datos_simulacion

            # Tabla de estadisticas
            ws4.merge_cells('B4:C4')
            c = ws4['B4']
            c.value = "ESTADISTICAS DE SIMULACION"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            stats_data = [
                ("Iteraciones", f"{ds['n_sim']:,}"),
                ("Variabilidad", f"+/-{ds['variabilidad']:.1f}%"),
                ("Ahorro Medio", f"S/ {ds['media']:,.0f}"),
                ("Mediana", f"S/ {ds['mediana']:,.0f}"),
                ("Desv. Estandar", f"S/ {ds['std']:,.0f}"),
                ("Coef. Variacion", f"{ds['cv']:.2f}%"),
                ("Minimo", f"S/ {ds['min']:,.0f}"),
                ("Maximo", f"S/ {ds['max']:,.0f}"),
                ("Percentil 5", f"S/ {ds['p5']:,.0f}"),
                ("Percentil 25", f"S/ {ds['p25']:,.0f}"),
                ("Percentil 75", f"S/ {ds['p75']:,.0f}"),
                ("Percentil 95", f"S/ {ds['p95']:,.0f}"),
            ]
            for i, (lab, val) in enumerate(stats_data):
                fila = 5 + i
                c = ws4.cell(row=fila, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                c = ws4.cell(row=fila, column=3)
                c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
            ws4.column_dimensions['B'].width = 24
            ws4.column_dimensions['C'].width = 28

            # Bootstrap IC
            fila_b = 5 + len(stats_data) + 1
            ws4.merge_cells(f'B{fila_b}:C{fila_b}')
            c = ws4[f'B{fila_b}']
            c.value = f"INTERVALO DE CONFIANZA - {ds.get('ic_metodo', 'Bootstrap Percentil')}"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            ic_data = [
                ("Metodo", ds.get('ic_metodo', 'Bootstrap Percentil')),
                ("N Bootstrap", f"{ds.get('n_bootstrap', 5000):,}"),
                (f"IC {ds['conf_nivel']:.0f}% Inferior", f"S/ {ds['ic_lower']:,.0f}"),
                (f"IC {ds['conf_nivel']:.0f}% Superior", f"S/ {ds['ic_upper']:,.0f}"),
            ]
            for i, (lab, val) in enumerate(ic_data):
                f = fila_b + 1 + i
                c = ws4.cell(row=f, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                c = ws4.cell(row=f, column=3)
                c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            # Prueba de hipotesis
            fila_h = fila_b + 1 + len(ic_data) + 1
            ws4.merge_cells(f'B{fila_h}:C{fila_h}')
            c = ws4[f'B{fila_h}']
            c.value = "PRUEBA DE HIPOTESIS (No Parametrica Bootstrap)"
            self._aplicar_estilo(c, self._estilo_subtitulo())

            hipotesis_data = [
                ("H0", "Ahorro medio = 0"),
                ("H1", "Ahorro medio > 0"),
                ("P-Valor (Bootstrap)", f"{ds['p_value']:.2e}"),
                ("t-stat (referencia)", f"{ds['t_stat']:.4f}"),
                ("P-Valor t (referencia)", f"{ds.get('p_value_t', ds['p_value']):.2e}"),
                ("Alfa", f"{ds['alfa']}"),
                ("Nivel Confianza", f"{ds['conf_nivel']:.1f}%"),
                ("DECISION", ds['decision']),
            ]
            for i, (lab, val) in enumerate(hipotesis_data):
                f = fila_h + 1 + i
                c = ws4.cell(row=f, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                c = ws4.cell(row=f, column=3)
                c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                if lab == "DECISION":
                    c.font = Font(name='Calibri', bold=True, color='006600', size=11)

            # Ajuste Log-Normal
            fila_ln_ws4 = fila_h + 1 + len(hipotesis_data) + 1  # default position
            ln_data_ws4 = []
            if ds.get('ln_ajustado'):
                fila_ln_ws4 = fila_h + 1 + len(hipotesis_data) + 1
                ws4.merge_cells(f'B{fila_ln_ws4}:C{fila_ln_ws4}')
                c = ws4[f'B{fila_ln_ws4}']
                c.value = "AJUSTE DISTRIBUCION LOG-NORMAL"
                self._aplicar_estilo(c, self._estilo_subtitulo())
                ln_data_ws4 = [
                    ("Distribucion", "Log-Normal"),
                    ("Shape (s)", f"{ds.get('ln_shape', 0):.4f}"),
                    ("Location", f"{ds.get('ln_loc', 0):.4f}"),
                    ("Scale", f"S/ {ds.get('ln_scale', 0):,.0f}"),
                ]
                for i, (lab, val) in enumerate(ln_data_ws4):
                    f = fila_ln_ws4 + 1 + i
                    c = ws4.cell(row=f, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws4.cell(row=f, column=3)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            # Notas de configuracion matriz
            if ds.get('var_matriz_costos') or ds.get('var_matriz_ingresos'):
                if ds.get('ln_ajustado') and ln_data_ws4:
                    fila_nota = fila_ln_ws4 + 1 + len(ln_data_ws4) + 1
                else:
                    fila_nota = fila_h + 1 + len(hipotesis_data) + 1
                ws4.merge_cells(f'B{fila_nota}:C{fila_nota}')
                c = ws4[f'B{fila_nota}']
                c.value = "CONFIGURACION VARIABILIDAD MATRIZ"
                self._aplicar_estilo(c, self._estilo_subtitulo())
                mat_notas = []
                if ds.get('var_matriz_costos'):
                    mat_notas.append(("Var. Costos Matriz", f"+/-{ds.get('pct_var_costos', 0):.1f}%"))
                if ds.get('var_matriz_ingresos'):
                    mat_notas.append(("Var. Ingresos Matriz", f"+/-{ds.get('pct_var_ingresos', 0):.1f}%"))
                mat_notas.append(("Distribucion", ds.get('distribucion', 'Log-normal')))
                mat_notas.append(("Correlacion (rho)", f"{ds.get('rho', 0):.2f}"))
                for i, (lab, val) in enumerate(mat_notas):
                    f = fila_nota + 1 + i
                    c = ws4.cell(row=f, column=2)
                    c.value = lab
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                    c.alignment = Alignment(horizontal='left')
                    c = ws4.cell(row=f, column=3)
                    c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))

            # Grafico
            if img_simulacion:
                self._insertar_imagen(ws4, img_simulacion, 'E4', ancho_cm=36, alto_cm=22)
        else:
            ws4['B4'] = "Ejecute la Simulacion Monte Carlo para generar datos y graficos."
            ws4['B4'].font = Font(italic=True, color='999999', size=11)

        # ========================================================

