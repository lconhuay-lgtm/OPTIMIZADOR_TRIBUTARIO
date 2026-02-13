import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWs1Mixin:
    def _exportar_ws1(self, wb, r, tasa_efectiva, roi):
        # HOJA 1: RESUMEN EJECUTIVO
        # ========================================================
        ws1 = wb.active
        ws1.title = "RESUMEN EJECUTIVO"
        ws1.sheet_properties.tabColor = "1F4E79"

        # Título principal
        ws1.merge_cells('B2:H2')
        c = ws1['B2']
        c.value = "REPORTE DE OPTIMIZACION TRIBUTARIA - PRESENTACION GERENCIAL"
        self._aplicar_estilo(c, {
            'font': Font(name='Calibri', bold=True, color='1F4E79', size=18),
            'alignment': Alignment(horizontal='center', vertical='center')
        })
        ws1.row_dimensions[2].height = 40

        # Subtítulo
        ws1.merge_cells('B3:H3')
        c = ws1['B3']
        c.value = f"Grupo {r['matriz']['nombre']} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        self._aplicar_estilo(c, {
            'font': Font(name='Calibri', italic=True, color='666666', size=11),
            'alignment': Alignment(horizontal='center', vertical='center')
        })

        # Línea separadora
        for col in range(2, 9):
            c = ws1.cell(row=4, column=col)
            c.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws1.row_dimensions[4].height = 4

        # KPIs principales (fila 6-7)
        kpis = [
            ("AHORRO TRIBUTARIO", f"S/ {r['grupo']['ahorro_tributario']:,.0f}"),
            ("AHORRO %", f"{r['grupo']['ahorro_porcentual']:.1f}%"),
            ("TASA EFECTIVA", f"{tasa_efectiva:.1f}%"),
            ("ROI ESTRUCTURA", f"{roi:.1f}%"),
        ]

        col_start = 2
        for i, (label, valor) in enumerate(kpis):
            col = col_start + (i * 2)
            # Label
            c = ws1.cell(row=6, column=col)
            c.value = label
            self._aplicar_estilo(c, self._estilo_kpi_label())
            ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col)
            ws1.column_dimensions[get_column_letter(col)].width = 22
            # Valor
            c = ws1.cell(row=7, column=col)
            c.value = valor
            self._aplicar_estilo(c, self._estilo_kpi_valor())
        ws1.row_dimensions[6].height = 30
        ws1.row_dimensions[7].height = 35

        # Sección: Escenario Sin Estructura
        fila = 9
        ws1.merge_cells(f'B{fila}:E{fila}')
        c = ws1[f'B{fila}']
        c.value = "ESCENARIO SIN ESTRUCTURA SATELITAL"
        self._aplicar_estilo(c, self._estilo_subtitulo())
        fila += 1

        datos_sin = [
            ("Ingresos Totales", f"S/ {r['matriz']['ingresos']:,.2f}"),
            ("Costos Externos", f"S/ {r['matriz']['costos_externos']:,.2f}"),
            ("Utilidad Gravable", f"S/ {r['matriz']['utilidad_sin_estructura']:,.2f}"),
            ("Impuesto (29.5%)", f"S/ {r['matriz']['impuesto_sin_estructura']:,.2f}"),
        ]
        for j, (lab, val) in enumerate(datos_sin):
            c = ws1.cell(row=fila + j, column=2)
            c.value = lab
            self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
            c.alignment = Alignment(horizontal='left')
            c = ws1.cell(row=fila + j, column=4)
            c.value = val
            self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))

        # Sección: Con Estructura
        fila += len(datos_sin) + 1
        ws1.merge_cells(f'B{fila}:E{fila}')
        c = ws1[f'B{fila}']
        c.value = "ESCENARIO CON ESTRUCTURA OPTIMIZADA"
        self._aplicar_estilo(c, self._estilo_subtitulo())
        fila += 1

        datos_con = [
            ("Impuesto Matriz (con estructura)", f"S/ {r['matriz']['impuesto_con_estructura']:,.2f}"),
            ("Impuesto Total Satelites", f"S/ {r['grupo']['total_impuesto_satelites']:,.2f}"),
            ("Impuesto Total Grupo", f"S/ {r['grupo']['impuesto_total']:,.2f}"),
            ("AHORRO NETO", f"S/ {r['grupo']['ahorro_tributario']:,.2f}"),
        ]
        for j, (lab, val) in enumerate(datos_con):
            c = ws1.cell(row=fila + j, column=2)
            c.value = lab
            self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
            c.alignment = Alignment(horizontal='left')
            c = ws1.cell(row=fila + j, column=4)
            c.value = val
            self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
            if j == len(datos_con) - 1:
                ws1.cell(row=fila + j, column=2).font = Font(name='Calibri', bold=True, color='006600', size=11)
                ws1.cell(row=fila + j, column=4).font = Font(name='Calibri', bold=True, color='006600', size=11)

        # Seccion: Equilibrio Financiero (si se calculo)
        if self.datos_pagos_cuenta:
            gpc = self.datos_pagos_cuenta['global']
            fila += len(datos_con) + 1
            ws1.merge_cells(f'B{fila}:E{fila}')
            c = ws1[f'B{fila}']
            c.value = "EQUILIBRIO FINANCIERO (IR = Pagos a Cuenta, Saldo = 0)"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            fila += 1

            datos_eq_res = [
                ("IR Grupo (Equilibrio)", f"S/ {gpc['ir_grupo_eq']:,.2f}"),
                ("PaC Grupo (Equilibrio)", f"S/ {gpc['pc_grupo_eq']:,.2f}"),
                ("Saldo Regularizacion", f"S/ {gpc['saldo_grupo_eq']:,.2f}"),
                ("Eficiencia PaC", f"{gpc['eficiencia_eq']:.1f}% (ideal=100%)"),
                ("Tasa Efectiva Grupo", f"{gpc['tasa_efectiva_eq']:.1f}%"),
                ("AHORRO vs Sin Estructura", f"S/ {gpc['ahorro_eq_total']:,.2f} ({gpc['ahorro_eq_pct']:.1f}%)"),
            ]
            for j, (lab, val) in enumerate(datos_eq_res):
                c = ws1.cell(row=fila + j, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                c = ws1.cell(row=fila + j, column=4)
                c.value = val
                self._aplicar_estilo(c, self._estilo_celda(j % 2 == 0))
                if j == len(datos_eq_res) - 1:
                    ws1.cell(row=fila + j, column=2).font = Font(name='Calibri', bold=True, color='2E75B6', size=11)
                    ws1.cell(row=fila + j, column=4).font = Font(name='Calibri', bold=True, color='2E75B6', size=11)

        # ========================================================

