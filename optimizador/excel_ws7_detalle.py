import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWs7DetalleMixin:
    def _exportar_ws7_detalle(self, ws7, dpc, gpc, fila, img_pagos_cuenta):
        """Continua generacion de hoja ws7: secciones IV-V."""
        # --- SECCION IV: INTERPRETACION GERENCIAL ---
        ws7.merge_cells(f'B{fila}:K{fila}')
        c = ws7[f'B{fila}']
        c.value = "IV. INTERPRETACION GERENCIAL - ESTRATEGIA DE EQUILIBRIO FINANCIERO"
        self._aplicar_estilo(c, self._estilo_subtitulo())
        fila += 1

        interp_lines = [
            f"PUNTO DE EQUILIBRIO: Cada satelite opera al margen donde IR = Pagos a Cuenta. "
            f"El saldo de regularizacion es cero: no hay dinero inmovilizado en SUNAT ni falta liquidez.",

            f"SIN ESTRUCTURA: {gpc['nombre_matriz']} paga IR de S/ {gpc['impuesto_sin_estructura']:,.0f} (29.5%). "
            f"Eficiencia PaC: {gpc['eficiencia_sin']:.1f}%. "
            + ("Saldo a favor inmovilizado." if gpc['saldo_matriz_sin'] < 0 else "Debe regularizar saldo."),

            f"EN EQUILIBRIO: El grupo paga IR de S/ {gpc['ir_grupo_eq']:,.0f} (tasa efectiva {gpc['tasa_efectiva_eq']:.1f}%). "
            f"Eficiencia PaC: {gpc['eficiencia_eq']:.1f}%. Ahorro: S/ {gpc['ahorro_eq_total']:,.0f} ({gpc['ahorro_eq_pct']:.1f}%).",

            f"REFERENCIA (Margen Optimo): Maximiza ahorro fiscal a S/ {gpc['ahorro_opt_total']:,.0f} ({gpc['ahorro_opt_pct']:.1f}%), "
            f"pero con eficiencia PaC de {gpc['eficiencia_opt']:.1f}% y saldo de regularizacion de S/ {gpc['saldo_grupo_opt']:,.0f}.",

            "CONCLUSION: El enfoque de equilibrio financiero prioriza que IR = PaC en cada satelite, "
            "logrando saldo cero y flujo de caja optimo. La gerencia puede elegir entre maximizar el ahorro "
            "fiscal (margen optimo) o maximizar la eficiencia financiera (punto de equilibrio)."
        ]
        for line in interp_lines:
            ws7.merge_cells(f'B{fila}:K{fila}')
            c = ws7[f'B{fila}']
            c.value = line
            c.font = Font(name='Calibri', size=10, color='333333')
            c.alignment = Alignment(horizontal='left', wrap_text=True)
            ws7.row_dimensions[fila].height = 40
            fila += 1
        ws7[f'B{fila - 1}'].font = Font(name='Calibri', size=10, color='1F4E79', bold=True)

        fila += 1

        ws7.column_dimensions['B'].width = 30
        ws7.column_dimensions['C'].width = 22
        ws7.column_dimensions['D'].width = 22
        ws7.column_dimensions['E'].width = 22
        ws7.column_dimensions['F'].width = 18

        # --- SECCION V: DETALLE POR SATELITE (equilibrio primero) ---
        ws7.merge_cells(f'B{fila}:K{fila}')
        c = ws7[f'B{fila}']
        c.value = "V. DETALLE POR SATELITE - PUNTO DE EQUILIBRIO FINANCIERO"
        self._aplicar_estilo(c, self._estilo_subtitulo())
        fila += 1

        fila_base = fila
        for idx_sat, sat_pc in enumerate(dpc['satelites']):
            ws7.merge_cells(f'B{fila_base}:K{fila_base}')
            c = ws7[f'B{fila_base}']
            c.value = f"{idx_sat + 1}. {sat_pc['nombre']} | Costo: {sat_pc['costo']:,.0f} | Gastos: {sat_pc['gastos']:,.0f}"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            fila_base += 1

            tiene_eq = sat_pc['margen_equilibrio'] is not None
            eq_str = f"{sat_pc['margen_equilibrio']:.4f}%" if tiene_eq else "N/A"
            reg_eq_str = f" [{sat_pc['regimen_equilibrio']}]" if tiene_eq else ""

            info_sat = [
                (">>> PUNTO DE EQUILIBRIO (IR=PaC) <<<", ""),
                ("Margen Equilibrio", f"{eq_str}{reg_eq_str}"),
                ("Precio Venta en Equilibrio", sat_pc['precio_equilibrio']),
                ("Utilidad Neta en Equilibrio", sat_pc['utilidad_neta_eq']),
                ("IR = PaC en Equilibrio", sat_pc['ir_equilibrio']),
                ("Saldo en Equilibrio", sat_pc['saldo_equilibrio']),
                ("Ahorro vs Reg. General", sat_pc['ahorro_equilibrio']),
                ("--- Referencia: Margen Optimo ---", ""),
                ("Margen Optimo", f"{sat_pc['margen_optimo_regimen']:.4f}%"),
                ("IR en Optimo", sat_pc['ir_optimo']),
                ("PaC en Optimo", sat_pc['pago_cuenta_optimo']),
                ("Saldo en Optimo", sat_pc['saldo_optimo']),
                ("Eficiencia PaC en Optimo", f"{sat_pc['eficiencia_optimo']:.1f}%"),
            ]
            for i, (lab, val) in enumerate(info_sat):
                c = ws7.cell(row=fila_base + i, column=2)
                c.value = lab
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                c.alignment = Alignment(horizontal='left')
                if lab.startswith(">>>"):
                    c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                elif lab.startswith("---"):
                    c.font = Font(name='Calibri', bold=True, color='888888', size=9)
                c = ws7.cell(row=fila_base + i, column=3)
                if isinstance(val, (int, float)):
                    c.value = val
                    c.number_format = '#,##0.00'
                else:
                    c.value = val
                self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                if lab == "Saldo en Equilibrio" and isinstance(val, (int, float)):
                    c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                    c2 = ws7.cell(row=fila_base + i, column=4)
                    c2.value = "(~0 EQUILIBRIO)" if abs(val) < 1 else ""
                    c2.font = Font(name='Calibri', bold=True, color='2E75B6', size=9)

            fila_base += len(info_sat) + 1

            # Tabla de sensibilidad
            ws7.merge_cells(f'B{fila_base}:K{fila_base}')
            c = ws7[f'B{fila_base}']
            c.value = "Sensibilidad del Margen - EQUILIBRIO es donde Saldo = 0"
            self._aplicar_estilo(c, {
                'font': Font(name='Calibri', bold=True, color='2E75B6', size=10),
                'alignment': Alignment(horizontal='left')
            })
            fila_base += 1

            headers_sens = ["Margen%", "Precio Venta", "Util. Neta", "Regimen",
                           "IR", "Pago Cta", "Saldo", "Ahorro Reg.", "Efic%", "Nota"]
            for j, h in enumerate(headers_sens, 2):
                c = ws7.cell(row=fila_base, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
            ws7.row_dimensions[fila_base].height = 24
            fila_base += 1

            for i, s in enumerate(sat_pc['sensibilidad']):
                nota = "<<< EQUILIBRIO" if s['es_equilibrio'] else ("(ref: optimo)" if s['es_optimo'] else "")
                datos_fila = [
                    s['margen'], s['precio_venta'], s['utilidad_neta'],
                    s['regimen'], s['ir'], s['pago_cuenta'],
                    s['saldo'], s['ahorro_regimen'], s['eficiencia_pc'], nota
                ]
                for j, val in enumerate(datos_fila, 2):
                    c = ws7.cell(row=fila_base, column=j)
                    if isinstance(val, (int, float)):
                        c.value = val
                        if j == 2:
                            c.number_format = '0.0000"%"'
                        elif j == 10:
                            c.number_format = '0.0"%"'
                        else:
                            c.number_format = '#,##0.00'
                    else:
                        c.value = val
                    estilo = self._estilo_celda(i % 2 == 0)
                    self._aplicar_estilo(c, estilo)
                    if s['es_equilibrio']:
                        c.font = Font(name='Calibri', bold=True, color='2E75B6', size=10)
                        c.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                    elif s['es_optimo']:
                        c.font = Font(name='Calibri', color='888888', size=9)
                fila_base += 1

            # Sensibilidad 2D: equilibrio ante cambios en tasa PaC
            fila_base += 1
            ws7.merge_cells(f'B{fila_base}:J{fila_base}')
            c = ws7[f'B{fila_base}']
            c.value = "Punto de Equilibrio ante cambios en la Tasa de Pago a Cuenta"
            self._aplicar_estilo(c, {
                'font': Font(name='Calibri', bold=True, color='2E75B6', size=10),
                'alignment': Alignment(horizontal='left')
            })
            fila_base += 1

            headers_2d = ["Tasa PaC%", "M. Equil%", "Precio Eq.", "Util. Neta",
                         "IR = PaC", "Saldo", "Ahorro Reg.", "Regimen"]
            for j, h in enumerate(headers_2d, 2):
                c = ws7.cell(row=fila_base, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
                if h in ("M. Equil%", "IR = PaC", "Saldo"):
                    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            fila_base += 1

            for i, s2 in enumerate(sat_pc['sensibilidad_2d']):
                datos_2d = [s2['tasa_pc'], s2['margen_equilibrio'], s2['precio_venta'],
                           s2['utilidad_neta'], s2['ir'], s2['saldo'],
                           s2['ahorro_regimen'], s2['regimen']]
                for j, val in enumerate(datos_2d, 2):
                    c = ws7.cell(row=fila_base, column=j)
                    if isinstance(val, (int, float)):
                        c.value = val
                        if j in (2, 3):
                            c.number_format = '0.0000"%"' if j == 3 else '0.0"%"'
                        else:
                            c.number_format = '#,##0.00'
                    else:
                        c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(i % 2 == 0))
                fila_base += 1

            fila_base += 2

        # Ajustar anchos para las columnas de sensibilidad
        for col_idx in range(2, 12):
            ws7.column_dimensions[get_column_letter(col_idx)].width = max(
                ws7.column_dimensions[get_column_letter(col_idx)].width or 0, 14)

        # Grafico pagos a cuenta
        if img_pagos_cuenta:
            self._insertar_imagen(ws7, img_pagos_cuenta, f'B{fila_base}', ancho_cm=36, alto_cm=20)
