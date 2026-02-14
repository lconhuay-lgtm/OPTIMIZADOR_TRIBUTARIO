import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWs2Ws3Mixin:
    def _exportar_ws2(self, wb, r):
        # HOJA 2: DETALLE SATELITES
        # ========================================================
        ws2 = wb.create_sheet("DETALLE SATELITES")
        ws2.sheet_properties.tabColor = "2E75B6"

        ws2.merge_cells('B2:L2')
        c = ws2['B2']
        c.value = "DETALLE POR EMPRESA SATELITE"
        self._aplicar_estilo(c, self._estilo_titulo())
        ws2.row_dimensions[2].height = 30

        # Headers
        headers = ["N", "Empresa", "Tipo", "Regimen", "Costos", "Gastos Op.",
                   "Margen %", "Precio Venta", "Utilidad Neta", "Impuesto",
                   "Ahorro Ind.", "Tasa Efect."]
        for j, h in enumerate(headers, 2):
            c = ws2.cell(row=4, column=j)
            c.value = h
            self._aplicar_estilo(c, self._estilo_header())
        ws2.row_dimensions[4].height = 28

        # Datos
        for i, sat in enumerate(r['satelites']):
            fila = 5 + i
            alt = i % 2 == 0
            datos_fila = [
                i + 1,
                sat['nombre'],
                sat['tipo'],
                sat['regimen'],
                sat['costo'],
                sat['gastos_operativos'],
                sat['margen_optimo'],
                sat['precio_venta'],
                sat['utilidad_neta'],
                sat['impuesto'],
                sat['ahorro_individual'],
                sat['tasa_efectiva']
            ]
            for j, val in enumerate(datos_fila, 2):
                c = ws2.cell(row=fila, column=j)
                if isinstance(val, (int, float)) and j >= 6:
                    if j == 8:  # Margen %
                        c.value = val
                        c.number_format = '0.00"%"'
                    elif j == 13:  # Tasa efectiva
                        c.value = val
                        c.number_format = '0.00"%"'
                    else:
                        c.value = val
                        c.number_format = '#,##0.00'
                else:
                    c.value = val
                self._aplicar_estilo(c, self._estilo_celda(alt))

        # Totales
        fila_total = 5 + len(r['satelites'])
        ws2.cell(row=fila_total, column=2).value = "TOTAL"
        ws2.cell(row=fila_total, column=2).font = Font(bold=True, size=11, color='1F4E79')
        for col_idx, val in [
            (6, sum(s['costo'] for s in r['satelites'])),
            (7, sum(s['gastos_operativos'] for s in r['satelites'])),
            (10, sum(s['utilidad_neta'] for s in r['satelites'])),
            (11, sum(s['impuesto'] for s in r['satelites'])),
            (12, sum(s['ahorro_individual'] for s in r['satelites']))
        ]:
            c = ws2.cell(row=fila_total, column=col_idx)
            c.value = val
            c.number_format = '#,##0.00'
            c.font = Font(bold=True, size=10, color='1F4E79')
            c.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')

        # --- METRICAS DE EQUILIBRIO FINANCIERO POR SATELITE ---
        if self.datos_pagos_cuenta and self.datos_pagos_cuenta.get('satelites'):
            fila_eq = fila_total + 2
            ws2.merge_cells(f'B{fila_eq}:Q{fila_eq}')
            c = ws2[f'B{fila_eq}']
            c.value = "METRICAS DE EQUILIBRIO FINANCIERO POR SATELITE (IR = PaC, Saldo = 0)"
            self._aplicar_estilo(c, self._estilo_subtitulo())
            c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
            c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            fila_eq += 1

            headers_eq = ["N", "Empresa", "Regimen Eq.", "Margen Eq.%",
                          "Precio Eq.", "Utilidad Eq.", "IR Eq.", "PaC Eq.",
                          "Saldo Eq.", "Ahorro Eq.", "Margen Opt.%", "IR Opt.",
                          "Saldo Opt.", "Efic.Opt.%"]
            for j, h in enumerate(headers_eq, 2):
                c = ws2.cell(row=fila_eq, column=j)
                c.value = h
                self._aplicar_estilo(c, self._estilo_header())
                if "Eq." in h:
                    c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
            ws2.row_dimensions[fila_eq].height = 28
            fila_eq += 1

            for i, sat_pc in enumerate(self.datos_pagos_cuenta['satelites']):
                alt = i % 2 == 0
                tiene_eq = sat_pc['margen_equilibrio'] is not None
                datos_eq_fila = [
                    i + 1,
                    sat_pc['nombre'],
                    sat_pc.get('regimen_equilibrio', 'N/A') if tiene_eq else 'N/A',
                    sat_pc['margen_equilibrio'] if tiene_eq else 'N/A',
                    sat_pc['precio_equilibrio'],
                    sat_pc['utilidad_neta_eq'],
                    sat_pc['ir_equilibrio'],
                    sat_pc['pago_cuenta_eq'],
                    sat_pc['saldo_equilibrio'],
                    sat_pc['ahorro_equilibrio'],
                    sat_pc['margen_optimo_regimen'],
                    sat_pc['ir_optimo'],
                    sat_pc['saldo_optimo'],
                    sat_pc['eficiencia_optimo'],
                ]
                for j, val in enumerate(datos_eq_fila, 2):
                    c = ws2.cell(row=fila_eq, column=j)
                    if isinstance(val, (int, float)):
                        c.value = val
                        if j in (5, 15):  # Margen %
                            c.number_format = '0.0000"%"'
                        elif j == 15:  # Eficiencia %
                            c.number_format = '0.00"%"'
                        else:
                            c.number_format = '#,##0.00'
                    else:
                        c.value = val
                    self._aplicar_estilo(c, self._estilo_celda(alt))
                fila_eq += 1

        # Ajustar anchos
        anchos = [4, 16, 12, 26, 14, 14, 10, 16, 14, 14, 14, 12]
        for j, w in enumerate(anchos, 2):
            ws2.column_dimensions[get_column_letter(j)].width = w
        for j in range(14, 17):
            ws2.column_dimensions[get_column_letter(j)].width = 14

        # ========================================================

    def _exportar_ws3(self, wb, imgs):
        # HOJA 3: GRAFICOS RESULTADOS
        # ========================================================
        ws3 = wb.create_sheet("GRAFICOS RESULTADOS")
        ws3.sheet_properties.tabColor = "2ECC71"

        ws3.merge_cells('B2:H2')
        c = ws3['B2']
        c.value = "VISUALIZACION ESTRATEGICA - ANALISIS TRIBUTARIO"
        self._aplicar_estilo(c, self._estilo_titulo())

        if img_resultados:
            self._insertar_imagen(ws3, img_resultados, 'B4', ancho_cm=36, alto_cm=22)
        else:
            ws3['B4'] = "Ejecute 'Calcular Margenes Optimos' para generar los graficos."
            ws3['B4'].font = Font(italic=True, color='999999', size=11)

        # ========================================================

